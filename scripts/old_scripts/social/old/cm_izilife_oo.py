"""
Agent Community Manager — izilife
Usage :
  python cm_izilife.py --zone lille-zone
  python cm_izilife.py --init  lille-zone
  python cm_izilife.py --list
  python cm_izilife.py --update MON_TYPE carrousel
"""

import os, sys, re, argparse, requests, anthropic, openpyxl, base64
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

# ─────────────────────────────────────────────
# CORE PATHS / CONFIG — migration safe
# ─────────────────────────────────────────────

def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return

_ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, ENV_GLOBAL, ENV_IZILIFE, izilife_social_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False


# ══════════════════════════════════════════════════════
# CHEMINS FIXES (locaux)
# ══════════════════════════════════════════════════════

if HAS_CORE_PATHS:
    DRIVE_ROOT    = izilife_social_zone("").parent
    CONTEXT_ROOT  = PROJECT_ROOT / "izilife" / "context" / "social"
    TEMPLATES_DIR = PROJECT_ROOT / "izilife" / "templates"
else:
    DRIVE_ROOT    = Path("G:/Mon Drive/agentic_workspace/izilife/social")
    CONTEXT_ROOT  = Path("C:/Users/alcamara/Documents/agentic_Workspace/izilife/context/social")
    TEMPLATES_DIR = Path("C:/Users/alcamara/Documents/agentic_Workspace/izilife/templates")

CONTEXT_COMMON = [
    "izilife-social-strategy.md",
    "community-manager.md",
    "izilife-style.md",
]

# Mapping TYPE → CONTENT_TYPE et TEMPLATE
TYPE_CONTENT = {
    "POST_AGENDA_SEMAINE":  "story",
    "POST_AGENDA_WEEKEND":  "story",
    "POST_TOP_LIEUX":       "carrousel",
    "POST_TOP_ACTEURS":     "carrousel",
    "POST_HISTOIRE":        "post",
    "POST_HUMOUR":          "post",
    "POST_ESCAPADE_VILLE":  "carrousel",
    "POST_ESCAPADE_NATURE": "carrousel",
    "POST_PEPITE":          "post",
    "EVENT_SERIE":          "post",
    "EVENT_ANIMATEUR":      "post",
    "PARTAGE_ACTEUR":       "post",
    "NOUVEAUTE_IZILIFE":    "post",
    "REEL":                 "reel",
    "VIDEO":                "video",
}

TYPE_TEMPLATE = {
    "POST_AGENDA_SEMAINE":  "agenda_semaine",
    "POST_AGENDA_WEEKEND":  "agenda_weekend",
    "POST_TOP_LIEUX":       "top",
    "POST_TOP_ACTEURS":     "top",
    "POST_HISTOIRE":        "histoire_lieu",
    "POST_HUMOUR":          "humour_local",
    "POST_ESCAPADE_VILLE":  "escapade_ville",
    "POST_ESCAPADE_NATURE": "escapade_ville",
    "POST_PEPITE":          "pepite_niche",
    "EVENT_SERIE":          "event_serie",
    "EVENT_ANIMATEUR":      "event_animateur",
}

# Colonnes Excel
COL_DATE=1; COL_RESEAU=2; COL_TYPE=3; COL_VILLE=4; COL_LIEU=5
COL_THEME=6; COL_SERIE=7; COL_ARTISTE=8; COL_STYLE=9; COL_DATA=10
COL_DATE_EVENT=11; COL_HEURE=12; COL_INPUT_ID=13
COL_IMAGE_PROMPT=14; COL_IMAGE_PROVIDER=15
COL_POST_OUTPUT=16; COL_STATUT=17

STATUT_A_FAIRE="À faire"; STATUT_RELANCER="Relancer"; STATUT_GENERE="Généré"

# ══════════════════════════════════════════════════════
# ZONES
# ══════════════════════════════════════════════════════

def get_zone_config(zone:str) -> dict:
    zone_drive = DRIVE_ROOT / zone
    return {
        "zone":     zone,
        "drive":    zone_drive,
        "excel":    zone_drive / f"planning_{zone}.xlsx",
        "outputs":  zone_drive / "outputs",
        "inputs":   zone_drive / "inputs",

        "context_zone": CONTEXT_ROOT / "zones" / f"{zone}.md",
    }

if not HAS_CORE_PATHS:
    ENV_GLOBAL   = Path("C:/Users/alcamara/Documents/agentic_Workspace/.env")
    ENV_IZILIFE  = Path("C:/Users/alcamara/Documents/agentic_Workspace/izilife/.env.izilife")

def load_zone_env(cfg:dict):
    """Charge les .env en cascade — JAMAIS de clés sur le Drive."""
    # 1. Clés API globales (local)
    if ENV_GLOBAL.exists():
        load_dotenv(ENV_GLOBAL, override=False)
    # 2. Config moteurs izilife (local)
    if ENV_IZILIFE.exists():
        load_dotenv(ENV_IZILIFE, override=False)
    # 3. Override zone si fichier local existe (PAS sur Drive)
    zone_override = cfg["drive"].parent / f".env.{cfg['zone']}"
    if zone_override.exists():
        load_dotenv(zone_override, override=True)
        print(f"   .env override zone : {zone_override}")

def list_zones():
    print("\n📍 Zones :\n")
    if not DRIVE_ROOT.exists():
        print(f"  ❌ Drive introuvable : {DRIVE_ROOT}"); return
    for d in sorted(DRIVE_ROOT.iterdir()):
        if d.is_dir():
            cfg = get_zone_config(d.name)
            env_ok  = "✅" if ENV_IZILIFE.exists()   else "⚠️ "
            xlsx_ok = "✅" if cfg["excel"].exists()  else "❌"
            print(f"  • {d.name}  {xlsx_ok} planning  {env_ok} .env")

# ══════════════════════════════════════════════════════
# INIT ZONE
# ══════════════════════════════════════════════════════

def init_zone(zone:str):
    cfg = get_zone_config(zone)
    for d in ["outputs","inputs"]:
        (cfg["drive"] / d).mkdir(parents=True, exist_ok=True)
    print(f"✅ Dossiers Drive créés : {cfg['drive']}")

    print(f"ℹ️  Config moteurs : {ENV_IZILIFE}")
    print(f"ℹ️  Clés API       : {ENV_GLOBAL}")
    if not ENV_GLOBAL.exists():
        print(f"⚠️  Crée ce fichier avec tes clés API : {ENV_GLOBAL}")
    if not ENV_IZILIFE.exists():
        print(f"⚠️  Crée ce fichier de config : {ENV_IZILIFE}")

    # Contexte zone
    zone_md = CONTEXT_ROOT / "zones" / f"{zone}.md"
    zone_md.parent.mkdir(parents=True, exist_ok=True)
    if not zone_md.exists():
        zone_md.write_text(f"""# Contexte zone — {zone}
## Spécificités
- Ville principale :
- Hashtags locaux : #{zone.replace('-zone','')}
- Compte Instagram : @izilife_{zone.replace('-zone','')}
- Événements locaux :
- Lieux emblématiques :
""", encoding="utf-8")
        print(f"✅ Contexte zone : {zone_md}")

    _create_planning_excel(cfg["excel"], zone)
    print(f"✅ Planning : {cfg['excel']}")
    print(f"\n🎉 Zone '{zone}' prête !")
    print(f"   1. Remplis : {ENV_IZILIFE}")
    print(f"   2. Lance   : python cm_izilife.py --zone {zone}")

def update_post_type(post_type:str, content_type:str):
    """Ajoute un type dans community-manager.md."""
    cm_path = CONTEXT_ROOT / "community-manager.md"
    if not cm_path.exists():
        print(f"❌ {cm_path} introuvable"); return
    content = cm_path.read_text(encoding="utf-8")
    if f"### {post_type}" in content:
        print(f"⚠️  {post_type} existe déjà"); return
    entry = f"""
### {post_type}
CONTENT_TYPE : {content_type}
TEMPLATE_LOCAL : templates/{post_type.lower()}/
CANVA_ID : —
VISUEL : à définir
TEXTE : à définir
"""
    cm_path.write_text(content + entry, encoding="utf-8")
    print(f"✅ {post_type} ajouté dans community-manager.md")
    print(f"   → Relance --init [zone] pour mettre à jour la liste Excel")

def _create_planning_excel(path:Path, zone:str):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning"
    ws.freeze_panes = "D3"

    thin = Side(style='thin', color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(sheet, r, c, v, bg="1A1A2E", ft="FFFFFF", sz=9):
        x = sheet.cell(row=r, column=c, value=v)
        x.font = Font(name="Arial", size=sz, bold=True, color=ft)
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = brd

    # Groupes
    groups = [
        (1,3,  "PLANIFICATION", "1A1A2E"),
        (4,7,  "CONTENU",       "0F3460"),
        (8,10, "DONNÉES",       "533483"),
        (11,12,"EVENT",         "1B4332"),
        (13,13,"INPUT",         "784212"),
        (14,15,"IMAGE",         "0D6986"),
        (16,16,"OUTPUT",        "1A5276"),
        (17,17,"STATUT",        "6C3483"),
    ]
    for s, e, label, bg in groups:
        if s < e: ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        hdr(ws, 1, s, label, bg=bg, sz=10)

    headers = [
        ("DATE",12),("RÉSEAU",11),("TYPE",22),
        ("VILLE",12),("LIEU",16),("THEME / SUJET",22),("SERIE",13),
        ("ARTISTE / DJ / ASSO",26),("STYLE",14),("DATA",36),
        ("DATE EVENT",12),("HEURE / DUREE",11),
        ("INPUT_ID",16),
        ("IMAGE_PROMPT",36),("IMAGE_PROVIDER\ngpt / canva",14),
        ("POST OUTPUT",60),
        ("STATUT",13),
    ]
    for i, (label, width) in enumerate(headers, 1):
        hdr(ws, 2, i, label)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 36

    all_types = '","'.join(TYPE_CONTENT.keys())
    dv_res  = DataValidation(type="list", formula1='"Instagram,Facebook,Les deux"', allow_blank=True)
    dv_type = DataValidation(type="list", formula1=f'"{all_types}"', allow_blank=True)
    dv_img  = DataValidation(type="list", formula1='"gpt,canva"', allow_blank=True)
    dv_stat = DataValidation(type="list",
        formula1='"À faire,Généré,Relancer,Validé,Publié,Skip"', allow_blank=True)

    for dv in [dv_res, dv_type, dv_img, dv_stat]:
        ws.add_data_validation(dv)
    for ri in range(3, 103):
        ws.row_dimensions[ri].height = 34
        for dv, col in [(dv_res,2),(dv_type,3),(dv_img,15),(dv_stat,17)]:
            dv.add(ws.cell(row=ri, column=col))

    # Onglet Post Types
    wt = wb.create_sheet("Post Types")
    wt.merge_cells("A1:F1")
    c = wt.cell(row=1, column=1,
        value="Configuration types de posts — comportement géré via community-manager.md")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(["POST_TYPE","CONTENT_TYPE","TEMPLATE_LOCAL","CANVA_ID","NOTES"], 1):
        hdr(wt, 2, i, h)
        wt.column_dimensions[get_column_letter(i)].width = [28,14,30,16,40][i-1]

    canva_ids = {
        "POST_AGENDA_SEMAINE":"DAGzaUvl3jI","POST_AGENDA_WEEKEND":"DAGzaUvl3jI",
        "POST_TOP_LIEUX":"DAG0oGZeu3M","POST_TOP_ACTEURS":"DAG0bj8ohvI",
        "POST_ESCAPADE_VILLE":"DAG0bhdKF14","POST_ESCAPADE_NATURE":"DAG0bq6g0yI",
    }
    notes_map = {
        "POST_HUMOUR":"Sondage / citation / liste",
        "POST_TOP_ACTEURS":"Photos via INPUT_ID",
        "POST_TOP_LIEUX":"Photos via INPUT_ID",
        "EVENT_SERIE":"Template par event dans templates/event_serie/",
        "REEL":"[futur] moteur video","VIDEO":"[futur] moteur video",
    }
    for ri, (pt, ct) in enumerate(TYPE_CONTENT.items(), 3):
        bg = "F5F7FA" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate([
            pt, ct,
            f"templates/{TYPE_TEMPLATE.get(pt,'—')}/",
            canva_ids.get(pt,"—"),
            notes_map.get(pt,""),
        ], 1):
            x = wt.cell(row=ri, column=ci, value=val)
            x.font = Font(name="Arial", size=9)
            x.fill = PatternFill("solid", fgColor=bg)
            x.alignment = Alignment(vertical="center")
            x.border = brd
        wt.row_dimensions[ri].height = 20

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

# ══════════════════════════════════════════════════════
# CONTEXTE
# ══════════════════════════════════════════════════════

def load_context(zone:str) -> str:
    parts = []
    for filename in CONTEXT_COMMON:
        path = CONTEXT_ROOT / filename
        if path.exists():
            parts.append(f"### {filename}\n{path.read_text(encoding='utf-8')}")
        else:
            print(f"  ⚠️  Manquant : {path}")
    zone_md = CONTEXT_ROOT / "zones" / f"{zone}.md"
    if zone_md.exists():
        parts.append(f"### {zone}.md\n{zone_md.read_text(encoding='utf-8')}")
    return "\n\n---\n\n".join(parts)

# ══════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════

def slugify(t:str) -> str:
    t = t.encode('ascii','ignore').decode()
    return re.sub(r"[^a-z0-9_]","",t.lower().replace(" ","_"))[:40]

def read_input(input_id:str, inputs_dir:Path) -> tuple[str, list[Path]]:
    if not input_id: return "", []
    folder = inputs_dir / input_id
    if not folder.exists(): return "", []
    desc = ""
    f = folder / "description.txt"
    if f.exists(): desc = f.read_text(encoding="utf-8").strip()
    images = sorted([x for x in folder.iterdir()
                     if x.suffix.lower() in (".jpg",".jpeg",".png")])
    return desc, images

# ══════════════════════════════════════════════════════
# TEXTE
# ══════════════════════════════════════════════════════

def build_prompt(row, content_type:str, input_desc:str) -> str:
    def v(col): val=row[col-1].value; return str(val).strip() if val else ""
    desc = f"\nDESCRIPTION :\n{input_desc}" if input_desc else ""
    slides_instr = "\n=== SLIDES ===\nSlide 1 : [texte]\nSlide 2 : [texte]\n..." \
                   if content_type == "carrousel" else ""
    return f"""Tu es l'agent Community Manager izilife. CONTENT_TYPE: {content_type}

DONNÉES :
TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
THEME: {v(COL_THEME)} | SERIE: {v(COL_SERIE)}
ARTISTE/DJ/ASSO: {v(COL_ARTISTE)} | STYLE: {v(COL_STYLE)}
DATA: {v(COL_DATA)}
DATE EVENT: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
RESEAU: {v(COL_RESEAU)}{desc}

Format OBLIGATOIRE :
=== TEXTE DU POST ===
[caption complète avec emojis + hashtags]{slides_instr}
"""

def call_llm(system:str, user:str) -> str:
    provider = os.getenv("LLM_PROVIDER","claude")
    if provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL","gpt-4o"),
            messages=[{"role":"system","content":system},{"role":"user","content":user}],
            max_tokens=1500)
        return r.choices[0].message.content
    else:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        msg = client.messages.create(
            model=os.getenv("CLAUDE_MODEL","claude-sonnet-4-6"),
            max_tokens=1500, system=system,
            messages=[{"role":"user","content":user}])
        return msg.content[0].text

# ══════════════════════════════════════════════════════
# IMAGE
# ══════════════════════════════════════════════════════

def get_template(post_type:str) -> Path | None:
    tpl = TYPE_TEMPLATE.get(post_type,"")
    if not tpl: return None
    folder = TEMPLATES_DIR / tpl
    if not folder.exists(): return None
    for ext in ["*.png","*.jpg"]:
        files = sorted(folder.glob(ext))
        if files: return files[0]
    return None

def generate_image_gpt(prompt:str, folder:Path, filename:str) -> Path | None:
    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        result = client.images.generate(
            model=os.getenv("IMAGE_MODEL","gpt-image-1"),
            prompt=f"{prompt}. Style photographique Instagram, format portrait.",
            n=1, size="1024x1536",
        )
        img_b64 = result.data[0].b64_json
        if img_b64:
            out = folder / f"{filename}_gpt.png"
            out.write_bytes(base64.b64decode(img_b64))
            print(f"    🖼  GPT → {out.name}"); return out
        url = result.data[0].url
        if url:
            out = folder / f"{filename}_gpt.jpg"
            out.write_bytes(requests.get(url).content)
            print(f"    🖼  GPT → {out.name}"); return out
    except Exception as e:
        print(f"    ❌ GPT Image : {e}")
    return None

def compose_pillow(bg:Path, text:str, folder:Path, filename:str) -> Path | None:
    try:
        from PIL import Image, ImageDraw, ImageFont
        img = Image.open(bg).convert("RGBA")
        W, H = img.size
        overlay = Image.new("RGBA",(W,H),(0,0,0,130))
        img = Image.alpha_composite(img, overlay)
        draw = ImageDraw.Draw(img)
        try: font = ImageFont.truetype("arial.ttf", max(36, W//20))
        except: font = ImageFont.load_default()
        lines = [l for l in text.split("\n") if l.strip()][:8]
        y = H // 4
        for line in lines:
            bb = draw.textbbox((0,0), line, font=font)
            x = (W - (bb[2]-bb[0])) // 2
            draw.text((x+2,y+2), line, font=font, fill=(0,0,0,180))
            draw.text((x,y),     line, font=font, fill=(255,255,255,255))
            y += bb[3]-bb[1]+14
        out = folder / f"{filename}_pillow.jpg"
        img.convert("RGB").save(out, quality=95)
        print(f"    🖼  Pillow → {out.name}"); return out
    except Exception as e:
        print(f"    ❌ Pillow : {e}")
    return None

def process_images(post_type:str, img_prompt:str, img_provider:str,
                   input_images:list, text:str,
                   folder:Path, filename:str) -> list[Path]:
    generated = []
    engine   = os.getenv("IMAGE_ENGINE","gpt")
    provider = img_provider or os.getenv("IMAGE_LLM","gpt")

    tpl = get_template(post_type)
    if tpl: print(f"    📁 Template : {tpl.name}")

    bg = input_images[0] if input_images else tpl

    # Pillow — si fond disponible
    if bg and engine in ("pillow","both"):
        # Extrait le texte court pour l'incrustation
        short_text = ""
        for line in text.split("\n"):
            if line.startswith("===") or not line.strip(): continue
            short_text += line + "\n"
            if len(short_text) > 300: break
        out = compose_pillow(bg, short_text.strip(), folder, filename)
        if out: generated.append(out)

    # Génération IA — si prompt fourni
    if img_prompt and engine in ("gpt","both"):
        if provider == "gpt":
            out = generate_image_gpt(img_prompt, folder, filename)
            if out: generated.append(out)
        elif provider == "canva":
            print("    ⚠️  Canva Magic API — non configuré")

    # Rien fourni et pas de prompt → log
    if not generated and not bg and not img_prompt:
        print("    ℹ️  Pas d'image — ajoute photos dans inputs/ ou un IMAGE_PROMPT")

    return generated

# ══════════════════════════════════════════════════════
# OUTPUT
# ══════════════════════════════════════════════════════

def save_post(text:str, input_images:list, generated:list,
              content_type:str, folder:Path):
    lines = [text.strip(), "", f"=== CONTENT_TYPE ===\n{content_type}"]
    if input_images:
        lines += ["", "=== IMAGES SOURCE ==="] + [f"  {i.name}" for i in input_images]
    if generated:
        lines += ["", "=== IMAGES GENEREES ==="] + [f"  {i.name}" for i in generated]
    (folder/"post.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"    📄 post.txt")

# ══════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════

def run_zone(zone:str):
    cfg = get_zone_config(zone)
    load_zone_env(cfg)

    print(f"\n🤖 Agent CM izilife — zone: {zone} — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"   LLM    : {os.getenv('LLM_PROVIDER','?').upper()}")
    print(f"   Image  : {os.getenv('IMAGE_ENGINE','?').upper()} / {os.getenv('IMAGE_LLM','?').upper()}")
    print(f"   Excel  : {cfg['excel']}\n")

    if not cfg["excel"].exists():
        print(f"❌ Planning introuvable. Lance : python cm_izilife.py --init {zone}"); return

    system = load_context(zone)
    if not system: print("❌ Aucun contexte."); return

    wb = openpyxl.load_workbook(cfg["excel"])
    ws = wb.active

    to_process = [r for r in ws.iter_rows(min_row=3)
                  if r[COL_STATUT-1].value in (STATUT_A_FAIRE, STATUT_RELANCER)]
    if not to_process: print("✅ Aucune ligne à traiter."); return
    print(f"📝 {len(to_process)} ligne(s)...\n")

    cfg["outputs"].mkdir(parents=True, exist_ok=True)
    cfg["inputs"].mkdir(parents=True, exist_ok=True)

    for i, row in enumerate(to_process, 1):
        def v(col): val=row[col-1].value; return str(val).strip() if val else ""

        post_type    = v(COL_TYPE)
        theme        = v(COL_THEME)
        ville        = v(COL_VILLE)
        input_id     = v(COL_INPUT_ID)
        img_prompt   = v(COL_IMAGE_PROMPT)
        img_provider = v(COL_IMAGE_PROVIDER)
        content_type = TYPE_CONTENT.get(post_type, "post")

        print(f"  [{i}/{len(to_process)}] {post_type} [{content_type}] — {ville} — {theme}")

        folder = cfg["outputs"] / f"{datetime.now().strftime('%Y%m%d')}_{slugify(post_type)}_{slugify(theme or ville)}"
        folder.mkdir(parents=True, exist_ok=True)

        input_desc, input_images = read_input(input_id, cfg["inputs"])
        if input_images: print(f"    📸 {len(input_images)} image(s) source")
        if input_desc:   print(f"    📝 description.txt")

        # Texte
        try:
            text = call_llm(system, build_prompt(row, content_type, input_desc))
            short = text[:400]+"..." if len(text)>400 else text
            row[COL_POST_OUTPUT-1].value = short
            row[COL_STATUT-1].value = STATUT_GENERE
            print(f"    ✅ Texte")
        except Exception as e:
            print(f"    ❌ Texte : {e}")
            row[COL_STATUT-1].value = "Erreur"; continue

        # Images
        generated = process_images(
            post_type, img_prompt, img_provider,
            input_images, text, folder,
            slugify(f"{post_type}_{theme or ville}")
        )

        save_post(text, input_images, generated, content_type, folder)
        print(f"    📁 {folder.name}")

    wb.save(cfg["excel"])
    print(f"\n💾 Sauvegardé.\n✅ Terminé.\n")

def main():
    parser = argparse.ArgumentParser(description="Agent CM izilife")
    parser.add_argument("--zone",   type=str)
    parser.add_argument("--init",   type=str)
    parser.add_argument("--update", type=str, nargs=2, metavar=("TYPE","CONTENT_TYPE"))
    parser.add_argument("--list",   action="store_true")
    args = parser.parse_args()

    if args.list:         list_zones()
    elif args.init:       init_zone(args.init)
    elif args.update:     update_post_type(args.update[0], args.update[1])
    elif args.zone:       run_zone(args.zone)
    else:
        print("Usage:")
        print("  python cm_izilife.py --zone lille-zone")
        print("  python cm_izilife.py --init valenciennes-zone")
        print("  python cm_izilife.py --update MON_TYPE carrousel")
        print("  python cm_izilife.py --list")

if __name__ == "__main__":
    main()