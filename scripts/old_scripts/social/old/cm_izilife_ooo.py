"""
Agent Community Manager — izilife v3
Usage :
  python cm_izilife.py --zone lille --init
  python cm_izilife.py --zone lille
  python cm_izilife.py --zone lille --dry-run
  python cm_izilife.py --list

V3 :
- ajoute TEMPLATE_SOURCE : owned | inspiration | none
- ajoute TEMPLATE_REF : chemin / Canva ID / URL / note
- ajoute IMAGE_STATUS
- --init crée ou met à jour le planning sans écraser les lignes existantes
- génération image GPT/Pillow branchée
- social reste prod par défaut, mais --env existe pour compatibilité
"""

import os, sys, re, argparse, requests, anthropic, openpyxl, base64
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts")); return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent)); return
_ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, ENV_GLOBAL, ENV_IZILIFE, izilife_social_zone, normalize_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False
    def normalize_zone(zone: str) -> str:
        zone = str(zone or "").strip().lower()
        return zone if zone.endswith("-zone") else f"{zone}-zone"
    PROJECT_ROOT = Path(os.getenv("AGENTIC_WORKSPACE_ROOT", Path.home() / "Documents" / "agentic_Workspace"))
    ENV_GLOBAL = PROJECT_ROOT / ".env"
    ENV_IZILIFE = PROJECT_ROOT / "izilife" / ".env.izilife"

DEFAULT_SOCIAL_ENV = os.getenv("SOCIAL_ENV", "prod")
if HAS_CORE_PATHS:
    DRIVE_ROOT = izilife_social_zone("", DEFAULT_SOCIAL_ENV).parent
    CONTEXT_ROOT = PROJECT_ROOT / "izilife" / "context" / "social"
    TEMPLATES_DIR = PROJECT_ROOT / "izilife" / "templates"
else:
    DRIVE_ROOT = Path("G:/Mon Drive/agentic_workspace/izilife/social")
    CONTEXT_ROOT = PROJECT_ROOT / "izilife" / "context" / "social"
    TEMPLATES_DIR = PROJECT_ROOT / "izilife" / "templates"

CONTEXT_COMMON = ["izilife-social-strategy.md", "community-manager.md", "izilife-style.md"]
TYPE_CONTENT = {
    "POST_AGENDA_SEMAINE":"story", "POST_AGENDA_WEEKEND":"story", "POST_TOP_LIEUX":"carrousel",
    "POST_TOP_ACTEURS":"carrousel", "POST_HISTOIRE":"post", "POST_HUMOUR":"post",
    "POST_ESCAPADE_VILLE":"carrousel", "POST_ESCAPADE_NATURE":"carrousel", "POST_PEPITE":"post",
    "EVENT_SERIE":"post", "EVENT_ANIMATEUR":"post", "PARTAGE_ACTEUR":"post", "NOUVEAUTE_IZILIFE":"post",
    "REEL":"reel", "VIDEO":"video",
}
TYPE_TEMPLATE = {
    "POST_AGENDA_SEMAINE":"agenda_semaine", "POST_AGENDA_WEEKEND":"agenda_weekend",
    "POST_TOP_LIEUX":"top", "POST_TOP_ACTEURS":"top", "POST_HISTOIRE":"histoire_lieu",
    "POST_HUMOUR":"humour_local", "POST_ESCAPADE_VILLE":"escapade_ville", "POST_ESCAPADE_NATURE":"escapade_ville",
    "POST_PEPITE":"pepite_niche", "EVENT_SERIE":"event_serie", "EVENT_ANIMATEUR":"event_animateur",
}

COL_DATE=1; COL_RESEAU=2; COL_TYPE=3; COL_VILLE=4; COL_LIEU=5
COL_THEME=6; COL_SERIE=7; COL_ARTISTE=8; COL_STYLE=9; COL_DATA=10
COL_DATE_EVENT=11; COL_HEURE=12; COL_INPUT_ID=13
COL_TEMPLATE_SOURCE=14; COL_TEMPLATE_REF=15; COL_IMAGE_PROMPT=16; COL_IMAGE_PROVIDER=17
COL_POST_OUTPUT=18; COL_IMAGE_OUTPUT=19; COL_IMAGE_STATUS=20; COL_STATUT=21
STATUT_A_FAIRE="À faire"; STATUT_RELANCER="Relancer"; STATUT_GENERE="Généré"


def get_zone_config(zone: str, env_name: str = DEFAULT_SOCIAL_ENV) -> dict:
    z = normalize_zone(zone)
    drive = izilife_social_zone(z, env_name) if HAS_CORE_PATHS else DRIVE_ROOT / z
    return {"zone": z, "drive": drive, "excel": drive / f"planning_{z}.xlsx", "outputs": drive / "outputs",
            "inputs": drive / "inputs", "context_zone": CONTEXT_ROOT / "zones" / f"{z}.md", "templates": TEMPLATES_DIR}


def load_zone_env(cfg: dict):
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_IZILIFE.exists(): load_dotenv(ENV_IZILIFE, override=False)


def slugify(t: str) -> str:
    t = str(t or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9_]", "", t.lower().replace(" ", "_").replace("-", "_"))[:45] or "post"


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style='thin', color="CCCCCC"); brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hdr(r,c,v,bg="1A1A2E",ft="FFFFFF",sz=9):
        x=ws.cell(row=r,column=c,value=v); x.font=Font(name="Arial",size=sz,bold=True,color=ft)
        x.fill=PatternFill("solid",fgColor=bg); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=brd
    groups=[(1,3,"PLANIFICATION","1A1A2E"),(4,7,"CONTENU","0F3460"),(8,10,"DONNÉES","533483"),(11,12,"EVENT","1B4332"),(13,13,"INPUT","784212"),(14,15,"TEMPLATE","8E44AD"),(16,17,"IMAGE","0D6986"),(18,20,"OUTPUT","1A5276"),(21,21,"STATUT","6C3483")]
    for s,e,label,bg in groups:
        if s<e: ws.merge_cells(start_row=1,start_column=s,end_row=1,end_column=e)
        hdr(1,s,label,bg=bg,sz=10)
    headers=[("DATE",12),("RÉSEAU",11),("TYPE",22),("VILLE",12),("LIEU",16),("THEME / SUJET",24),("SERIE",16),("ARTISTE / DJ / ASSO",26),("STYLE",14),("DATA",38),("DATE EVENT",12),("HEURE / DURÉE",14),("INPUT_ID",16),("TEMPLATE_SOURCE\nowned/inspiration/none",22),("TEMPLATE_REF\nchemin / Canva ID / URL",32),("IMAGE_PROMPT",38),("IMAGE_PROVIDER\ngpt/canva/pillow",16),("POST OUTPUT",60),("IMAGE_OUTPUT",38),("IMAGE_STATUS",16),("STATUT",13)]
    for i,(label,width) in enumerate(headers,1):
        hdr(2,i,label); ws.column_dimensions[get_column_letter(i)].width=width
    ws.freeze_panes="D3"; ws.row_dimensions[2].height=42


def _add_validations(ws):
    from openpyxl.worksheet.datavalidation import DataValidation
    pairs=[
        (DataValidation(type="list",formula1='"Instagram,Facebook,Les deux"',allow_blank=True), COL_RESEAU),
        (DataValidation(type="list",formula1='"' + ','.join(TYPE_CONTENT.keys()) + '"',allow_blank=True), COL_TYPE),
        (DataValidation(type="list",formula1='"owned,inspiration,none"',allow_blank=True), COL_TEMPLATE_SOURCE),
        (DataValidation(type="list",formula1='"gpt,canva,pillow"',allow_blank=True), COL_IMAGE_PROVIDER),
        (DataValidation(type="list",formula1='"non demandé,généré,erreur,source fournie,template utilisé"',allow_blank=True), COL_IMAGE_STATUS),
        (DataValidation(type="list",formula1='"À faire,Généré,Relancer,Validé,Publié,Skip"',allow_blank=True), COL_STATUT),
    ]
    for dv,col in pairs:
        ws.add_data_validation(dv)
        for r in range(3,503): dv.add(ws.cell(row=r,column=col))


def ensure_planning(path: Path):
    wb = openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    ws.title = "Planning"
    _style_header(ws); _add_validations(ws)
    for r in range(3, max(ws.max_row,3)+1):
        if not ws.cell(r,COL_TEMPLATE_SOURCE).value: ws.cell(r,COL_TEMPLATE_SOURCE).value="none"
        if not ws.cell(r,COL_IMAGE_STATUS).value: ws.cell(r,COL_IMAGE_STATUS).value="non demandé"
    if "Post Types" in wb.sheetnames: del wb["Post Types"]
    wt=wb.create_sheet("Post Types"); wt.append(["POST_TYPE","CONTENT_TYPE","TEMPLATE_LOCAL","TEMPLATE_SOURCE","NOTES"])
    for pt,ct in TYPE_CONTENT.items(): wt.append([pt,ct,f"templates/{TYPE_TEMPLATE.get(pt,'—')}/","owned|inspiration|none","owned=template fixe ; inspiration=référence non copiée ; none=libre"])
    path.parent.mkdir(parents=True, exist_ok=True); wb.save(path)


def update_docs():
    cm = CONTEXT_ROOT / "community-manager.md"
    block = """
---

## Règle TEMPLATE_SOURCE v3

Dans le planning, deux colonnes pilotent les visuels :

- TEMPLATE_SOURCE = owned
  Le template est un asset izilife ou un design Canva appartenant à izilife.
  L'agent peut l'utiliser comme base fixe.

- TEMPLATE_SOURCE = inspiration
  La référence vient d'ailleurs ou d'une création qui ne doit pas être copiée.
  L'agent doit seulement s'en inspirer : ambiance, cadrage, rythme, intention.

- TEMPLATE_SOURCE = none
  Aucun template. L'image est générée librement depuis IMAGE_PROMPT, INPUT_ID ou DATA.

TEMPLATE_REF contient le chemin, Canva ID, URL ou note courte associée.
"""
    cm.parent.mkdir(parents=True, exist_ok=True)
    if cm.exists():
        txt=cm.read_text(encoding="utf-8")
        if "## Règle TEMPLATE_SOURCE v3" not in txt: cm.write_text(txt.rstrip()+"\n"+block,encoding="utf-8")


def init_zone(zone: str, env_name: str = DEFAULT_SOCIAL_ENV):
    cfg=get_zone_config(zone, env_name)
    for d in ["outputs","inputs"]: (cfg["drive"]/d).mkdir(parents=True,exist_ok=True)
    cfg["context_zone"].parent.mkdir(parents=True,exist_ok=True)
    if not cfg["context_zone"].exists():
        cfg["context_zone"].write_text(f"""# Contexte zone — {cfg['zone']}
## Spécificités
- Ville principale :
- Hashtags locaux : #{cfg['zone'].replace('-zone','')}
- Compte Instagram : @izilife_{cfg['zone'].replace('-zone','')}
- Événements locaux :
- Lieux emblématiques :

## Templates & inspirations
- TEMPLATE_SOURCE = owned : template/design izilife que l'on peut utiliser comme base fixe.
- TEMPLATE_SOURCE = inspiration : référence externe ou ancienne créa servant uniquement d'inspiration. Ne pas copier.
- TEMPLATE_SOURCE = none : pas de référence, génération libre selon le brief.
""", encoding="utf-8")
    ensure_planning(cfg["excel"]); update_docs()
    print(f"✅ Zone prête : {cfg['zone']}\n   Planning : {cfg['excel']}\n   Contexte : {cfg['context_zone']}")


def load_context(zone: str) -> str:
    parts=[]
    for fn in CONTEXT_COMMON:
        p=CONTEXT_ROOT/fn
        if p.exists(): parts.append(f"### {fn}\n{p.read_text(encoding='utf-8')}")
    zp=CONTEXT_ROOT/"zones"/f"{zone}.md"
    if zp.exists(): parts.append(f"### {zone}.md\n{zp.read_text(encoding='utf-8')}")
    return "\n\n---\n\n".join(parts)


def read_input(input_id: str, inputs_dir: Path):
    if not input_id: return "", []
    folder=inputs_dir/input_id
    if not folder.exists(): return "", []
    desc=(folder/"description.txt").read_text(encoding="utf-8").strip() if (folder/"description.txt").exists() else ""
    imgs=sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg",".jpeg",".png",".webp")])
    return desc,imgs


def build_prompt(row, content_type: str, input_desc: str) -> str:
    def v(c):
        val=row[c-1].value; return str(val).strip() if val else ""
    slides="\n=== SLIDES ===\nSlide 1 : ...\nSlide 2 : ..." if content_type=="carrousel" else ""
    return f"""Tu es l'agent Community Manager izilife. CONTENT_TYPE: {content_type}

DONNÉES :
TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
THEME: {v(COL_THEME)} | SERIE: {v(COL_SERIE)}
ARTISTE/DJ/ASSO: {v(COL_ARTISTE)} | STYLE: {v(COL_STYLE)}
DATA: {v(COL_DATA)}
DATE EVENT: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
RESEAU: {v(COL_RESEAU)}

TEMPLATE_SOURCE: {v(COL_TEMPLATE_SOURCE) or 'none'}
TEMPLATE_REF: {v(COL_TEMPLATE_REF)}
IMAGE_PROMPT: {v(COL_IMAGE_PROMPT)}

Règle template : owned = template izilife fixe utilisable ; inspiration = référence à ne pas copier ; none = libre.

DESCRIPTION INPUT :
{input_desc}

Format OBLIGATOIRE :
=== TEXTE DU POST ===
[caption complète avec emojis + hashtags]{slides}
"""


def call_llm(system: str, user: str) -> str:
    provider=os.getenv("LLM_PROVIDER","claude").lower()
    if provider=="openai":
        from openai import OpenAI
        client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r=client.chat.completions.create(model=os.getenv("OPENAI_MODEL","gpt-4o"),messages=[{"role":"system","content":system},{"role":"user","content":user}],max_tokens=1500)
        return r.choices[0].message.content
    client=anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    msg=client.messages.create(model=os.getenv("CLAUDE_MODEL","claude-sonnet-4-6"),max_tokens=1500,system=system,messages=[{"role":"user","content":user}])
    return msg.content[0].text


def resolve_template_ref(template_source: str, template_ref: str, post_type: str, cfg: dict) -> Path | None:
    if template_source != "owned": return None
    candidates=[]
    if template_ref:
        p=Path(template_ref); candidates += [p, cfg["templates"]/template_ref, cfg["drive"]/template_ref]
    tpl=TYPE_TEMPLATE.get(post_type)
    if tpl: candidates.append(cfg["templates"]/tpl)
    for c in candidates:
        if c.is_file(): return c
        if c.is_dir():
            files=sorted(list(c.glob("*.png"))+list(c.glob("*.jpg"))+list(c.glob("*.jpeg")))
            if files: return files[0]
    return None


def generate_image_gpt(prompt: str, folder: Path, filename: str) -> Path | None:
    if not prompt: return None
    try:
        from openai import OpenAI
        client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        result=client.images.generate(model=os.getenv("IMAGE_MODEL","gpt-image-1"),prompt=f"{prompt}. Style Instagram, vertical, haute qualité.",n=1,size=os.getenv("IMAGE_SIZE","1024x1536"))
        b64=getattr(result.data[0],"b64_json",None)
        if b64:
            out=folder/f"{filename}_gpt.png"; out.write_bytes(base64.b64decode(b64)); return out
        url=getattr(result.data[0],"url",None)
        if url:
            out=folder/f"{filename}_gpt.jpg"; out.write_bytes(requests.get(url,timeout=60).content); return out
    except Exception as e: print(f"    ❌ GPT Image : {e}")
    return None


def compose_pillow(bg: Path, text: str, folder: Path, filename: str) -> Path | None:
    try:
        from PIL import Image, ImageDraw, ImageFont
        img=Image.open(bg).convert("RGBA"); W,H=img.size
        img=Image.alpha_composite(img, Image.new("RGBA",(W,H),(0,0,0,125)))
        draw=ImageDraw.Draw(img)
        try: font=ImageFont.truetype("arial.ttf",max(34,W//22))
        except Exception: font=ImageFont.load_default()
        y=H//4
        for line in [l.strip() for l in text.splitlines() if l.strip() and not l.startswith("===")][:7]:
            line=line[:55]; bb=draw.textbbox((0,0),line,font=font); x=max(30,(W-(bb[2]-bb[0]))//2)
            draw.text((x+2,y+2),line,font=font,fill=(0,0,0,180)); draw.text((x,y),line,font=font,fill=(255,255,255,255)); y+=(bb[3]-bb[1])+18
        out=folder/f"{filename}_pillow.jpg"; img.convert("RGB").save(out,quality=95); return out
    except Exception as e: print(f"    ❌ Pillow : {e}")
    return None


def process_images(row, text: str, input_images: list[Path], folder: Path, filename: str, cfg: dict):
    def v(c): val=row[c-1].value; return str(val).strip() if val else ""
    src=(v(COL_TEMPLATE_SOURCE) or "none").lower(); ref=v(COL_TEMPLATE_REF); prompt=v(COL_IMAGE_PROMPT)
    provider=(v(COL_IMAGE_PROVIDER) or os.getenv("IMAGE_LLM","gpt")).lower(); engine=os.getenv("IMAGE_ENGINE","gpt").lower()
    if src=="inspiration" and ref: prompt=f"{prompt}\nInspiration visuelle à respecter sans copier : {ref}".strip()
    tpl=resolve_template_ref(src, ref, v(COL_TYPE), cfg); bg=input_images[0] if input_images else tpl
    generated=[]
    if bg and engine in ("pillow","both"):
        out=compose_pillow(bg,text,folder,filename)
        if out: generated.append(out)
    if provider=="gpt" and engine in ("gpt","both") and prompt:
        out=generate_image_gpt(prompt,folder,filename)
        if out: generated.append(out)
    return generated


def save_post(text,input_images,generated,content_type,folder):
    lines=[text.strip(),"",f"=== CONTENT_TYPE ===\n{content_type}"]
    if input_images: lines += ["","=== IMAGES SOURCE ==="]+[f"  {i.name}" for i in input_images]
    if generated: lines += ["","=== IMAGES GENEREES ==="]+[f"  {i.name}" for i in generated]
    (folder/"post.txt").write_text("\n".join(lines),encoding="utf-8")


def run_zone(zone: str, env_name: str = DEFAULT_SOCIAL_ENV, dry_run: bool = False):
    cfg=get_zone_config(zone, env_name); load_zone_env(cfg)
    if not cfg["excel"].exists(): print(f"❌ Planning introuvable. Lance : python cm_izilife.py --zone {zone} --init"); return
    system=load_context(cfg["zone"])
    wb=openpyxl.load_workbook(cfg["excel"]); ws=wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    rows=[r for r in ws.iter_rows(min_row=3) if r[COL_STATUT-1].value in (STATUT_A_FAIRE, STATUT_RELANCER)]
    print(f"\n🤖 CM izilife v3 — {cfg['zone']} — {len(rows)} ligne(s)" + (" [DRY RUN]" if dry_run else ""))
    cfg["outputs"].mkdir(parents=True,exist_ok=True); cfg["inputs"].mkdir(parents=True,exist_ok=True)
    for idx,row in enumerate(rows,1):
        def v(c): val=row[c-1].value; return str(val).strip() if val else ""
        post_type=v(COL_TYPE); theme=v(COL_THEME); ville=v(COL_VILLE); ct=TYPE_CONTENT.get(post_type,"post")
        print(f"  [{idx}/{len(rows)}] {post_type} — {ville} — {theme}")
        if dry_run: continue
        folder=cfg["outputs"]/f"{datetime.now().strftime('%Y%m%d')}_{slugify(post_type)}_{slugify(theme or ville)}"; folder.mkdir(parents=True,exist_ok=True)
        desc,imgs=read_input(v(COL_INPUT_ID), cfg["inputs"])
        text=call_llm(system, build_prompt(row,ct,desc)); gen=process_images(row,text,imgs,folder,slugify(theme or post_type),cfg)
        save_post(text,imgs,gen,ct,folder)
        row[COL_POST_OUTPUT-1].value=text[:32000]; row[COL_IMAGE_OUTPUT-1].value="\n".join(str(p) for p in gen)
        row[COL_IMAGE_STATUS-1].value="généré" if gen else ("source fournie" if imgs else "non demandé"); row[COL_STATUT-1].value=STATUT_GENERE
    if not dry_run: wb.save(cfg["excel"]); print(f"✅ Planning mis à jour : {cfg['excel']}")


def list_zones(env_name: str = DEFAULT_SOCIAL_ENV):
    print("\n📍 Zones CM izilife :\n")
    if not DRIVE_ROOT.exists(): print(f"  ❌ Drive introuvable : {DRIVE_ROOT}"); return
    for d in sorted(DRIVE_ROOT.iterdir()):
        if d.is_dir(): print(f"  • {d.name}")


def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--zone","-z"); ap.add_argument("--env",default=DEFAULT_SOCIAL_ENV,choices=["local","staging","prod"])
    ap.add_argument("--init",action="store_true"); ap.add_argument("--list",action="store_true"); ap.add_argument("--update-md",action="store_true"); ap.add_argument("--dry-run",action="store_true")
    args=ap.parse_args()
    if args.list: list_zones(args.env); return
    if args.update_md: update_docs(); return
    if not args.zone: ap.error("--zone requis sauf --list/--update-md")
    if args.init: init_zone(args.zone,args.env); return
    run_zone(args.zone,args.env,args.dry_run)
if __name__=="__main__": main()
