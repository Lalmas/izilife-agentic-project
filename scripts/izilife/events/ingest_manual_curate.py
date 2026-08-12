"""
ingest_manual_curate.py
-----------------------
Curation manuelle — lit curate_events.xlsx depuis Google Drive par zone
et dispatche vers le bon handler selon l'onglet.

Onglets Excel : Facebook | Shotgun | HelloAsso | Eventbrite | Meetup | Instagram

Colonnes : url | status | notes | added_at | processed_at
Status   : pending → done | error | skip

Usage :
    python ingest_manual_curate.py --zone=lille --env=local --city=lille --dry-run
    python ingest_manual_curate.py --zone=lille --env=local --city=lille
    python ingest_manual_curate.py --zone=lille --env=local --city=lille --sheet=Facebook
    python ingest_manual_curate.py --zone=lille --env=local --city=lille --init

Prérequis :
    pip install playwright requests playwright-stealth openpyxl
    python -m playwright install chromium
    Variable : IZILIFE_AGENT_TOKEN
"""

import os
import sys

import re
import json
import time
import random
import argparse
import requests
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

# ─────────────────────────────────────────────
# CHEMINS — cross-platform Windows / Linux
# ─────────────────────────────────────────────

def get_drive_root() -> Path:
    value = os.environ.get("AGENTIC_DRIVE_ROOT", "").strip()
    if not value:
        raise RuntimeError("AGENTIC_DRIVE_ROOT non défini.")
    return Path(value).expanduser()
    # Ancienne détection conservée temporairement mais désormais inaccessible.
    """Retourne la racine Google Drive selon l'OS."""
    if sys.platform == "win32":
        # Google Drive for Desktop sur Windows
        candidates = [
            Path(value),
            Path.home() / "Google Drive",
            Path.home() / "Mon Drive",
        ]
    else:
        # Linux / Mac
        candidates = [
            Path.home() / "GoogleDrive",
            Path.home() / "Google Drive",
            Path.home() / "gdrive",
            Path(value),
        ]

    for p in candidates:
        if p.exists():
            return p

    # Fallback — dossier local si Drive pas trouvé
    fallback = Path(__file__).parent.parent.parent.parent / "izilife-agent-workspace" / "curate"
    fallback.mkdir(parents=True, exist_ok=True)
    print(f"⚠️  Google Drive non trouvé — fallback local : {fallback}")
    return fallback


def get_curate_file(zone: str) -> Path:
    """Retourne le chemin du fichier Excel pour une zone/env."""
    return event_curate_file(zone, CURRENT_ENV)


def get_download_dir(zone: str) -> Path:
    """Retourne l'ancien dossier de téléchargement intermédiaire."""
    return event_download_dir(zone, CURRENT_ENV, downloads=True)


def get_image_dir(zone: str) -> Path:
    """
    Retourne directement le dossier racine des images de la zone.
    Exemple : images/lille-zone/
    """
    return event_download_dir(zone, CURRENT_ENV, downloads=False)


def get_cookies_path(platform: str) -> Path:
    """Retourne le chemin des cookies pour une plateforme."""
    scripts_dir = Path(__file__).parent
    return scripts_dir / "profiles" / platform / f"{platform}_cookies.json"


# ─────────────────────────────────────────────
# CONFIGURATION — envs via core.paths avec fallback local
# ─────────────────────────────────────────────

CORE_ROOT = Path(__file__).resolve().parents[2]
if str(CORE_ROOT) not in sys.path:
    sys.path.append(str(CORE_ROOT))
try:
    from core.paths import (
        IZILIFE_ENVS, normalize_zone, workspace_root, local_agent_workspace_root,
        event_curate_file, event_download_dir, event_source_dir, event_sent_log_file
    )
    ENVS = IZILIFE_ENVS
except Exception:
    def normalize_zone(zone: str) -> str:
        zone = str(zone or "").strip().lower()
        return zone if zone.endswith("-zone") else f"{zone}-zone"
    ENVS = {
        "local":   {"base_url": "https://localhost:4443/izilife-admin",          "verify_ssl": False},
        "staging": {"base_url": "https://www.staging.izilife.co/izilife-admin", "verify_ssl": True},
        "prod":    {"base_url": "https://www.izilife.co/izilife-admin",          "verify_ssl": True},
    }
    def local_agent_workspace_root(env_name="prod"):
        suffix = {"local":"-local", "staging":"-staging", "prod":""}.get(env_name, "")
        return Path.home() / "Documents" / "agentic_Workspace" / "izilife" / f"izilife-agent-workspace{suffix}"
    def workspace_root(env_name="prod"):
        folder = {"local":"agentic_workspace_local", "staging":"agentic_workspace_staging", "prod":"agentic_workspace"}.get(env_name, "agentic_workspace")
        return get_drive_root() / folder
    def event_curate_file(zone, env_name="prod"):
        return workspace_root(env_name) / "izilife" / "events" / normalize_zone(zone) / "curate_events.xlsx"
    def event_download_dir(zone, env_name="prod", downloads=True):
        d = local_agent_workspace_root(env_name) / "images" / normalize_zone(zone) / ("downloads" if downloads else "")
        d.mkdir(parents=True, exist_ok=True)
        return d
    def event_source_dir(platform, zone, env_name="prod"):
        d = local_agent_workspace_root(env_name) / platform / normalize_zone(zone) / "events"
        d.mkdir(parents=True, exist_ok=True)
        return d
    def event_sent_log_file(env_name="prod"):
        d = local_agent_workspace_root(env_name) / "logs"
        d.mkdir(parents=True, exist_ok=True)
        return d / "event_images_sent.txt"

AGENT_TOKEN = os.environ.get("IZILIFE_AGENT_TOKEN", "METTRE_TOKEN_ICI")
CURRENT_ENV = "prod"

def set_current_env(env_name: str):
    global CURRENT_ENV
    CURRENT_ENV = str(env_name or "prod").lower()

FB_EMAIL    = os.environ.get("FB_EMAIL", "")
FB_PASSWORD = os.environ.get("FB_PASSWORD", "")
UPLOAD_PATH = "/scraper/agentUploadEventSources/{city_id}"
IMAGE_PATH  = "/scraper/agentUploadEventImages/{city_id}"

WAIT_SELECTORS = {
    "shotgun.live":   "h1",
    "helloasso.com":  "h1",
    "billetweb.fr":   "h1",
    "eventbrite.fr":  '[data-testid="event-title"], h1',
    "meetup.com":     "h1",
    "facebook.com":   "h1",
    "weezevent.com":  "h1",
    "festik.net":     "h1",
}

FB_LOGIN_URL = "https://www.facebook.com/login"

IGRAM_SITES = [
    {
        "name": "igram.world",
        "url": "https://igram.world/fr/",
        "input_sel": "input[type='text']",
        "result_timeout": 90000,
    },
    {
        "name": "snapinsta.app",
        "url": "https://snapinsta.app/fr",
        "input_sel": "input[type='text']",
        "btn_sel": "button[type='submit']",
        "wait_sel": "a[href*='.jpg'], a[download]",
        "result_timeout": 30000,
    },
    {
        "name": "instafinsta.com",
        "url": "https://instafinsta.com",
        "input_sel": "input[name='url'], input[type='text']",
        "btn_sel": "button[type='submit']",
        "wait_sel": "a[download], a[href*='.jpg']",
        "result_timeout": 30000,
    },
]

SHEET_HANDLERS = {
    "Facebook":   "html",
    "Shotgun":    "html",
    "HelloAsso":  "html",
    "Eventbrite": "html",
    "Meetup":     "html",
    "Instagram":  "instagram",
}

DELAY = (2, 5)

# ─────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def sleep_random(mn=2, mx=5):
    time.sleep(random.uniform(mn, mx))


def apply_stealth(page):
    if HAS_STEALTH:
        stealth_sync(page)
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['fr-FR', 'fr']});
        window.chrome = { runtime: {} };
    """)


def detect_domain(url: str) -> str:
    return urlparse(url).netloc.lower().replace("www.", "")


def load_cookies(context, platform: str):
    path = get_cookies_path(platform)
    if path.exists():
        cookies = json.loads(path.read_text())
        valid   = [c for c in cookies if c.get("name") and c.get("value")]
        context.add_cookies(valid)
        log(f"  🍪 {len(valid)} cookies chargés ({platform})")
    else:
        log(f"  ⚠️  Cookies non trouvés : {path}")


def send_html(html: str, filename: str, base_url: str, city_id: int,
              verify_ssl: bool, source_url: str) -> bool:
    url = base_url + UPLOAD_PATH.format(city_id=city_id)
    try:
        r = requests.post(
            url,
            files={"sources[]": (filename, html.encode("utf-8"), "text/html")},
            data={"source_urls[]": source_url},
            headers={"X-Agent-Token": AGENT_TOKEN},
            verify=verify_ssl, timeout=30
        )
        if r.status_code == 200:
            resp = r.json()
            log(f"  ✅ inserted={resp.get('inserted',0)} skipped={resp.get('skipped',0)}")
            return True
        log(f"  ❌ HTTP {r.status_code} : {r.text[:200]}")
        return False
    except Exception as e:
        log(f"  ❌ Erreur envoi : {e}")
        return False


def send_image(image_path: Path, base_url: str, city_id: int, verify_ssl: bool) -> bool:
    url  = base_url + IMAGE_PATH.format(city_id=city_id)
    mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png",  "webp": "image/webp"}.get(
                image_path.suffix.lstrip(".").lower(), "image/jpeg")
    try:
        with open(image_path, "rb") as f:
            r = requests.post(url,
                              files={"images[]": (image_path.name, f, mime)},
                              headers={"X-Agent-Token": AGENT_TOKEN},
                              verify=verify_ssl, timeout=60)
        if r.status_code == 200:
            log(f"  ✅ Image envoyée → OCR")
            return True
        log(f"  ❌ HTTP {r.status_code} : {r.text[:150]}")
        return False
    except Exception as e:
        log(f"  ❌ Erreur envoi image : {e}")
        return False


def resolve_city_id(city_slug: str, base_url: str, verify_ssl: bool) -> int:
    url = f"{base_url}/scraper/cityByStringId/{city_slug}"
    try:
        r = requests.get(url, headers={"X-Agent-Token": AGENT_TOKEN},
                         verify=verify_ssl, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data.get('success') and data.get('city'):
                city_id = int(data['city']['id'])
                log(f"Ville résolue : {city_slug} → city_id={city_id}")
                return city_id
    except Exception as e:
        log(f"  ⚠️  Erreur résolution ville : {e}")
    log(f"❌ Ville introuvable : {city_slug}")
    sys.exit(1)


# ─────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────

def read_pending_rows(sheet) -> list:
    headers = [str(c.value).strip().lower() if c.value else "" for c in sheet[1]]
    rows    = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {headers[i]: (cell.value or "") for i, cell in enumerate(row) if i < len(headers)}
        status = str(values.get("status", "")).strip().lower()
        url    = str(values.get("url", "")).strip()
        if status == "pending" and url.startswith("http"):
            rows.append({"row_idx": row_idx, "url": url,
                         "notes": values.get("notes", "")})
    return rows


def update_row(sheet, row_idx: int, status: str):
    headers = [str(c.value).strip().lower() if c.value else "" for c in sheet[1]]
    try:
        status_col    = headers.index("status") + 1
        processed_col = headers.index("processed_at") + 1
        sheet.cell(row=row_idx, column=status_col).value     = status
        sheet.cell(row=row_idx, column=processed_col).value  = datetime.now().strftime('%Y-%m-%d %H:%M')
    except ValueError:
        pass


def create_excel_template(zone: str):
    curate_file = get_curate_file(zone)
    curate_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers  = ["url", "status", "notes", "added_at", "processed_at"]
    examples = {
        "Facebook":   "https://www.facebook.com/events/123456789",
        "Shotgun":    "https://shotgun.live/fr/events/mon-event",
        "HelloAsso":  "https://www.helloasso.com/associations/asso/evenements/event",
        "Eventbrite": "https://www.eventbrite.fr/e/mon-event-123456",
        "Meetup":     "https://www.meetup.com/fr-FR/groupe/events/123456789/",
        "Instagram":  "https://www.instagram.com/p/ABC123DEF/",
    }

    for sheet_name in SHEET_HANDLERS.keys():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        ws.append([examples.get(sheet_name, ""), "pending", "exemple",
                   datetime.now().strftime('%Y-%m-%d'), ""])
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18

    wb.save(curate_file)
    log(f"✅ Fichier créé : {curate_file}")


# ─────────────────────────────────────────────
# HANDLERS
# ─────────────────────────────────────────────

def login_facebook(page) -> bool:
    """Login automatique Facebook depuis .env."""
    if not FB_EMAIL or not FB_PASSWORD:
        return False
    log(f"Tentative login auto : {FB_EMAIL}")
    try:
        page.goto(FB_LOGIN_URL, timeout=30000)
        sleep_random(2, 3)
        page.fill('#email', FB_EMAIL)
        sleep_random(0.5, 1)
        page.fill('#pass', FB_PASSWORD)
        sleep_random(0.5, 1)
        page.click('[name="login"]')
        sleep_random(4, 6)
        if "login" not in page.url and "checkpoint" not in page.url:
            log("✅ Login auto réussi")
            return True
        log("⚠️  Login auto échoué (2FA ou erreur)")
        return False
    except Exception as e:
        log(f"  ❌ Erreur login : {e}")
        return False


def handle_html(page, row: dict, base_url: str, city_id: int,
                verify_ssl: bool, dry_run: bool) -> bool:
    url      = row["url"]
    domain   = detect_domain(url)
    wait_sel = next((v for k, v in WAIT_SELECTORS.items() if k in domain), "h1")

    log(f"  [{domain}] {url}")

    if dry_run:
        log(f"  [DRY RUN] serait envoyé")
        return True

    try:
        page.goto(url, timeout=30000)
        try:
            page.wait_for_selector(wait_sel, timeout=12000)
        except:
            pass
        sleep_random(2, 5)

        html     = page.content()
        slug     = re.sub(r'[^a-z0-9\-]', '-', url.split('/')[-1].lower())[:50] or "event"
        prefix   = domain.split('.')[0]
        filename = f"{prefix}_{slug}.html"

        return send_html(html, filename, base_url, city_id, verify_ssl, url)

    except Exception as e:
        log(f"  ❌ {e}")
        return False



def _visible_indexes(locator):
    visible = []
    try:
        count = locator.count()
    except Exception:
        return visible
    for index in range(count):
        try:
            if locator.nth(index).is_visible():
                visible.append(index)
        except Exception:
            pass
    return visible


def _save_playwright_image_response(response, local_path: Path) -> bool:
    try:
        body = response.body()
        content_type = (response.headers.get("content-type") or "").lower()
        if not body or "image/" not in content_type:
            return False
        local_path.write_bytes(body)
        return True
    except Exception:
        return False

def handle_igram_consent(page, appearance_timeout: int = 15000) -> bool:
    """
    Attend l'apparition différée du CMP Google Funding Choices puis le refuse.
    `locator.is_visible()` ne sert pas d'attente fiable : on utilise `wait_for()`.
    """
    consent_root = page.locator(".fc-consent-root").first

    try:
        log("  ⏳ Attente éventuelle du consentement igram...")
        consent_root.wait_for(
            state="visible",
            timeout=appearance_timeout
        )
    except Exception:
        log("  ℹ️ Aucun consentement igram affiché")
        return False

    log("  🍪 Fenêtre de consentement igram détectée")

    # Le texte et les classes internes peuvent varier : on cible d'abord
    # tous les boutons du CMP puis on filtre leur libellé.
    buttons = consent_root.locator("button")

    deadline = time.time() + 10

    while time.time() < deadline:
        try:
            count = buttons.count()
        except Exception:
            count = 0

        for index in range(count):
            button = buttons.nth(index)

            try:
                if not button.is_visible():
                    continue

                label = " ".join(
                    filter(
                        None,
                        [
                            button.inner_text(timeout=1000).strip(),
                            (button.get_attribute("aria-label") or "").strip(),
                            (button.get_attribute("title") or "").strip(),
                        ]
                    )
                ).lower()
            except Exception:
                continue

            if any(
                expected in label
                for expected in (
                    "refuser",
                    "tout refuser",
                    "refuser tout",
                    "continuer sans accepter",
                )
            ):
                button.scroll_into_view_if_needed()
                button.click(force=True, timeout=5000)

                try:
                    consent_root.wait_for(
                        state="hidden",
                        timeout=10000
                    )
                except Exception:
                    # Certains CMP retirent l'overlay mais conservent brièvement
                    # la racine dans le DOM.
                    page.wait_for_timeout(1000)

                log("  🍪 Consentement igram refusé")
                return True

        page.wait_for_timeout(500)

    # Dernier secours : croix en haut à droite.
    close_selectors = [
        "button.fc-close",
        "button[aria-label*='fermer' i]",
        "button[aria-label*='close' i]",
        "[role='button'][aria-label*='fermer' i]",
        "[role='button'][aria-label*='close' i]",
    ]

    for selector in close_selectors:
        try:
            close_button = consent_root.locator(selector).first
            close_button.wait_for(state="visible", timeout=1500)
            close_button.click(force=True, timeout=3000)
            page.wait_for_timeout(1000)
            log("  🍪 Fenêtre de consentement igram fermée")
            return True
        except Exception:
            pass

    log("  ⚠️ Bouton Refuser du CMP introuvable")
    return False


def close_igram_ad_popup(page) -> bool:
    """Ferme uniquement l'interstitiel publicitaire, jamais son bouton Ouvrir."""
    selectors = [
        "text=Fermer",
        "[aria-label='Fermer' i]",
        "[aria-label='Close' i]",
        "[title='Fermer' i]",
        "[title='Close' i]",
    ]

    for scope in [page, *page.frames]:
        for selector in selectors:
            try:
                close_button = scope.locator(selector).first
                if close_button.is_visible(timeout=250):
                    close_button.click(force=True, timeout=1500)
                    page.wait_for_timeout(500)
                    log("  ✅ Publicité interstitielle fermée")
                    return True
            except Exception:
                pass

    if "google_vignette" in page.url:
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(300)
        except Exception:
            pass

    return False


def download_from_igram_world(
    page,
    instagram_url: str,
    image_dir: Path,
    pid: str
) -> list:
    """
    Parcours igram.world comme un utilisateur :
      1. ouvrir la page d'accueil française ;
      2. traiter le consentement différé ;
      3. cliquer sur Photo ;
      4. coller l'URL Instagram ;
      5. cliquer sur le premier bouton Télécharger ;
      6. attendre le résultat ;
      7. récupérer directement le href du bouton final ;
      8. enregistrer l'image dans images/<zone>/.
    """
    page.goto(
        "https://igram.world/fr/",
        wait_until="domcontentloaded",
        timeout=30000
    )

    handle_igram_consent(page, appearance_timeout=15000)

    consent_root = page.locator(".fc-consent-root").first
    try:
        consent_root.wait_for(state="hidden", timeout=3000)
    except Exception:
        if consent_root.is_visible():
            raise RuntimeError(
                "Le bandeau de consentement igram bloque toujours la page"
            )

    # Accéder à l'outil Photo depuis la page d'accueil.
    photo_link = page.locator(
        "a[href*='/fr/photo'], "
        "a:has-text('Photo'), "
        "button:has-text('Photo')"
    ).first

    try:
        photo_link.wait_for(state="visible", timeout=15000)
        photo_link.click()
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        # Secours pragmatique si le menu change.
        page.goto(
            "https://igram.world/fr/photo",
            wait_until="domcontentloaded",
            timeout=30000
        )

    # Le CMP peut parfois réapparaître après la navigation interne.
    handle_igram_consent(page, appearance_timeout=5000)

    input_box = page.locator(
        "form.search-form input[type='text'], "
        "input[type='text']"
    ).first

    input_box.wait_for(state="visible", timeout=15000)
    input_box.fill(instagram_url)

    sleep_random(1, 2)

    # Premier bouton Télécharger : submit du formulaire.
    submit_button = page.locator(
        "form.search-form button.search-form__button[type='submit'], "
        "form.search-form button[type='submit']"
    ).first

    submit_button.wait_for(state="visible", timeout=15000)
    submit_button.click()

    log("  ⏳ URL envoyée à igram.world, attente du résultat...")

    # Le résultat réel contient un lien .button__download.
    result_link = page.locator(
        ".output-list a.button__download[href], "
        ".output-component a.button__download[href], "
        ".search-result a.button__download[href]"
    ).last

    href = None
    for _ in range(20):
        close_igram_ad_popup(page)
        try:
            if result_link.is_visible(timeout=250):
                href = result_link.get_attribute("href")
                break
        except Exception:
            pass
        page.wait_for_timeout(1000)

    if not href:
        log("  ⚠️ Le bouton de téléchargement du résultat n'est pas apparu")
        close_igram_ad_popup(page)
        candidates = page.locator("img").evaluate_all("""
            images => images.map(img => ({
                src: img.currentSrc || img.src || '',
                width: img.naturalWidth || 0,
                height: img.naturalHeight || 0,
                context: `${img.alt || ''} ${img.className || ''} ${img.closest('a,article,section,div')?.className || ''}`.toLowerCase()
            })).filter(item =>
                /^https?:\/\//i.test(item.src)
                && item.width >= 300
                && item.height >= 300
                && !/(logo|avatar|icon|favicon|advert|publicit|banner|doubleclick|googleads)/i.test(`${item.src} ${item.context}`)
            ).sort((a, b) => (b.width * b.height) - (a.width * a.height))
        """)
        if not candidates:
            log("  ⚠️ Aucune grande image de résultat exploitable")
            return []
        href = candidates[0]["src"]
        log(
            f"  ✅ Image de résultat détectée directement "
            f"({candidates[0]['width']}x{candidates[0]['height']})"
        )

    if not href or not href.startswith("http"):
        log("  ⚠️ Le lien de téléchargement igram est absent ou invalide")
        return []

    log("  ✅ Résultat igram.world prêt")

    response = page.request.get(
        href,
        headers={"Referer": page.url},
        timeout=90000
    )

    if not response.ok:
        log(f"  ⚠️ Téléchargement HTTP échoué : {response.status}")
        return []

    content_type = (
        response.headers.get("content-type") or ""
    ).lower()

    if "png" in content_type:
        suffix = ".png"
    elif "webp" in content_type:
        suffix = ".webp"
    else:
        suffix = ".jpg"

    image_dir.mkdir(parents=True, exist_ok=True)
    local_path = image_dir / f"ig_{pid}_0{suffix}"
    local_path.write_bytes(response.body())

    if local_path.stat().st_size < 10_000:
        log(
            f"  ⚠️ Fichier suspect, seulement "
            f"{local_path.stat().st_size} octets"
        )
        local_path.unlink(missing_ok=True)
        return []

    log(
        f"  📥 Image enregistrée : {local_path.name} "
        f"({local_path.stat().st_size // 1024} Ko)"
    )

    return [local_path]


def handle_instagram(page, row: dict, base_url: str, city_id: int,
                     verify_ssl: bool, dry_run: bool, zone: str) -> bool:
    url = row["url"]
    log(f"  [INSTAGRAM] {url}")

    if dry_run:
        log("  [DRY RUN] serait téléchargé et envoyé")
        return True

    image_dir = get_image_dir(zone)
    post_id = re.search(r'/p/([A-Za-z0-9_-]+)', url)
    pid = post_id.group(1) if post_id else "unknown"

    for igram in IGRAM_SITES:
        log(f"  Tentative via {igram['name']}...")

        try:
            if igram["name"] == "igram.world":
                local_files = download_from_igram_world(
                    page,
                    url,
                    image_dir,
                    pid
                )

                if not local_files:
                    log("  ⚠️ igram.world : aucun fichier récupéré")
                    continue

                ok = False

                for local_path in local_files:
                    log(
                        f"  📤 Envoi OCR : {local_path.name} "
                        f"({local_path.stat().st_size // 1024} Ko)"
                    )

                    if send_image(
                        local_path,
                        base_url,
                        city_id,
                        verify_ssl
                    ):
                        ok = True

                if ok:
                    return True

                continue

            # Sites de secours : on conserve le flux générique,
            # mais les images restent également à la racine de la zone.
            page.goto(
                igram["url"],
                wait_until="domcontentloaded",
                timeout=30000
            )

            sleep_random(2, 3)

            page.locator(
                igram["input_sel"]
            ).first.fill(url)

            sleep_random(1, 2)

            page.locator(
                igram["btn_sel"]
            ).first.click()

            try:
                page.wait_for_selector(
                    igram["wait_sel"],
                    timeout=igram.get("result_timeout", 30000)
                )
            except Exception:
                log(f"  ⚠️ {igram['name']} : pas de résultat")
                continue

            img_urls = page.evaluate("""
                () => Array.from(document.querySelectorAll(
                    'a[href*=".jpg"], a[href*=".jpeg"], '
                    + 'a[href*=".png"], a[href*=".webp"], a[download]'
                ))
                .map(a => a.href)
                .filter(h => h.startsWith('http'))
            """)

            if not img_urls:
                log(f"  ⚠️ {igram['name']} : aucune image")
                continue

            log(f"  → {len(img_urls)} image(s)")
            ok = False

            for i, img_url in enumerate(img_urls):
                try:
                    response = page.request.get(
                        img_url,
                        timeout=30000
                    )

                    content_type = (
                        response.headers.get("content-type") or ""
                    ).lower()

                    suffix = (
                        ".png" if "png" in content_type
                        else ".webp" if "webp" in content_type
                        else ".jpg"
                    )

                    fname = f"ig_{pid}_{i}{suffix}"
                    local_path = image_dir / fname

                    if not _save_playwright_image_response(
                        response,
                        local_path
                    ):
                        continue

                    log(
                        f"  📥 Image enregistrée : {fname} "
                        f"({local_path.stat().st_size // 1024} Ko)"
                    )

                    if send_image(
                        local_path,
                        base_url,
                        city_id,
                        verify_ssl
                    ):
                        ok = True

                except Exception as exc:
                    log(f"  ⚠️ image {i} : {exc}")

            if ok:
                return True

        except Exception as exc:
            log(f"  ⚠️ {igram['name']} : {exc}")
            continue

    log("  ❌ Tous les sites igram ont échoué")
    return False


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Curation manuelle izilife")
    parser.add_argument("--zone",     type=str, required=True,
                        help="Zone cible : lille, valenciennes, dunkerque, paris, ...")
    parser.add_argument("--env",      choices=ENVS.keys(), default="local")
    parser.add_argument("--city",     type=str, required=False, default="lille")
    parser.add_argument("--sheet",    type=str, default="all",
                        help="Onglet : Facebook|Shotgun|HelloAsso|Eventbrite|Meetup|Instagram|all")
    parser.add_argument("--dry-run",  action="store_true")
    parser.add_argument("--init",     action="store_true",
                        help="Créer le fichier Excel template pour la zone")
    args = parser.parse_args()
    set_current_env(args.env)

    if args.init:
        create_excel_template(args.zone)
        log(f"Lance : python ingest_manual_curate.py --zone={args.zone} --env=local --city=lille --dry-run")
        return

    env        = ENVS[args.env]
    base_url   = env["base_url"]
    verify_ssl = env["verify_ssl"]
    dry_run    = args.dry_run
    zone       = args.zone.lower().strip()

    curate_file = get_curate_file(zone)

    city_id = resolve_city_id(args.city, base_url, verify_ssl)

    log(f"=== ingest_manual_curate.py — zone={zone} env={args.env} city={args.city}" +
        (" [DRY RUN]" if dry_run else "") + " ===")
    log(f"Fichier : {curate_file}")

    if AGENT_TOKEN == "METTRE_TOKEN_ICI" and not dry_run:
        log("❌ IZILIFE_AGENT_TOKEN non défini.")
        sys.exit(1)

    if not curate_file.exists():
        log(f"❌ Fichier introuvable : {curate_file}")
        log(f"   Lance : python ingest_manual_curate.py --zone={zone} --init")
        sys.exit(1)

    wb = openpyxl.load_workbook(curate_file)

    sheets_to_process = (
        [s for s in SHEET_HANDLERS.keys() if s in wb.sheetnames]
        if args.sheet.lower() == "all"
        else [args.sheet] if args.sheet in wb.sheetnames else []
    )

    if not sheets_to_process:
        log(f"❌ Onglet(s) introuvable(s). Disponibles : {wb.sheetnames}")
        sys.exit(1)

    all_rows = {}
    for sheet_name in sheets_to_process:
        rows = read_pending_rows(wb[sheet_name])
        if rows:
            all_rows[sheet_name] = rows
            log(f"  {sheet_name} : {len(rows)} URLs pending")

    if not all_rows:
        log("✅ Aucune URL pending.")
        return

    log(f"\nTotal : {sum(len(r) for r in all_rows.values())} URLs")

    stats = {"done": 0, "errors": 0}

    with sync_playwright() as p:

        needs_chrome = any(s in all_rows for s in ["Facebook", "Instagram"])
        launch_args  = {"headless": False,
                        "args": ["--disable-blink-features=AutomationControlled",
                                 "--disable-notifications"]}
        if needs_chrome:
            launch_args["channel"] = "chrome"

        browser = p.chromium.launch(**launch_args)
        context = browser.new_context(locale="fr-FR")

        if "Facebook" in all_rows:
            load_cookies(context, "facebook")

        page = context.new_page()
        apply_stealth(page)

        # Vérifier connexion Facebook si nécessaire
        if "Facebook" in all_rows:
            page.goto("https://www.facebook.com", timeout=30000)
            sleep_random(2, 3)
            if "login" in page.url or "checkpoint" in page.url:
                log("⚠️  Facebook non connecté — tentative login auto...")
                logged_in = login_facebook(page)
                if not logged_in:
                    print("")
                    print("=" * 60)
                    print("  Connecte-toi manuellement dans Chrome.")
                    print("  Quand tu es connecté → appuie sur ENTREE")
                    print("=" * 60)
                    input("")
                    if "login" in page.url:
                        log("❌ Toujours non connecté — onglet Facebook ignoré")
                        all_rows.pop("Facebook", None)
            else:
                log("✅ Facebook connecté")

        for sheet_name, rows in all_rows.items():
            handler_type = SHEET_HANDLERS.get(sheet_name, "html")
            ws           = wb[sheet_name]

            log(f"\n{'='*50}")
            log(f"Onglet : {sheet_name} ({len(rows)} URLs)")
            log(f"{'='*50}")

            for row in rows:
                log(f"\n  → {row['url'][:80]}")

                if handler_type == "instagram":
                    ok = handle_instagram(page, row, base_url, city_id,
                                          verify_ssl, dry_run, zone)
                else:
                    ok = handle_html(page, row, base_url, city_id,
                                     verify_ssl, dry_run)

                if not dry_run:
                    new_status = "done" if ok else "failed"
                    update_row(ws, row["row_idx"], new_status)
                    wb.save(curate_file)

                stats["done" if ok else "errors"] += 1
                sleep_random(2, 4)

        browser.close()

    log(f"\n=== Résultat ===")
    log(f"  Traités : {stats['done']}")
    log(f"  Erreurs : {stats['errors']}")


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from agent_excel_logger import run_logged
    run_logged("curate_manuel_events", "events", main)
