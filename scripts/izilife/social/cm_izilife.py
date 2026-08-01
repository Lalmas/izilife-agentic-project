"""
Agent Community Manager — izilife V12.1 mono-post final — Drive safe

Usage :
  python scripts/izilife/social/cm_izilife.py --zone=lille --env=local --init
  python ./scripts/izilife/social/cm_izilife.py --zone=lille --env=local --dry-run
  python ./scripts/izilife/social/cm_izilife.py --zone=lille --env=local

Principes V12.1 :
- Excel Planning = contenu + directives LLM uniquement.
- Onglet Post Types = configuration éditoriale/template uniquement.
- Aucun provider, modèle, engine ou clé API dans Excel.
- Config serveur unique par défaut : scripts/izilife/.env.izilife.
- Surcharge facultative : scripts/config/izilife/zones/<zone>/social.env.
- Aucun autre .env n'est requis par ce script.
- TEXT_PROVIDER/TEXT_MODEL pilotent le texte.
- IMAGE_ENGINE accepte none | gpt | llm | both | canva.
- IMAGE_PROVIDER/IMAGE_MODEL pilotent le fournisseur et le modèle image.
- VIDEO_ENGINE/VIDEO_PROVIDER/VIDEO_MODEL sont réservés à la vidéo future.
- Pillow est définitivement supprimé.
- Les templates peuvent contenir plusieurs exemples mono-post ou plusieurs exemples complets de carrousel.
- Le logo Izilife est global : izilife/assets/logo.png, surchargeable par BRAND_LOGO_PATH.
"""

from __future__ import annotations

import argparse
import random
import base64
import json
import mimetypes
import time
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import anthropic
import openpyxl
import requests
from dotenv import load_dotenv
from openpyxl.cell.cell import MergedCell


def _ensure_core_import_path() -> Path:
    here = Path(__file__).resolve()
    # Retourne scripts root si détecté.
    for parent in here.parents:
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return parent
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return parent / "scripts"
    # fallback : si script est scripts/izilife/social/cm_izilife.py
    try:
        return here.parents[2]
    except Exception:
        return Path.cwd()


SCRIPTS_ROOT = _ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, izilife_social_zone, normalize_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False
    PROJECT_ROOT = SCRIPTS_ROOT.parent

    def normalize_zone(zone: str) -> str:
        z = str(zone or "").strip().lower()
        return z if z.endswith("-zone") else f"{z}-zone"


CONTEXT_ROOT = PROJECT_ROOT / "izilife" / "context" / "social"
TEMPLATES_DIR = PROJECT_ROOT / "izilife" / "templates"
ASSETS_DIR = PROJECT_ROOT / "izilife" / "assets"
IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp")

ENV_IZILIFE_SCRIPTS = SCRIPTS_ROOT / "izilife" / ".env.izilife"
CONFIG_ZONE_ROOT = SCRIPTS_ROOT / "config" / "izilife" / "zones"

TYPE_CONTENT = {
    "POST_AGENDA_SEMAINE": "story",
    "POST_AGENDA_WEEKEND": "story",
    "POST_TOP_LIEUX": "carrousel",
    "POST_TOP_ACTEURS": "carrousel",
    "POST_HISTOIRE": "carrousel",
    "POST_HUMOUR": "post",
    "POST_ESCAPADE_VILLE": "carrousel",
    "POST_ESCAPADE_NATURE": "carrousel",
    "POST_PEPITE": "post",
    "EVENT_SERIE": "post",
    "EVENT_ANIMATEUR": "post",
    "PARTAGE_ACTEUR": "post",
    "NOUVEAUTE_IZILIFE": "post",
    "REEL": "reel",
    "VIDEO": "video",
}

TYPE_TEMPLATE = {
    "POST_AGENDA_SEMAINE": "templates/agenda_semaine/",
    "POST_AGENDA_WEEKEND": "templates/agenda_weekend/",
    "POST_TOP_LIEUX": "templates/top/",
    "POST_TOP_ACTEURS": "templates/top/",
    "POST_HISTOIRE": "templates/histoire_lieu/",
    "POST_HUMOUR": "templates/humour_local/",
    "POST_ESCAPADE_VILLE": "templates/escapade_ville/",
    "POST_ESCAPADE_NATURE": "templates/escapade_ville/",
    "POST_PEPITE": "templates/pepite_niche/",
    "EVENT_SERIE": "templates/event_serie/",
    "EVENT_ANIMATEUR": "templates/event_animateur/",
}

# Planning — colonnes contenu/directives.
COL_DATE = 1
COL_RESEAU = 2
COL_TYPE = 3
COL_VILLE = 4
COL_LIEU = 5
COL_SUJET = 6
COL_SERIE = 7
COL_ARTISTES = 8
COL_STYLE = 9
COL_DATA = 10
COL_DATE_EVENT = 11
COL_HEURE = 12
COL_INPUT_ID = 13
COL_IMAGE_PROMPT = 14
COL_POST_OUTPUT = 15
COL_IMAGE_OUTPUT = 16
COL_IMAGE_STATUS = 17
COL_STATUT = 18

STATUT_A_FAIRE = "À faire"
STATUT_RELANCER = "Relancer"
STATUT_GENERE = "Généré"

POST_TYPES_HEADERS = [
    "POST_TYPE",
    "CONTENT_TYPE",
    "TEMPLATE_LOCAL",
    "TEMPLATE_SOURCE",
    "TEMPLATE_REF",
    "TITLE_FONT",
    "TEXT_FONT",
    "NOTES",
]


def zone_key(zone: str) -> str:
    return normalize_zone(zone).replace("-zone", "")


def social_env_default() -> str:
    return os.getenv("SOCIAL_ENV", os.getenv("DEFAULT_ENV", "prod")).strip().lower() or "prod"


def zone_config(zone: str, env_name: str | None = None) -> dict[str, Any]:
    env_name = (env_name or social_env_default()).strip().lower()
    z = normalize_zone(zone)
    if HAS_CORE_PATHS:
        drive = izilife_social_zone(z, env_name)
    else:
        workspace = {"local": "agentic_workspace_local", "staging": "agentic_workspace_staging", "prod": "agentic_workspace"}.get(env_name, "agentic_workspace")
        drive = Path(os.getenv("AGENTIC_DRIVE_ROOT", "G:/Mon Drive")) / workspace / "izilife" / "social" / z
    return {
        "env": env_name,
        "zone": z,
        "zone_key": z.replace("-zone", ""),
        "drive": drive,
        "excel": drive / f"planning_{z}.xlsx",
        "inputs": drive / "inputs",
        "outputs": drive / "outputs",
        "context_zone": CONTEXT_ROOT / "zones" / f"{z}.md",
        "zone_env": CONFIG_ZONE_ROOT / z.replace("-zone", "") / "social.env",
    }


def ensure_directory(path: Path) -> None:
    """
    Crée un dossier de façon robuste, y compris sur Google Drive for Desktop.

    Sur un filesystem synchronisé, mkdir(parents=True, exist_ok=True) peut parfois
    voir le parent comme absent puis comme déjà créé entre deux appels. Cette
    fonction crée la hiérarchie niveau par niveau et tolère ce cas uniquement
    lorsque le chemin final est bien un dossier.
    """
    path = Path(path)

    if path.is_dir():
        return

    if path.exists():
        raise RuntimeError(f"Le chemin existe mais n'est pas un dossier : {path}")

    parent = path.parent
    if parent != path and not parent.is_dir():
        ensure_directory(parent)

    try:
        path.mkdir()
    except FileExistsError:
        if not path.is_dir():
            raise
    except FileNotFoundError:
        # Google Drive peut ne pas exposer immédiatement le parent créé.
        # On revérifie puis on retente une fois sans parents=True.
        if not parent.is_dir():
            ensure_directory(parent)
        try:
            path.mkdir()
        except FileExistsError:
            if not path.is_dir():
                raise


def ensure_default_env_files(cfg: dict[str, Any]) -> None:
    """Crée uniquement des exemples de configuration, sans écraser les vrais fichiers."""
    ENV_IZILIFE_SCRIPTS.parent.mkdir(parents=True, exist_ok=True)
    if not ENV_IZILIFE_SCRIPTS.exists():
        example = ENV_IZILIFE_SCRIPTS.with_name(".env.izilife.example")
        if not example.exists():
            example.write_text(
                "# API\n"
                "OPENAI_API_KEY=\n"
                "ANTHROPIC_API_KEY=\n\n"
                "# Texte\n"
                "TEXT_PROVIDER=anthropic\n"
                "TEXT_MODEL=claude-sonnet-4-6\n\n"
                "# Images\n"
                "IMAGE_ENGINE=gpt\n"
                "IMAGE_PROVIDER=openai\n"
                "IMAGE_MODEL=gpt-image-1\n"
                "IMAGE_SIZE_POST=1024x1024\n"
                "IMAGE_SIZE_STORY=1024x1536\n\n"
                "# Canva\n"
                "CANVA_ENABLED=0\n\n"
                "# Vidéo\n"
                "VIDEO_ENGINE=none\n"
                "VIDEO_PROVIDER=\n"
                "VIDEO_MODEL=\n\n"
                "# Branding\n"
                "BRAND_LOGO_ENABLED=1\n"
                "BRAND_LOGO_PATH=../../izilife/assets/logo.png\n",
                encoding="utf-8",
            )
            print(f"ℹ️ Crée et renseigne : {ENV_IZILIFE_SCRIPTS} (exemple : {example})")

    cfg["zone_env"].parent.mkdir(parents=True, exist_ok=True)
    if not cfg["zone_env"].exists():
        cfg["zone_env"].write_text(
            f"# Override social zone {cfg['zone']} — facultatif et non sensible\n"
            "# Décommente uniquement les valeurs qui diffèrent des defaults Izilife.\n"
            "# TEXT_PROVIDER=anthropic\n"
            "# TEXT_MODEL=claude-sonnet-4-6\n"
            "# IMAGE_ENGINE=gpt\n"
            "# IMAGE_PROVIDER=openai\n"
            "# IMAGE_MODEL=gpt-image-1\n"
            "# BRAND_LOGO_ENABLED=1\n"
            "# VIDEO_ENGINE=none\n",
            encoding="utf-8",
        )
        print(f"✅ Config zone créée : {cfg['zone_env']}")

def load_env_files(cfg: dict[str, Any]) -> None:
    """Charge defaults Izilife puis surcharge zone."""
    if not ENV_IZILIFE_SCRIPTS.exists():
        raise RuntimeError(f"Configuration manquante : {ENV_IZILIFE_SCRIPTS}")
    load_dotenv(ENV_IZILIFE_SCRIPTS, override=True)
    print(f"   config chargée : {ENV_IZILIFE_SCRIPTS}")
    if cfg["zone_env"].exists():
        load_dotenv(cfg["zone_env"], override=True)
        print(f"   override zone chargé : {cfg['zone_env']}")

def slugify(t: str) -> str:
    t = str(t or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9_]", "", t.lower().replace(" ", "_").replace("-", "_"))[:45] or "post"


def _safe_cell(ws, r: int, c: int):
    x = ws.cell(r, c)
    return None if isinstance(x, MergedCell) else x


def ensure_planning(path: Path) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    ws.title = "Planning"

    # Ne jamais supprimer les données. On ne fait que restyler/ajouter colonnes manquantes.
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(sheet, r, c, v, bg="1A1A2E", sz=9):
        x = _safe_cell(sheet, r, c)
        if x is None:
            return
        x.value = v
        x.font = Font(name="Arial", size=sz, bold=True, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = brd

    groups = [
        (1, 3, "PLANIFICATION", "1A1A2E"),
        (4, 10, "CONTENU / DONNÉES LLM", "0F3460"),
        (11, 12, "EVENT", "1B4332"),
        (13, 14, "INPUT / IMAGE", "0D6986"),
        (15, 17, "OUTPUT", "1A5276"),
        (18, 18, "STATUT", "6C3483"),
    ]
    for s, e, label, bg in groups:
        if s < e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        hdr(ws, 1, s, label, bg=bg, sz=10)

    headers = [
        ("DATE", 12),
        ("RÉSEAU", 12),
        ("TYPE", 24),
        ("VILLE", 14),
        ("LIEU", 18),
        ("THÈME / SUJET", 30),
        ("SÉRIE", 18),
        ("ARTISTE / DJ / ASSO", 30),
        ("STYLE", 40),
        ("DATA", 60),
        ("DATE EVENT", 14),
        ("HEURE / DURÉE", 16),
        ("INPUT_ID", 18),
        ("IMAGE_PROMPT", 48),
        ("POST OUTPUT", 70),
        ("IMAGE_OUTPUT", 40),
        ("IMAGE_STATUS", 20),
        ("STATUT", 14),
    ]
    for i, (label, width) in enumerate(headers, 1):
        hdr(ws, 2, i, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 42

    validations = [
        (DataValidation(type="list", formula1='"Instagram,Facebook,Les deux"', allow_blank=True), COL_RESEAU),
        (DataValidation(type="list", formula1='"' + ",".join(TYPE_CONTENT.keys()) + '"', allow_blank=True), COL_TYPE),
        (DataValidation(type="list", formula1='"non demandé,généré,généré gpt,généré canva,généré gpt+canva,erreur,référence trouvée,canva demande créée"', allow_blank=True), COL_IMAGE_STATUS),
        (DataValidation(type="list", formula1='"À faire,Généré,Relancer,Validé,Publié,Skip,Erreur"', allow_blank=True), COL_STATUT),
    ]
    for dv, col in validations:
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}500")

    # Post Types = éditorial/template seulement. Pas de providers, pas de modèles, pas de moteurs.
    wt = wb["Post Types"] if "Post Types" in wb.sheetnames else wb.create_sheet("Post Types")
    existing_headers = [str(c.value or "").strip() for c in wt[1]] if wt.max_row else []
    for h in POST_TYPES_HEADERS:
        if h not in existing_headers:
            wt.cell(row=1, column=len(existing_headers) + 1, value=h)
            existing_headers.append(h)
    header_map = {h: i + 1 for i, h in enumerate(existing_headers)}

    for i, h in enumerate(existing_headers, 1):
        hdr(wt, 1, i, h, "1A1A2E")
        wt.column_dimensions[get_column_letter(i)].width = {
            "POST_TYPE": 28,
            "CONTENT_TYPE": 14,
            "TEMPLATE_LOCAL": 30,
            "TEMPLATE_SOURCE": 18,
            "TEMPLATE_REF": 42,
            "TITLE_FONT": 28,
            "TEXT_FONT": 28,
            "NOTES": 80,
        }.get(h, 20)

    existing_post_types = set()
    for row in wt.iter_rows(min_row=2):
        val = row[header_map["POST_TYPE"] - 1].value if "POST_TYPE" in header_map else None
        if val:
            existing_post_types.add(str(val).strip())

    for pt, ct in TYPE_CONTENT.items():
        if pt in existing_post_types:
            continue
        r = wt.max_row + 1
        values = {
            "POST_TYPE": pt,
            "CONTENT_TYPE": ct,
            "TEMPLATE_LOCAL": TYPE_TEMPLATE.get(pt, ""),
            "TEMPLATE_SOURCE": "owned" if pt in TYPE_TEMPLATE else "none",
            "TEMPLATE_REF": "",
            "TITLE_FONT": "",
            "TEXT_FONT": "",
            "NOTES": "owned=template interne utilisable comme référence directe ; reference=image tierce envoyée comme inspiration sans copie ; text=inspiration texte seulement ; none=libre. TITLE_FONT/TEXT_FONT sont des préférences typographiques données au LLM. Providers/moteurs/config IA côté serveur uniquement.",
        }
        for h, v in values.items():
            wt.cell(row=r, column=header_map[h], value=v)

    if "TEMPLATE_SOURCE" in header_map:
        dv = DataValidation(type="list", formula1='"owned,reference,text,none"', allow_blank=True)
        wt.add_data_validation(dv)
        dv.add(f"{get_column_letter(header_map['TEMPLATE_SOURCE'])}2:{get_column_letter(header_map['TEMPLATE_SOURCE'])}200")

    # Mode d'emploi clair.
    if "Mode d'emploi" in wb.sheetnames:
        del wb["Mode d'emploi"]
    wi = wb.create_sheet("Mode d'emploi", 0)
    lines = [
        "CM izilife — mode d'emploi",
        "Planning = contenu à produire. Tu peux tout piloter depuis Excel si tu veux.",
        "STYLE = directives LLM : ton, angle, contrainte, prompt direct.",
        "DATA = infos factuelles à injecter : punchlines, lieux, artistes, prix, horaires, liens, conditions.",
        "ARTISTE / DJ / ASSO = liste imposée par le fondateur ; l'agent ne l'invente jamais.",
        "INPUT_ID = dossier inputs/[INPUT_ID]/ avec description.txt et/ou images. Complément ou fallback de l'Excel.",
        "IMAGE_PROMPT = directive ponctuelle sur le visuel ou le rôle des images INPUT_ID. Optionnel.",
        "Post Types = configuration éditoriale/template une seule fois par TYPE.",
        "Aucun provider, modèle, engine ou clé API dans Excel. Tout est côté scripts/.env, scripts/izilife/.env.izilife, scripts/config/izilife/zones/<zone>/social.env.",
        "TEMPLATE_SOURCE : owned = template interne envoyé au LLM comme référence de composition ; reference = image tierce envoyée comme inspiration sans copie ; text = inspiration texte ; none = libre.",
        "TITLE_FONT / TEXT_FONT = polices préférées pour le titre et le texte visuel. Ce sont des directives au LLM, pas des polices injectées techniquement.",
        "Le script ne touche jamais Validé / Publié / Skip. Pour relancer : STATUT = Relancer.",
    ]
    for r, line in enumerate(lines, 1):
        wi.cell(r, 1, line)
        wi.column_dimensions["A"].width = 145
        if r == 1:
            wi.cell(r, 1).font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            wi.cell(r, 1).fill = PatternFill("solid", fgColor="1A1A2E")

    ensure_directory(path.parent)
    wb.save(path)
    print(f"✅ Planning CM izilife prêt : {path}")


def load_post_type_config(wb, post_type: str) -> dict[str, Any]:
    if "Post Types" not in wb.sheetnames:
        return {}
    ws = wb["Post Types"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if str(d.get("POST_TYPE") or "").strip() == post_type:
            return d
    return {}



def ensure_description_template(inputs_dir: Path) -> Path:
    """Crée un squelette description_template.txt sans écraser l'existant."""
    target = inputs_dir / "description_template.txt"
    if target.exists():
        return target
    target.write_text(
        """# DESCRIPTION DU POST
# Décris ici le sujet, l'intention ou le contexte général.


# DIRECTIVES GLOBALES
# Ton, angle, contraintes communes au texte et au visuel.


# DIRECTIVES IMAGE
# Exemples : utiliser photo.jpg comme fond principal ; conserver la personne ;
# utiliser l'image uniquement comme inspiration d'ambiance ; changer complètement le fond.


# TEXTE VISUEL IMPOSÉ
# Laisser vide si le LLM peut composer/reformuler le texte du visuel.
# Si rempli, ce texte doit être reproduit strictement, avec orthographe et casse correctes.


# TEXTE DU POST IMPOSÉ
# Caption Instagram/Facebook. Laisser vide pour génération par le LLM texte.

""",
        encoding="utf-8",
    )
    return target

def init_zone(zone: str, env_name: str) -> None:
    cfg = zone_config(zone, env_name)

    ensure_default_env_files(cfg)
    load_env_files(cfg)

    print(f"📁 Workspace social : {cfg['drive']}")

    ensure_directory(cfg["drive"])
    ensure_directory(cfg["inputs"])
    ensure_directory(cfg["outputs"])

    template_path = ensure_description_template(cfg["inputs"])
    print(f"✅ Squelette description : {template_path}")

    ensure_directory(cfg["context_zone"].parent)
    if not cfg["context_zone"].exists():
        cfg["context_zone"].write_text(
            f"""# Contexte zone — {cfg['zone']}
## Spécificités
- Ville principale :
- Hashtags locaux : #{cfg['zone'].replace('-zone','')}
- Compte Instagram : @izilife_{cfg['zone'].replace('-zone','')}
- Événements locaux :
- Lieux emblématiques :
""",
            encoding="utf-8",
        )

    ensure_planning(cfg["excel"])
    print(f"✅ Init CM terminé : {cfg['zone']} / {cfg['env']}")


def load_context(zone: str) -> str:
    parts: list[str] = []
    for fn in ["izilife-social-strategy.md", "community-manager.md", "izilife-style.md"]:
        p = CONTEXT_ROOT / fn
        if p.exists():
            parts.append(f"### {fn}\n{p.read_text(encoding='utf-8')}")
    zp = CONTEXT_ROOT / "zones" / f"{zone}.md"
    if zp.exists():
        parts.append(f"### {zone}.md\n{zp.read_text(encoding='utf-8')}")
    return "\n\n---\n\n".join(parts)



LEGACY_IMAGE_PROMPT_TOKENS = {"owned", "reference", "text", "none", "inspiration"}


def clean_image_prompt(value: str) -> str:
    """Ignore les anciennes valeurs de configuration restées dans IMAGE_PROMPT."""
    raw = str(value or "").strip()
    return "" if raw.lower() in LEGACY_IMAGE_PROMPT_TOKENS else raw


def extract_forced_visual_text(*sources: str) -> str:
    """Extrait un bloc TEXTE VISUEL IMPOSÉ / TEXTE IMAGE IMPOSÉ / TEXTE À AFFICHER."""
    marker_re = re.compile(
        r"(?is)(?:TEXTE\s+VISUEL\s+IMPOS[ÉE]|TEXTE\s+IMAGE\s+IMPOS[ÉE]|TEXTE\s+[ÀA]\s+AFFICHER)\s*:\s*(.+)"
    )
    stop_re = re.compile(r"(?m)^\s*[A-ZÀ-ÖØ-Ý0-9 _/\-]{4,}\s*:\s*$")
    for source in sources:
        text = str(source or "").strip()
        if not text:
            continue
        match = marker_re.search(text)
        if not match:
            continue
        block = match.group(1).strip()
        stop = stop_re.search(block)
        if stop:
            block = block[:stop.start()].strip()
        if block:
            return block
    return ""

def read_input(input_id: str, inputs_dir: Path):
    if not input_id:
        return "", []
    folder = inputs_dir / input_id
    if not folder.exists():
        return "", []
    desc = (folder / "description.txt").read_text(encoding="utf-8").strip() if (folder / "description.txt").exists() else ""
    imgs = sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")])
    return desc, imgs


def build_prompt(row, content_type: str, input_desc: str, post_cfg: dict[str, Any], expected_slides: int = 0) -> str:
    def v(c: int) -> str:
        val = row[c - 1].value
        return str(val).strip() if val else ""

    if content_type == "carrousel":
        count_rule = expected_slides if expected_slides > 0 else 5
        slides = f"""
=== SLIDES ===
Tu dois produire exactement {count_rule} blocs :
Slide 1 : ...
Slide 2 : ...
...
Slide {count_rule} : ...
Chaque slide doit avoir un texte visuel court, grammaticalement correct, avec accents et ponctuation.
"""
    else:
        slides = ""

    return f"""Tu es l'agent Community Manager izilife. CONTENT_TYPE: {content_type}

CONFIG ÉDITORIALE DU TYPE :
TEMPLATE_SOURCE: {post_cfg.get('TEMPLATE_SOURCE','none')}
TEMPLATE_REF: {post_cfg.get('TEMPLATE_REF','')}
TEMPLATE_LOCAL: {post_cfg.get('TEMPLATE_LOCAL','')}

DONNÉES EXCEL — prioritaires sur description.txt :
TYPE: {v(COL_TYPE)}
RÉSEAU: {v(COL_RESEAU)}
VILLE: {v(COL_VILLE)}
LIEU: {v(COL_LIEU)}
THÈME / SUJET: {v(COL_SUJET)}
SÉRIE: {v(COL_SERIE)}
ARTISTE / DJ / ASSO: {v(COL_ARTISTES)}
STYLE / DIRECTIVES LLM: {v(COL_STYLE)}
DATA / INFOS À INJECTER: {v(COL_DATA)}
DATE EVENT: {v(COL_DATE_EVENT)}
HEURE / DURÉE: {v(COL_HEURE)}
IMAGE_PROMPT: {v(COL_IMAGE_PROMPT)}

DESCRIPTION INPUT — complément ou fallback :
{input_desc}

Règles :
- Les colonnes Excel sont prioritaires. description.txt complète ce qui manque et peut aussi contenir des directives explicites.
- Le script doit fonctionner avec Excel seul, description.txt seul, ou un mélange des deux.
- STYLE peut contenir un prompt direct, un ton, un angle, une contrainte visuelle ou éditoriale.
- Si description.txt contient « TEXTE DU POST IMPOSÉ: » ou « CAPTION IMPOSÉE: », reprendre ce texte tel quel comme caption et ne pas le réécrire.
- DATA contient les faits à injecter : punchlines, artistes, DJ, assos, lieux, prix, horaires, liens, infos pratiques.
- Ne jamais inventer un lieu, un acteur, un artiste, un DJ, une asso ou un événement non fourni.
- Pour les tops et events, la liste ARTISTE / DJ / ASSO est imposée.
- Orthographe française irréprochable : accents, apostrophes, majuscules et ponctuation corrects.
- 1 seul CTA vers izilife.
- Ton local, direct, pas corporate.

Format OBLIGATOIRE :
=== TEXTE DU POST ===
[caption complète avec emojis + hashtags]{slides}
"""



def extract_forced_caption(input_desc: str) -> str:
    marker_re = re.compile(r"(?is)(?:TEXTE\s+DU\s+POST\s+IMPOS[ÉE]|CAPTION\s+IMPOS[ÉE]E?)\s*:\s*(.+)")
    match = marker_re.search(str(input_desc or ""))
    if not match:
        return ""
    block = match.group(1).strip()
    stop = re.search(r"(?m)^\s*[A-ZÀ-ÖØ-Ý0-9 _/\-]{4,}\s*:\s*$", block)
    return (block[:stop.start()] if stop else block).strip()

def get_text_provider() -> str:
    return str(os.getenv("TEXT_PROVIDER", "anthropic")).strip().lower()


def get_text_model(provider: str) -> str:
    if provider in ("openai", "gpt"):
        return os.getenv("TEXT_MODEL", "gpt-4o")
    return os.getenv("TEXT_MODEL", "claude-sonnet-4-6")


def call_llm(system: str, user: str) -> str:
    provider = get_text_provider()
    if provider in ("openai", "gpt"):
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r = client.chat.completions.create(
            model=get_text_model(provider),
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            max_tokens=1500,
        )
        return r.choices[0].message.content or ""

    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    msg = client.messages.create(
        model=get_text_model(provider),
        max_tokens=1500,
        system=system,
        messages=[{"role": "user", "content": user}],
    )
    return msg.content[0].text


def _image_files(directory: Path, recursive: bool = False) -> list[Path]:
    if not directory.exists() or not directory.is_dir():
        return []
    iterator = directory.rglob("*") if recursive else directory.iterdir()
    return sorted([p for p in iterator if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS], key=lambda p: p.name.lower())


def _resolve_local_reference(ref: str) -> Path | None:
    ref = str(ref or "").strip()
    if not ref or ref.lower().startswith(("http://", "https://", "canva:", "brandtemplate:")):
        return None
    candidates = [Path(ref), TEMPLATES_DIR / ref]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def resolve_template_root(post_cfg: dict[str, Any]) -> Path | None:
    template_local = str(post_cfg.get("TEMPLATE_LOCAL") or "").strip()
    if not template_local:
        return None
    rel = template_local.replace("templates/", "").replace("templates\\", "")
    root = TEMPLATES_DIR / rel
    return root if root.exists() else None


def select_template_bundle(post_cfg: dict[str, Any], content_type: str) -> dict[str, Any]:
    """Sélectionne un exemple de template selon la convention simple.

    Mono-post :
      - fichiers directement dans le dossier = exemples alternatifs ; un seul est choisi ;
      - sous-dossiers = exemples alternatifs ; tous les fichiers du sous-dossier choisi
        sont envoyés ensemble comme références complémentaires.

    Carrousel :
      - chaque sous-dossier = un exemple complet de carrousel ;
      - chaque fichier du sous-dossier choisi = référence d'une slide ;
      - sans sous-dossier, les fichiers directs constituent un seul exemple complet.

    TEMPLATE_REF peut nommer explicitement un fichier ou sous-dossier. Sinon choix aléatoire.
    """
    root = resolve_template_root(post_cfg)
    if not root:
        return {"root": None, "selected": None, "references": [], "slide_references": []}

    direct = _image_files(root)
    subdirs = sorted([d for d in root.iterdir() if d.is_dir() and _image_files(d)], key=lambda p: p.name.lower())
    requested = str(post_cfg.get("TEMPLATE_REF") or "").strip()
    selected_path = None

    if requested and not requested.lower().startswith(("http://", "https://", "canva:", "brandtemplate:")):
        candidate = root / requested
        if candidate.exists():
            selected_path = candidate
        else:
            external = _resolve_local_reference(requested)
            if external:
                selected_path = external

    rng = random.SystemRandom()
    ct = str(content_type or "post").lower()

    if ct == "carrousel":
        if selected_path and selected_path.is_dir():
            slides = _image_files(selected_path)
            selected = selected_path
        elif selected_path and selected_path.is_file():
            slides = [selected_path]
            selected = selected_path
        elif subdirs:
            selected = rng.choice(subdirs)
            slides = _image_files(selected)
        else:
            selected = root
            slides = direct
        return {"root": root, "selected": selected, "references": [], "slide_references": slides}

    # Mono-post
    if selected_path and selected_path.is_dir():
        refs = _image_files(selected_path)
        selected = selected_path
    elif selected_path and selected_path.is_file():
        refs = [selected_path]
        selected = selected_path
    elif subdirs:
        selected = rng.choice(subdirs)
        refs = _image_files(selected)
    elif direct:
        selected = rng.choice(direct)
        refs = [selected]
    else:
        selected = root
        refs = []
    return {"root": root, "selected": selected, "references": refs, "slide_references": []}


def brand_logo_path() -> Path | None:
    enabled = str(os.getenv("BRAND_LOGO_ENABLED", "1")).strip().lower()
    if enabled in ("0", "false", "no", "off"):
        return None
    configured = str(os.getenv("BRAND_LOGO_PATH", "")).strip()
    candidates = []
    if configured:
        configured_path = Path(configured)
        if configured_path.is_absolute():
            candidates.append(configured_path)
        else:
            # Les chemins relatifs du .env sont résolus depuis scripts/izilife/.
            candidates.append((ENV_IZILIFE_SCRIPTS.parent / configured_path).resolve())
            candidates.append((PROJECT_ROOT / configured_path).resolve())
    candidates += [ASSETS_DIR / "logo.png", ASSETS_DIR / "logo.webp", ASSETS_DIR / "logo.jpg"]
    for p in candidates:
        if p.exists() and p.is_file():
            return p
    return None


def materialize_url_reference(url: str, folder: Path) -> Path | None:
    try:
        ext = Path(url.split("?", 1)[0]).suffix.lower()
        if ext not in IMAGE_EXTENSIONS:
            ext = ".png"
        out = folder / f"external_reference{ext}"
        r = requests.get(url, timeout=60, headers={"User-Agent": "Izilife CM Agent/1.0"})
        r.raise_for_status()
        out.write_bytes(r.content)
        return out
    except Exception as exc:
        print(f"    ⚠️ référence URL non récupérée : {exc}")
        return None


def target_canvas_size_str(content_type: str) -> str:
    ct = str(content_type or "post").lower()
    if ct in ("story", "reel", "video"):
        return os.getenv("IMAGE_SIZE_STORY", os.getenv("IMAGE_SIZE", "1024x1536"))
    return os.getenv("IMAGE_SIZE_POST", os.getenv("IMAGE_SIZE", "1024x1024"))


def _row_value(row, column: int) -> str:
    value = row[column - 1].value
    return str(value).strip() if value not in (None, "") else ""


def build_image_prompt(
    row,
    post_cfg: dict[str, Any],
    content_type: str,
    input_desc: str = "",
    slide_text: str = "",
    slide_index: int | None = None,
    slide_count: int | None = None,
    reference_count: int = 0,
    has_logo: bool = False,
) -> str:
    template_source = str(post_cfg.get("TEMPLATE_SOURCE") or "none").strip().lower()
    image_prompt = clean_image_prompt(_row_value(row, COL_IMAGE_PROMPT))
    style = _row_value(row, COL_STYLE)
    data = _row_value(row, COL_DATA)
    forced_visual_text = extract_forced_visual_text(data, image_prompt, input_desc)
    title_font = str(post_cfg.get("TITLE_FONT") or "").strip()
    text_font = str(post_cfg.get("TEXT_FONT") or "").strip()

    refs = []
    if reference_count:
        refs.append(
            f"Les images 1 à {reference_count} sont des références visuelles. "
            "Inspire-toi de leur composition, rythme, hiérarchie et identité, sans recopier leur texte ni leur photo de fond."
        )
    if has_logo:
        refs.append(
            f"L'image {reference_count + 1} est le logo officiel Izilife. "
            "Utilise exactement ce logo une seule fois, petit, net, sans le redessiner ni le modifier."
        )

    carousel_rule = ""
    if slide_index is not None and slide_count:
        carousel_rule = f"Cette image est la slide {slide_index}/{slide_count} d'un carrousel. Elle doit rester cohérente avec les autres slides."

    return f"""Tu es un directeur artistique senior spécialisé en posts Instagram et Facebook pour Izilife.

OBJECTIF
Créer un visuel final publiable, lisible et original. Il ne s'agit pas d'une capture de texte ni d'une simple affiche automatique.

RÈGLES ABSOLUES
- Orthographe française parfaite : accents, apostrophes, majuscules et ponctuation corrects.
- Toute phrase, tout item de liste et tout libellé commence par une majuscule quand la grammaire française l’exige.
- Ne produis jamais une liste entière en minuscules par défaut.
- Si un TEXTE VISUEL IMPOSÉ est fourni plus bas, reproduis-le strictement, sans reformulation.
- Sinon, corrige la casse, les accents et la ponctuation du texte visuel sans en changer le sens.
- Aucun texte tronqué, hors cadre, trop petit ou caché.
- Hiérarchie nette, marges de sécurité, contraste suffisant.
- Ne génère pas la caption du post dans l'image.
- N'ajoute pas de CTA du type « réponds en commentaire » sauf demande explicite dans STYLE, DATA ou IMAGE_PROMPT.
- Ne crée jamais un faux logo.
- Si une référence contient une photo, ne la réutilise pas automatiquement : choisis une nouvelle image de fond cohérente sauf instruction contraire.
- TEMPLATE_SOURCE={template_source}. Même si la référence est owned, on s'en inspire : on ne recopie pas servilement.
- Police préférée pour le titre : {title_font or "libre, cohérente avec la référence"}.
- Police préférée pour le texte : {text_font or "libre, cohérente avec la référence"}.
- Si une police nommée précisément n'est pas disponible, reproduis au plus près son style typographique sans inventer son nom dans l'image.

{carousel_rule}
{' '.join(refs)}

SUJET
{_row_value(row, COL_SUJET)}

VILLE / LIEU
{_row_value(row, COL_VILLE)} / {_row_value(row, COL_LIEU)}

STYLE / DIRECTIVES PRIORITAIRES
{style}

DATA / INFORMATIONS À INTÉGRER
{data}

DIRECTIVE IMAGE SPÉCIFIQUE
{image_prompt}

DESCRIPTION.TXT — directives et complément éventuel
{input_desc}

TEXTE VISUEL IMPOSÉ — vide = tu peux composer/corriger la forme
{forced_visual_text}

TEXTE DE CETTE SLIDE / CE VISUEL
{forced_visual_text or slide_text or data or _row_value(row, COL_SUJET)}

RÔLE DES IMAGES INPUT
- Les images du dossier INPUT_ID sont des références de travail.
- IMAGE_PROMPT ou DESCRIPTION.TXT peut préciser : fond principal, sujet à conserver, produit/personne à intégrer, ou simple inspiration d’ambiance.
- Sans directive explicite, utilise-les comme références visuelles complémentaires et ne les recopie pas servilement.

FORMAT
{content_type}. Taille demandée : {target_canvas_size_str(content_type)}.
""".strip()


def _decode_or_download_image(data, out: Path) -> Path | None:
    if getattr(data, "b64_json", None):
        out.write_bytes(base64.b64decode(data.b64_json))
        return out
    if getattr(data, "url", None):
        out.write_bytes(requests.get(data.url, timeout=90).content)
        return out
    return None


def generate_image_gpt(prompt: str, folder: Path, filename: str, content_type: str = "post") -> Path | None:
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    result = client.images.generate(
        model=os.getenv("IMAGE_MODEL", "gpt-image-1"),
        prompt=prompt,
        n=1,
        size=target_canvas_size_str(content_type),
        quality=os.getenv("IMAGE_QUALITY", "high"),
    )
    return _decode_or_download_image(result.data[0], folder / f"{filename}.png")


def generate_image_gpt_with_references(
    prompt: str,
    references: list[Path],
    folder: Path,
    filename: str,
    content_type: str = "post",
) -> Path | None:
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    refs = [p for p in references if p and p.exists()][:16]
    if not refs:
        return generate_image_gpt(prompt, folder, filename, content_type)

    handles = [open(p, "rb") for p in refs]
    try:
        image_arg = handles[0] if len(handles) == 1 else handles
        result = client.images.edit(
            model=os.getenv("IMAGE_MODEL", "gpt-image-1"),
            image=image_arg,
            prompt=prompt,
            n=1,
            size=target_canvas_size_str(content_type),
            quality=os.getenv("IMAGE_QUALITY", "high"),
            input_fidelity=os.getenv("IMAGE_INPUT_FIDELITY", "high"),
        )
    finally:
        for handle in handles:
            handle.close()
    return _decode_or_download_image(result.data[0], folder / f"{filename}.png")


def parse_slide_texts(text: str, expected_count: int, row) -> list[str]:
    matches = re.findall(r"(?im)^\s*slide\s+(\d+)\s*[:\-–]\s*(.+?)(?=^\s*slide\s+\d+\s*[:\-–]|\Z)", text, flags=re.S | re.M)
    parsed = [content.strip() for _, content in sorted(matches, key=lambda x: int(x[0]))]
    if not parsed:
        raw = _row_value(row, COL_DATA)
        items = [x.strip(" -•\t") for x in re.split(r"[\n;]+", raw) if x.strip(" -•\t")]
        title = _row_value(row, COL_SUJET)
        if expected_count <= 1:
            return [raw or title]
        parsed = [title] + items
    while len(parsed) < expected_count:
        parsed.append("")
    return parsed[:expected_count]


def process_images(
    post_cfg: dict[str, Any],
    row,
    text: str,
    input_desc: str,
    input_images: list[Path],
    template_bundle: dict[str, Any],
    folder: Path,
    filename: str,
    content_type: str = "post",
) -> tuple[list[Path], str]:
    engine = str(os.getenv("IMAGE_ENGINE", "gpt")).strip().lower()
    provider = str(os.getenv("IMAGE_PROVIDER", "openai")).strip().lower()
    if engine == "none":
        return [], "non demandé"
    if engine == "canva":
        raise RuntimeError("IMAGE_ENGINE=canva demandé, mais Canva n'est pas encore branché.")
    if engine == "both":
        print("    ℹ️ IMAGE_ENGINE=both : génération GPT active ; Canva sera ajouté quand CANVA_ENABLED=1 sera réellement branché.")
    elif engine not in ("gpt", "llm"):
        raise RuntimeError(f"IMAGE_ENGINE inconnu : {engine}. Utilise none|gpt|llm|both|canva.")
    if provider not in ("openai", "gpt"):
        raise RuntimeError(f"IMAGE_PROVIDER={provider} non disponible dans cette version. Utilise openai.")

    logo = brand_logo_path()
    generated: list[Path] = []
    ct = str(content_type or "post").lower()

    if ct == "carrousel":
        slide_refs = list(template_bundle.get("slide_references") or [])
        expected_count = len(slide_refs)
        if expected_count == 0:
            expected_count = max(2, int(os.getenv("DEFAULT_CAROUSEL_SLIDES", "5")))
            slide_refs = [None] * expected_count
        slide_texts = parse_slide_texts(text, expected_count, row)
        for index in range(expected_count):
            refs: list[Path] = []
            if slide_refs[index]:
                refs.append(slide_refs[index])
            refs.extend(input_images[:4])
            reference_count = len(refs)
            if logo:
                refs.append(logo)
            prompt = build_image_prompt(
                row,
                post_cfg,
                content_type,
                input_desc=input_desc,
                slide_text=slide_texts[index],
                slide_index=index + 1,
                slide_count=expected_count,
                reference_count=reference_count,
                has_logo=bool(logo),
            )
            out = generate_image_gpt_with_references(
                prompt,
                refs,
                folder,
                f"{filename}_slide_{index + 1:02d}",
                content_type,
            )
            if out:
                generated.append(out)
                print(f"    🖼️ slide {index + 1}/{expected_count} → {out.name}")
    else:
        refs = list(template_bundle.get("references") or [])
        ref_url = str(post_cfg.get("TEMPLATE_REF") or "").strip()
        if ref_url.lower().startswith(("http://", "https://")):
            downloaded = materialize_url_reference(ref_url, folder)
            if downloaded:
                refs.append(downloaded)
        refs.extend(input_images[:8])
        reference_count = len(refs)
        if logo:
            refs.append(logo)
        prompt = build_image_prompt(
            row,
            post_cfg,
            content_type,
            input_desc=input_desc,
            slide_text="",
            reference_count=reference_count,
            has_logo=bool(logo),
        )
        out = generate_image_gpt_with_references(prompt, refs, folder, f"{filename}_gpt", content_type)
        if out:
            generated.append(out)
            print(f"    🖼️ visuel → {out.name}")

    return generated, "généré gpt" if generated else "erreur"


def save_post(text: str, generated: list[Path], folder: Path, content_type: str):
    lines = [
        text.strip(),
        "",
        f"=== CONTENT_TYPE ===\n{content_type}",
        "",
        "=== CONFIG_RUNTIME ===",
        f"TEXT_PROVIDER={get_text_provider()}",
        f"TEXT_MODEL={get_text_model(get_text_provider())}",
        f"IMAGE_ENGINE={os.getenv('IMAGE_ENGINE', 'gpt')}",
        f"IMAGE_PROVIDER={os.getenv('IMAGE_PROVIDER', 'openai')}",
        f"IMAGE_MODEL={os.getenv('IMAGE_MODEL', 'gpt-image-1')}",
        f"VIDEO_ENGINE={os.getenv('VIDEO_ENGINE', 'none')}",
    ]
    if generated:
        lines += ["", "=== IMAGES GENEREES ==="] + [f"  {p.name}" for p in generated if p]
    (folder / "post.txt").write_text("\n".join(lines), encoding="utf-8")


def run_zone(zone: str, env_name: str, dry_run=False) -> None:
    cfg = zone_config(zone, env_name)
    load_env_files(cfg)
    if not cfg["excel"].exists():
        print(f"❌ Planning introuvable. Lance --init : {cfg['excel']}")
        return

    system = load_context(cfg["zone"])
    wb = openpyxl.load_workbook(cfg["excel"])
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    rows = [r for r in ws.iter_rows(min_row=3) if str(r[COL_STATUT - 1].value or "").strip() in (STATUT_A_FAIRE, STATUT_RELANCER)]
    print(f"🤖 CM izilife V12.1 — {cfg['zone']} — env={env_name} — {len(rows)} ligne(s)" + (" [DRY RUN]" if dry_run else ""))
    print(f"   TEXT={get_text_provider()}:{get_text_model(get_text_provider())} | IMAGE={os.getenv('IMAGE_ENGINE','gpt')}:{os.getenv('IMAGE_PROVIDER','openai')}:{os.getenv('IMAGE_MODEL','gpt-image-1')} | VIDEO={os.getenv('VIDEO_ENGINE','none')}")
    ensure_directory(cfg["outputs"])
    ensure_directory(cfg["inputs"])

    for idx, row in enumerate(rows, 1):
        def v(c: int) -> str:
            val = row[c - 1].value
            return str(val).strip() if val else ""

        post_type = v(COL_TYPE)
        subject = v(COL_SUJET)
        city = v(COL_VILLE)
        post_cfg = load_post_type_config(wb, post_type)
        content_type = str(post_cfg.get("CONTENT_TYPE") or TYPE_CONTENT.get(post_type, "post")).strip().lower()
        template_bundle = select_template_bundle(post_cfg, content_type)
        expected_slides = len(template_bundle.get("slide_references") or []) if content_type == "carrousel" else 0
        print(f"  [{idx}/{len(rows)}] {post_type} [{content_type}] — {city} — {subject}")
        input_desc, input_images = read_input(v(COL_INPUT_ID), cfg["inputs"])
        raw_image_prompt = v(COL_IMAGE_PROMPT)
        if raw_image_prompt.lower() in LEGACY_IMAGE_PROMPT_TOKENS:
            print(f"    ⚠️ IMAGE_PROMPT={raw_image_prompt!r} ignoré : ancienne valeur de config. Mets ici une directive ponctuelle ou laisse vide.")

        if dry_run:
            print(f"    TEMPLATE_SOURCE={post_cfg.get('TEMPLATE_SOURCE')} TEMPLATE_LOCAL={post_cfg.get('TEMPLATE_LOCAL')}")
            print(f"    TEMPLATE_SELECTED={template_bundle.get('selected')}")
            print(f"    TEMPLATE_REFS={len(template_bundle.get('references') or [])} SLIDES={len(template_bundle.get('slide_references') or [])}")
            print(f"    LOGO={brand_logo_path()}")
            print(f"    INPUT_ID={v(COL_INPUT_ID)} desc={'yes' if input_desc else 'no'} images={len(input_images)}")
            continue

        folder = cfg["outputs"] / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{slugify(post_type)}_{slugify(subject or city)}"
        ensure_directory(folder)
        try:
            forced_caption = extract_forced_caption(input_desc)
            if forced_caption:
                text = forced_caption
                print("    📝 caption imposée depuis description.txt")
            else:
                text = call_llm(system, build_prompt(row, content_type, input_desc, post_cfg, expected_slides=expected_slides))
            generated, image_status = process_images(
                post_cfg,
                row,
                text,
                input_desc,
                input_images,
                template_bundle,
                folder,
                slugify(post_type),
                content_type,
            )
            save_post(text, generated, folder, content_type)
            row[COL_POST_OUTPUT - 1].value = text[:30000]
            row[COL_IMAGE_OUTPUT - 1].value = "; ".join([str(p) for p in generated if p])
            row[COL_IMAGE_STATUS - 1].value = image_status
            row[COL_STATUT - 1].value = STATUT_GENERE
            print(f"    ✅ généré → {folder}")
        except Exception as exc:
            row[COL_STATUT - 1].value = "Erreur"
            row[COL_IMAGE_STATUS - 1].value = "erreur"
            row[COL_POST_OUTPUT - 1].value = str(exc)[:1000]
            print(f"    ❌ {exc}")
    wb.save(cfg["excel"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zone", required=True)
    ap.add_argument("--env", default=social_env_default(), choices=["local", "staging", "prod"])
    ap.add_argument("--init", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    cfg = zone_config(args.zone, args.env)
    if args.init:
        init_zone(args.zone, args.env)
    else:
        run_zone(args.zone, args.env, args.dry_run)


if __name__ == "__main__":
    main()
