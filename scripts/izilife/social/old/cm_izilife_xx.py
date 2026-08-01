"""
Agent Community Manager — izilife v4 final

Usage :
  python cm_izilife.py --zone=lille --env=local --init
  python cm_izilife.py --zone=lille --env=local --dry-run
  python cm_izilife.py --zone=lille --env=local

Principes :
- Planning simple : calendrier éditorial uniquement.
- Onglet Post Types = configuration permanente par type de post.
- .env = fallback global / izilife / zone.
- Image : IMAGE_ENGINE peut valoir both, gpt, pillow, canva, none.
- Pour la phase de test, both = GPT image + composition Pillow quand une image/template existe.
- Canva est préparé mais non branché : statut "canva à faire".
"""

from __future__ import annotations

import os
import sys
import re
import argparse
import base64
import requests
import anthropic
import openpyxl
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.cell.cell import MergedCell


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return


_ensure_core_import_path()

try:
    from core.paths import (
        PROJECT_ROOT,
        ENV_GLOBAL,
        ENV_IZILIFE,
        izilife_social_zone,
        normalize_zone,
    )
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False
    PROJECT_ROOT = Path(os.getenv("AGENTIC_WORKSPACE_ROOT", Path.home() / "Documents" / "agentic_Workspace"))
    ENV_GLOBAL = PROJECT_ROOT / ".env"
    ENV_IZILIFE = PROJECT_ROOT / "izilife" / ".env.izilife"

    def normalize_zone(zone: str) -> str:
        z = str(zone or "").strip().lower()
        return z if z.endswith("-zone") else f"{z}-zone"


CONTEXT_ROOT = PROJECT_ROOT / "izilife" / "context" / "social"
TEMPLATES_DIR = PROJECT_ROOT / "izilife" / "templates"

TYPE_CONTENT = {
    "POST_AGENDA_SEMAINE": "story",
    "POST_AGENDA_WEEKEND": "story",
    "POST_TOP_LIEUX": "carrousel",
    "POST_TOP_ACTEURS": "carrousel",
    "POST_HISTOIRE": "post",
    "POST_HUMOUR": "post",
    "POST_ESCAPADE_VILLE": "carrousel",
    "POST_ESCAPADE_NATURE": "carrousel",
    "POST_PEPITE": "post",
    "EVENT_SERIE": "post",
    "EVENT_ANIMATEUR": "post",
    "PARTAGE_ACTEUR": "post",
    "NOUVEAUTE_IZILIFE": "post",
    "REEL": "reel",
    "VIDEO": "video",
}

TYPE_TEMPLATE = {
    "POST_AGENDA_SEMAINE": "templates/agenda_semaine/",
    "POST_AGENDA_WEEKEND": "templates/agenda_weekend/",
    "POST_TOP_LIEUX": "templates/top/",
    "POST_TOP_ACTEURS": "templates/top/",
    "POST_HISTOIRE": "templates/histoire_lieu/",
    "POST_HUMOUR": "templates/humour_local/",
    "POST_ESCAPADE_VILLE": "templates/escapade_ville/",
    "POST_ESCAPADE_NATURE": "templates/escapade_ville/",
    "POST_PEPITE": "templates/pepite_niche/",
    "EVENT_SERIE": "templates/event_serie/",
    "EVENT_ANIMATEUR": "templates/event_animateur/",
}

# Planning simplifié
COL_DATE = 1
COL_RESEAU = 2
COL_TYPE = 3
COL_VILLE = 4
COL_LIEU = 5
COL_SUJET = 6
COL_DATA = 7
COL_DATE_EVENT = 8
COL_HEURE = 9
COL_INPUT_ID = 10
COL_IMAGE_PROMPT = 11
COL_POST_OUTPUT = 12
COL_IMAGE_OUTPUT = 13
COL_IMAGE_STATUS = 14
COL_STATUT = 15

STATUT_A_FAIRE = "À faire"
STATUT_RELANCER = "Relancer"
STATUT_GENERE = "Généré"

POST_TYPES_HEADERS = [
    "POST_TYPE",
    "CONTENT_TYPE",
    "TEMPLATE_LOCAL",
    "TEMPLATE_SOURCE",
    "TEMPLATE_REF",
    "TEXT_PROVIDER",
    "IMAGE_ENGINE",
    "IMAGE_LLM",
    "IMAGE_MODE",
    "VIDEO_ENGINE",
    "NOTES",
]


def load_env_files(zone: str | None = None, env_name: str | None = None, drive: Path | None = None):
    """Charge les configs sans écraser ce qui est déjà défini par l'OS.

    Ordre :
      1. .env global projet
      2. izilife/.env.izilife
      3. izilife/context/social/.env.social si présent
      4. dossier social de zone/.env si présent
      5. dossier social de zone/.env.{env} si présent
    """
    candidates = [
        ENV_GLOBAL,
        ENV_IZILIFE,
        PROJECT_ROOT / "izilife" / "context" / "social" / ".env.social",
    ]
    if drive is not None:
        candidates.append(drive / ".env")
        if env_name:
            candidates.append(drive / f".env.{env_name}")
    for p in candidates:
        if p and Path(p).exists():
            load_dotenv(p, override=False)


def social_env_default() -> str:
    return os.getenv("SOCIAL_ENV", "prod").strip().lower() or "prod"


def zone_config(zone: str, env_name: str | None = None) -> dict:
    env_name = env_name or social_env_default()
    z = normalize_zone(zone)
    if HAS_CORE_PATHS:
        drive = izilife_social_zone(z, env_name)
    else:
        root_name = {
            "local": "agentic_workspace_local",
            "staging": "agentic_workspace_staging",
            "prod": "agentic_workspace",
        }.get(env_name, "agentic_workspace")
        drive = Path(os.getenv("AGENTIC_DRIVE_ROOT", "G:/Mon Drive")) / root_name / "izilife" / "social" / z
    return {
        "env": env_name,
        "zone": z,
        "drive": drive,
        "excel": drive / f"planning_{z}.xlsx",
        "outputs": drive / "outputs",
        "inputs": drive / "inputs",
        "context_zone": CONTEXT_ROOT / "zones" / f"{z}.md",
    }


def slugify(t: str) -> str:
    t = str(t or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9_]", "", t.lower().replace(" ", "_").replace("-", "_"))[:45] or "post"


def _safe_cell(ws, r: int, c: int):
    x = ws.cell(r, c)
    return None if isinstance(x, MergedCell) else x


def ensure_planning(path: Path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    ws.title = "Planning"

    # Nettoie les anciennes fusions avant restyle.
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(sheet, r, c, v, bg="1A1A2E", sz=9):
        x = _safe_cell(sheet, r, c)
        if x is None:
            return
        x.value = v
        x.font = Font(name="Arial", size=sz, bold=True, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = brd

    groups = [
        (1, 3, "PLANIFICATION", "1A1A2E"),
        (4, 7, "CONTENU", "0F3460"),
        (8, 9, "EVENT", "1B4332"),
        (10, 11, "INPUT / IMAGE", "0D6986"),
        (12, 14, "OUTPUT", "1A5276"),
        (15, 15, "STATUT", "6C3483"),
    ]
    for s, e, label, bg in groups:
        if s < e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        hdr(ws, 1, s, label, bg=bg, sz=10)

    headers = [
        ("DATE", 12),
        ("RÉSEAU", 12),
        ("TYPE", 24),
        ("VILLE", 14),
        ("LIEU", 18),
        ("SUJET", 30),
        ("DATA / INFOS", 52),
        ("DATE EVENT", 14),
        ("HEURE / DURÉE", 16),
        ("INPUT_ID", 18),
        ("IMAGE_PROMPT", 48),
        ("POST OUTPUT", 70),
        ("IMAGE_OUTPUT", 40),
        ("IMAGE_STATUS", 18),
        ("STATUT", 14),
    ]
    for i, (label, width) in enumerate(headers, 1):
        hdr(ws, 2, i, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 42

    validations = [
        (DataValidation(type="list", formula1='"Instagram,Facebook,Les deux"', allow_blank=True), COL_RESEAU),
        (DataValidation(type="list", formula1='"' + ",".join(TYPE_CONTENT.keys()) + '"', allow_blank=True), COL_TYPE),
        (DataValidation(type="list", formula1='"non demandé,généré,erreur,source fournie,template utilisé,canva à faire"', allow_blank=True), COL_IMAGE_STATUS),
        (DataValidation(type="list", formula1='"À faire,Généré,Relancer,Validé,Publié,Skip,Erreur"', allow_blank=True), COL_STATUT),
    ]
    for dv, col in validations:
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}500")

    # Post Types : on conserve les réglages existants. On ajoute seulement les colonnes/lignes manquantes.
    if "Post Types" in wb.sheetnames:
        wt = wb["Post Types"]
    else:
        wt = wb.create_sheet("Post Types")

    existing_headers = [str(c.value or "").strip() for c in wt[1]] if wt.max_row else []
    for h in POST_TYPES_HEADERS:
        if h not in existing_headers:
            wt.cell(row=1, column=len(existing_headers) + 1, value=h)
            existing_headers.append(h)
    header_map = {h: i + 1 for i, h in enumerate(existing_headers)}

    for i, h in enumerate(existing_headers, 1):
        hdr(wt, 1, i, h, "1A1A2E")
        wt.column_dimensions[get_column_letter(i)].width = {
            "POST_TYPE": 28,
            "CONTENT_TYPE": 14,
            "TEMPLATE_LOCAL": 30,
            "TEMPLATE_SOURCE": 18,
            "TEMPLATE_REF": 36,
            "TEXT_PROVIDER": 18,
            "IMAGE_ENGINE": 18,
            "IMAGE_LLM": 18,
            "IMAGE_MODE": 18,
            "VIDEO_ENGINE": 18,
            "NOTES": 72,
        }.get(h, 20)

    existing_post_types = set()
    for row in wt.iter_rows(min_row=2):
        val = row[header_map["POST_TYPE"] - 1].value if "POST_TYPE" in header_map else None
        if val:
            existing_post_types.add(str(val).strip())

    defaults = {
        "TEXT_PROVIDER": os.getenv("LLM_PROVIDER", "claude"),
        "IMAGE_ENGINE": os.getenv("IMAGE_ENGINE", "both"),
        "IMAGE_LLM": os.getenv("IMAGE_LLM", "gpt"),
        "IMAGE_MODE": "auto",
        "VIDEO_ENGINE": os.getenv("VIDEO_ENGINE", "none"),
        "NOTES": "owned=template interne ; inspiration=référence non copiée ; none=libre. IMAGE_ENGINE both=gpt+pillow.",
    }

    for pt, ct in TYPE_CONTENT.items():
        if pt in existing_post_types:
            continue
        r = wt.max_row + 1
        values = {
            "POST_TYPE": pt,
            "CONTENT_TYPE": ct,
            "TEMPLATE_LOCAL": TYPE_TEMPLATE.get(pt, ""),
            "TEMPLATE_SOURCE": "owned" if pt in TYPE_TEMPLATE else "none",
            "TEMPLATE_REF": "",
            **defaults,
        }
        for h, v in values.items():
            wt.cell(row=r, column=header_map[h], value=v)

    for col_name, values in [
        ("TEMPLATE_SOURCE", "owned,inspiration,none"),
        ("TEXT_PROVIDER", "claude,openai"),
        ("IMAGE_ENGINE", "both,gpt,pillow,canva,none"),
        ("IMAGE_LLM", "gpt,canva,stable,none"),
        ("IMAGE_MODE", "auto,template,input,generate,none"),
        ("VIDEO_ENGINE", "none,runwayml,pika,kling"),
    ]:
        if col_name in header_map:
            dv = DataValidation(type="list", formula1='"' + values + '"', allow_blank=True)
            wt.add_data_validation(dv)
            dv.add(f"{get_column_letter(header_map[col_name])}2:{get_column_letter(header_map[col_name])}200")

    # Mode d'emploi
    if "Mode d'emploi" in wb.sheetnames:
        del wb["Mode d'emploi"]
    wi = wb.create_sheet("Mode d'emploi", 0)
    lines = [
        "CM izilife — mode d'emploi rapide",
        "1. Remplir uniquement l'onglet Planning pour les posts à produire.",
        "2. Les réglages fixes sont dans Post Types, une seule fois par TYPE.",
        "3. TEMPLATE_SOURCE : owned = template interne ; inspiration = référence non copiée ; none = génération libre.",
        "4. TEMPLATE_REF : chemin local, Canva ID, URL ou note. Jamais à répéter ligne par ligne.",
        "5. TEXT_PROVIDER : claude/openai. Vide = fallback .env.",
        "6. IMAGE_ENGINE : both/gpt/pillow/canva/none. Pendant les tests, both = GPT image + Pillow quand une source existe.",
        "7. IMAGE_LLM : moteur image LLM. Pour l'instant gpt. Canva sera branché plus tard.",
        "8. VIDEO_ENGINE : none pour le moment. Prévu pour runwayml/pika/kling.",
        "9. INPUT_ID pointe vers inputs/[INPUT_ID]/description.txt et/ou des images.",
        "10. Le script ne touche pas aux lignes Validé / Publié / Skip.",
        "11. Pour relancer une ligne : STATUT = Relancer.",
        "12. Config .env : global, izilife/.env.izilife, context/social/.env.social, puis dossier zone/.env ou .env.local.",
    ]
    for r, line in enumerate(lines, 1):
        wi.cell(r, 1, line)
        wi.column_dimensions["A"].width = 140
        if r == 1:
            wi.cell(r, 1).font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            wi.cell(r, 1).fill = PatternFill("solid", fgColor="1A1A2E")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"✅ Planning CM izilife prêt : {path}")


def load_post_type_config(wb, post_type: str) -> dict:
    if "Post Types" not in wb.sheetnames:
        return {}
    ws = wb["Post Types"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if str(d.get("POST_TYPE") or "").strip() == post_type:
            return d
    return {}


def init_zone(zone: str, env_name: str):
    cfg = zone_config(zone, env_name)
    load_env_files(zone=cfg["zone"], env_name=env_name, drive=cfg["drive"])
    for d in [cfg["outputs"], cfg["inputs"]]:
        d.mkdir(parents=True, exist_ok=True)
    cfg["context_zone"].parent.mkdir(parents=True, exist_ok=True)
    if not cfg["context_zone"].exists():
        cfg["context_zone"].write_text(
            f"""# Contexte zone — {cfg['zone']}
## Spécificités
- Ville principale :
- Hashtags locaux : #{cfg['zone'].replace('-zone','')}
- Compte Instagram : @izilife_{cfg['zone'].replace('-zone','')}
- Événements locaux :
- Lieux emblématiques :
""",
            encoding="utf-8",
        )
    ensure_planning(cfg["excel"])


def load_context(zone: str) -> str:
    parts = []
    for fn in ["izilife-social-strategy.md", "community-manager.md", "izilife-style.md"]:
        p = CONTEXT_ROOT / fn
        if p.exists():
            parts.append(f"### {fn}\n{p.read_text(encoding='utf-8')}")
    zp = CONTEXT_ROOT / "zones" / f"{zone}.md"
    if zp.exists():
        parts.append(f"### {zone}.md\n{zp.read_text(encoding='utf-8')}")
    return "\n\n---\n\n".join(parts)


def read_input(input_id: str, inputs_dir: Path):
    if not input_id:
        return "", []
    folder = inputs_dir / input_id
    if not folder.exists():
        return "", []
    desc = (folder / "description.txt").read_text(encoding="utf-8").strip() if (folder / "description.txt").exists() else ""
    imgs = sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")])
    return desc, imgs


def build_prompt(row, content_type: str, input_desc: str, cfg: dict) -> str:
    def v(c):
        val = row[c - 1].value
        return str(val).strip() if val else ""

    slides = "\n=== SLIDES ===\nSlide 1 : ...\nSlide 2 : ..." if content_type == "carrousel" else ""
    return f"""Tu es l'agent Community Manager izilife. CONTENT_TYPE: {content_type}

CONFIG VISUELLE DU TYPE :
TEMPLATE_SOURCE: {cfg.get('TEMPLATE_SOURCE','none')}
TEMPLATE_REF: {cfg.get('TEMPLATE_REF','')}
IMAGE_ENGINE: {cfg.get('IMAGE_ENGINE','')}
IMAGE_LLM: {cfg.get('IMAGE_LLM','')}
IMAGE_MODE: {cfg.get('IMAGE_MODE','auto')}
VIDEO_ENGINE: {cfg.get('VIDEO_ENGINE','none')}

DONNÉES :
TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
SUJET: {v(COL_SUJET)}
DATA: {v(COL_DATA)}
DATE EVENT: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
DESCRIPTION INPUT:
{input_desc}

Règles :
- 1 seul CTA vers izilife.
- Ne jamais inventer un lieu, un acteur ou un événement non fourni.
- Ton local, direct, pas corporate.

Format OBLIGATOIRE :
=== TEXTE DU POST ===
[caption complète avec emojis + hashtags]{slides}
"""


def choose_text_provider(cfg: dict) -> str:
    return str(cfg.get("TEXT_PROVIDER") or os.getenv("LLM_PROVIDER") or "claude").strip().lower()


def call_llm(system: str, user: str, cfg: dict) -> str:
    provider = choose_text_provider(cfg)
    if provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o"),
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            max_tokens=1500,
        )
        return r.choices[0].message.content or ""

    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    msg = client.messages.create(
        model=os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6"),
        max_tokens=1500,
        system=system,
        messages=[{"role": "user", "content": user}],
    )
    return msg.content[0].text


def find_template(template_local: str, template_ref: str):
    candidates = []
    ref = str(template_ref or "").strip()
    if ref and not ref.lower().startswith(("http://", "https://", "dag")):
        candidates.append(Path(ref))
        candidates.append(TEMPLATES_DIR / ref)
    if template_local:
        candidates.append(TEMPLATES_DIR / str(template_local).replace("templates/", ""))
    for c in candidates:
        if c.is_file():
            return c
        if c.is_dir():
            files = []
            for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
                files += sorted(c.glob(ext))
            if files:
                return files[0]
    return None


def generate_image_gpt(prompt: str, folder: Path, filename: str):
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    result = client.images.generate(
        model=os.getenv("IMAGE_MODEL", "gpt-image-1"),
        prompt=f"{prompt}. Style Instagram, format portrait, composition originale.",
        n=1,
        size=os.getenv("IMAGE_SIZE", "1024x1536"),
    )
    data = result.data[0]
    if getattr(data, "b64_json", None):
        out = folder / f"{filename}_gpt.png"
        out.write_bytes(base64.b64decode(data.b64_json))
        return out
    if getattr(data, "url", None):
        out = folder / f"{filename}_gpt.jpg"
        out.write_bytes(requests.get(data.url, timeout=60).content)
        return out
    return None


def compose_pillow(bg: Path, text: str, folder: Path, filename: str):
    from PIL import Image, ImageDraw, ImageFont
    img = Image.open(bg).convert("RGBA")
    W, H = img.size
    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 130))
    img = Image.alpha_composite(img, overlay)
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", max(34, W // 22))
    except Exception:
        font = ImageFont.load_default()
    lines = [l.strip() for l in text.split("\n") if l.strip() and not l.startswith("===")][:8]
    y = H // 4
    for line in lines:
        bb = draw.textbbox((0, 0), line, font=font)
        x = (W - (bb[2] - bb[0])) // 2
        draw.text((x + 2, y + 2), line, font=font, fill=(0, 0, 0, 180))
        draw.text((x, y), line, font=font, fill=(255, 255, 255, 255))
        y += (bb[3] - bb[1]) + 16
    out = folder / f"{filename}_pillow.jpg"
    img.convert("RGB").save(out, quality=95)
    return out


def process_images(post_type: str, cfg: dict, img_prompt: str, input_images: list[Path], text: str, folder: Path, filename: str):
    engine = str(cfg.get("IMAGE_ENGINE") or os.getenv("IMAGE_ENGINE") or "both").strip().lower()
    image_llm = str(cfg.get("IMAGE_LLM") or os.getenv("IMAGE_LLM") or "gpt").strip().lower()
    mode = str(cfg.get("IMAGE_MODE") or "auto").strip().lower()
    tpl = find_template(str(cfg.get("TEMPLATE_LOCAL") or ""), str(cfg.get("TEMPLATE_REF") or ""))
    bg = input_images[0] if input_images else tpl
    generated: list[Path] = []

    if engine == "none" or mode == "none":
        return generated, "non demandé"

    if engine == "canva":
        return generated, "canva à faire"

    # Phase de test : both = sortie LLM + sortie Pillow si une source existe.
    if engine in ("pillow", "both") and bg:
        try:
            p = compose_pillow(bg, text, folder, filename)
            if p:
                generated.append(p)
        except Exception as e:
            print(f"    ❌ Pillow : {e}")

    if engine in ("gpt", "both") and image_llm in ("gpt", "openai"):
        prompt = img_prompt or "\n".join([
            str(cfg.get("TEMPLATE_REF") or ""),
            str(cfg.get("TEMPLATE_SOURCE") or ""),
            text[:700],
        ])
        if prompt.strip():
            try:
                p = generate_image_gpt(prompt, folder, filename)
                if p:
                    generated.append(p)
            except Exception as e:
                print(f"    ❌ GPT Image : {e}")

    if generated:
        return generated, "généré"
    if bg:
        return [], "template utilisé"
    return [], "non demandé"


def save_post(text: str, generated: list[Path], folder: Path, content_type: str, runtime: dict):
    lines = [
        text.strip(),
        "",
        f"=== CONTENT_TYPE ===\n{content_type}",
        "",
        "=== CONFIG_RUNTIME ===",
        f"TEXT_PROVIDER={runtime.get('TEXT_PROVIDER')}",
        f"IMAGE_ENGINE={runtime.get('IMAGE_ENGINE')}",
        f"IMAGE_LLM={runtime.get('IMAGE_LLM')}",
        f"VIDEO_ENGINE={runtime.get('VIDEO_ENGINE')}",
    ]
    if generated:
        lines += ["", "=== IMAGES GENEREES ==="] + [f"  {p.name}" for p in generated if p]
    (folder / "post.txt").write_text("\n".join(lines), encoding="utf-8")


def runtime_from_cfg(cfg: dict) -> dict:
    return {
        "TEXT_PROVIDER": str(cfg.get("TEXT_PROVIDER") or os.getenv("LLM_PROVIDER") or "claude").lower(),
        "IMAGE_ENGINE": str(cfg.get("IMAGE_ENGINE") or os.getenv("IMAGE_ENGINE") or "both").lower(),
        "IMAGE_LLM": str(cfg.get("IMAGE_LLM") or os.getenv("IMAGE_LLM") or "gpt").lower(),
        "VIDEO_ENGINE": str(cfg.get("VIDEO_ENGINE") or os.getenv("VIDEO_ENGINE") or "none").lower(),
    }


def run_zone(zone: str, env_name: str, dry_run=False):
    cfg = zone_config(zone, env_name)
    load_env_files(zone=cfg["zone"], env_name=env_name, drive=cfg["drive"])
    zone = cfg["zone"]
    if not cfg["excel"].exists():
        print(f"❌ Planning introuvable. Lance --init : {cfg['excel']}")
        return

    system = load_context(zone)
    wb = openpyxl.load_workbook(cfg["excel"])
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    rows = [r for r in ws.iter_rows(min_row=3) if str(r[COL_STATUT - 1].value or "").strip() in (STATUT_A_FAIRE, STATUT_RELANCER)]
    print(f"🤖 CM izilife V4 final — {zone} — env={env_name} — {len(rows)} ligne(s)" + (" [DRY RUN]" if dry_run else ""))
    cfg["outputs"].mkdir(parents=True, exist_ok=True)
    cfg["inputs"].mkdir(parents=True, exist_ok=True)

    for idx, row in enumerate(rows, 1):
        def v(c):
            val = row[c - 1].value
            return str(val).strip() if val else ""

        post_type = v(COL_TYPE)
        subject = v(COL_SUJET)
        city = v(COL_VILLE)
        pt_cfg = load_post_type_config(wb, post_type)
        runtime = runtime_from_cfg(pt_cfg)
        content_type = str(pt_cfg.get("CONTENT_TYPE") or TYPE_CONTENT.get(post_type, "post"))
        print(f"  [{idx}/{len(rows)}] {post_type} [{content_type}] — {city} — {subject}")
        input_desc, input_images = read_input(v(COL_INPUT_ID), cfg["inputs"])

        if dry_run:
            print(
                "    CONFIG: "
                f"text={runtime['TEXT_PROVIDER']} "
                f"image_engine={runtime['IMAGE_ENGINE']} "
                f"image_llm={runtime['IMAGE_LLM']} "
                f"video={runtime['VIDEO_ENGINE']} "
                f"template_source={pt_cfg.get('TEMPLATE_SOURCE')} "
                f"template_ref={pt_cfg.get('TEMPLATE_REF')}"
            )
            continue

        folder = cfg["outputs"] / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{slugify(post_type)}_{slugify(subject or city)}"
        folder.mkdir(parents=True, exist_ok=True)
        try:
            text = call_llm(system, build_prompt(row, content_type, input_desc, pt_cfg), pt_cfg)
            generated, image_status = process_images(post_type, pt_cfg, v(COL_IMAGE_PROMPT), input_images, text, folder, slugify(post_type))
            save_post(text, generated, folder, content_type, runtime)
            row[COL_POST_OUTPUT - 1].value = text[:30000]
            row[COL_IMAGE_OUTPUT - 1].value = "; ".join([str(p) for p in generated if p])
            row[COL_IMAGE_STATUS - 1].value = image_status
            row[COL_STATUT - 1].value = STATUT_GENERE
            print(f"    ✅ généré → {folder}")
        except Exception as e:
            row[COL_STATUT - 1].value = "Erreur"
            row[COL_IMAGE_STATUS - 1].value = "erreur"
            row[COL_POST_OUTPUT - 1].value = str(e)[:1000]
            print(f"    ❌ {e}")
    wb.save(cfg["excel"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zone", required=True)
    ap.add_argument("--env", default=social_env_default(), choices=["local", "staging", "prod"])
    ap.add_argument("--init", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if args.init:
        init_zone(args.zone, args.env)
    else:
        run_zone(args.zone, args.env, args.dry_run)


if __name__ == "__main__":
    main()
