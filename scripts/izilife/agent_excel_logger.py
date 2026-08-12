"""Journal Excel commun des vrais agents iziLife (aucune dependance a pilotage)."""
from __future__ import annotations

import io
import os
import re
import sys
import time
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "run_id", "started_at", "ended_at", "duration_seconds", "agent", "environment",
    "zone", "status", "processed", "created", "skipped", "errors", "command", "details",
]
PILOTAGE_HEADERS = [
    "started_at", "etat", "agent", "environment", "zone", "duration_seconds",
    "processed", "created", "skipped", "errors", "detail_file", "message",
]


def _load_workbook_or_new(path: Path, sheet_name: str, headers: list[str]):
    """Charge un XLSX valide ou recrée un fichier Drive incomplet."""
    if path.exists():
        try:
            return openpyxl.load_workbook(path), False
        except (KeyError, zipfile.BadZipFile, EOFError, OSError):
            try:
                path.unlink()
            except OSError:
                pass
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    return workbook, True


def _save_workbook_atomic(workbook, path: Path) -> None:
    """Écrit un XLSX complet avant de remplacer le fichier visible par Google Drive."""
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            try:
                temporary.unlink()
            except OSError:
                pass


def _arg(name: str, default: str = "") -> str:
    prefix = "--" + name + "="
    for index, value in enumerate(sys.argv[1:]):
        if value.startswith(prefix):
            return value[len(prefix):]
        if value == "--" + name and index + 2 <= len(sys.argv) - 1:
            return sys.argv[index + 2]
    return default


def _workspace(env_name: str) -> Path:
    drive = os.environ.get("AGENTIC_DRIVE_ROOT", "").strip()
    if not drive:
        raise RuntimeError("AGENTIC_DRIVE_ROOT non defini : impossible d'ecrire le journal Excel.")
    suffix = {"local": "agentic_workspace_local", "staging": "agentic_workspace_staging", "prod": "agentic_workspace"}
    return Path(drive).expanduser() / suffix.get(env_name, suffix["prod"])


def _zone(value: str) -> str:
    value = (value or "global").strip().lower().replace("_", "-")
    return value if value.endswith("-zone") else value + "-zone"


def log_path(domain: str, agent: str, env_name: str, zone: str, day: datetime) -> Path:
    return _workspace(env_name) / "izilife" / domain / _zone(zone) / "logs" / f"{agent}_{day:%Y-%m-%d}.xlsx"


class _Tee(io.TextIOBase):
    def __init__(self, original):
        self.original = original
        self.buffer = io.StringIO()

    def write(self, value):
        self.buffer.write(value)
        return self.original.write(value)

    def flush(self):
        self.original.flush()


def _count(text: str, labels: list[str]) -> int | None:
    for label in labels:
        matches = re.findall(rf"(?im){re.escape(label)}\s*[:=]\s*(\d+)", text)
        if matches:
            return int(matches[-1])
    return None


def _append(path: Path, values: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(5):
        try:
            workbook, created = _load_workbook_or_new(path, "Executions", HEADERS)
            sheet = workbook["Executions"] if "Executions" in workbook.sheetnames else workbook.active
            if created:
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="24364B")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                widths = [20, 20, 20, 18, 25, 14, 18, 14, 12, 12, 12, 12, 55, 100]
                for index, width in enumerate(widths, 1):
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = "A1:N1"
            sheet.append(values)
            row = sheet.max_row
            sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row, 3).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row, 14).alignment = Alignment(wrap_text=True, vertical="top")
            status = str(values[7]).upper()
            sheet.cell(row, 8).fill = PatternFill("solid", fgColor={"DONE": "C6EFCE", "ERROR": "FFC7CE", "INIT": "DDEBF7"}.get(status, "FFF2CC"))
            _save_workbook_atomic(workbook, path)
            return
        except PermissionError:
            if attempt == 4:
                raise
            time.sleep(1.5)


def _append_pilotage(path: Path, values: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(5):
        try:
            workbook, created = _load_workbook_or_new(path, "Pilotage", PILOTAGE_HEADERS)
            sheet = workbook["Pilotage"] if "Pilotage" in workbook.sheetnames else workbook.active
            if created:
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="24364B")
                for index, width in enumerate([20, 12, 26, 14, 18, 18, 12, 12, 12, 12, 85, 80], 1):
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = "A1:L1"
            sheet.append(values)
            row = sheet.max_row
            sheet.cell(row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row, 2).font = Font(bold=True)
            sheet.cell(row, 2).fill = PatternFill("solid", fgColor={"VERT": "C6EFCE", "ORANGE": "FFF2CC", "ROUGE": "FFC7CE"}[values[1]])
            sheet.cell(row, 12).alignment = Alignment(wrap_text=True, vertical="top")
            _save_workbook_atomic(workbook, path)
            return
        except PermissionError:
            if attempt == 4:
                raise
            time.sleep(1.5)


def run_logged(agent: str, domain: str, callback) -> None:
    started = datetime.now()
    started_clock = time.monotonic()
    env_name = _arg("env", "prod").lower()
    zone = _arg("zone", _arg("city", "global"))
    tee_out, tee_err = _Tee(sys.stdout), _Tee(sys.stderr)
    old_out, old_err = sys.stdout, sys.stderr
    status, failure = ("INIT" if "--init" in sys.argv else "DONE"), None
    sys.stdout, sys.stderr = tee_out, tee_err
    try:
        callback()
    except BaseException as exc:
        status, failure = "ERROR", exc
    finally:
        sys.stdout, sys.stderr = old_out, old_err
        ended = datetime.now()
        output = (tee_out.buffer.getvalue() + "\n" + tee_err.buffer.getvalue()).strip()
        if failure:
            output = (output + f"\n{type(failure).__name__}: {failure}").strip()
        values = [
            uuid.uuid4().hex[:16], started, ended, round(time.monotonic() - started_clock, 2), agent,
            env_name, _zone(zone), status,
            _count(output, ["Traites", "Traités", "Total", "Events traites", "Events traités"]),
            _count(output, ["Crees", "Créés", "Inseres", "Insérés", "Occurrences créées"]),
            _count(output, ["Skippes", "Skippés", "Ignores", "Ignorés"]),
            _count(output, ["Erreurs", "Errors"]), " ".join(sys.argv), output[-30000:],
        ]
        try:
            path = log_path(domain, agent, env_name, zone, started)
            _append(path, values)
            error_count = values[11] or 0
            state = "ROUGE" if status == "ERROR" else ("ORANGE" if error_count else "VERT")
            central = _workspace(env_name) / "izilife" / "logs" / "pilotage_global.xlsx"
            message = (f"{type(failure).__name__}: {failure}" if failure else ("Execution terminee avec erreurs" if error_count else "OK"))
            _append_pilotage(central, [started, state, agent, env_name, _zone(zone), values[3], values[8], values[9], values[10], values[11], str(path), message])
            old_out.write(f"\nJournal Excel : {path}\n")
            old_out.write(f"Pilotage global : {central}\n")
        except Exception as log_error:
            old_err.write(f"\nERREUR journal Excel : {log_error}\n")
    if failure:
        raise failure
