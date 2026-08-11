from __future__ import annotations

import json, os, re, time, random
from pathlib import Path
from typing import Any
import requests
import openpyxl
from dotenv import load_dotenv


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            import sys
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            import sys
            sys.path.insert(0, str(parent))
            return

_ensure_core_import_path()

try:
    from core.paths import (
        PROJECT_ROOT, ENV_GLOBAL, ENV_IZILIFE, IZILIFE_ENVS,
        drive_workspace_root, normalize_zone, normalize_env,
    )
except Exception:
    PROJECT_ROOT = Path.home() / "Documents" / "agentic_Workspace"
    ENV_GLOBAL = PROJECT_ROOT / ".env"
    ENV_IZILIFE = PROJECT_ROOT / "izilife" / ".env.izilife"
    IZILIFE_ENVS = {
        "local": {"base_url": os.getenv("IZILIFE_LOCAL_URL", "https://localhost:4443/izilife-admin"), "verify_ssl": False},
        "staging": {"base_url": os.getenv("IZILIFE_STAGING_URL", "https://www.staging.izilife.co/izilife-admin"), "verify_ssl": True},
        "prod": {"base_url": os.getenv("IZILIFE_PROD_URL", "https://www.izilife.co/izilife-admin"), "verify_ssl": True},
    }
    def normalize_env(e):
        e = str(e or "prod").strip().lower()
        return e if e in ("local", "staging", "prod") else "prod"
    def normalize_zone(z):
        z = str(z or "").strip().lower()
        return z if z.endswith("-zone") else f"{z}-zone"
    def drive_workspace_root(env_name="prod"):
        root = Path(os.environ["AGENTIC_DRIVE_ROOT"]).expanduser()
        return root / {"local":"agentic_workspace_local","staging":"agentic_workspace_staging","prod":"agentic_workspace"}[normalize_env(env_name)]


def load_env():
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_IZILIFE.exists(): load_dotenv(ENV_IZILIFE, override=False)


def object_zone_dir(object_name: str, zone: str, env_name: str = "prod") -> Path:
    return drive_workspace_root(env_name) / "izilife" / object_name / normalize_zone(zone)


def log(msg: str):
    print(msg, flush=True)


def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text[:150] or "item"


def now_iso():
    import datetime as dt
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def read_rows(path: Path, sheet_name="Sources") -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2):
        d = {headers[i]: r[i].value for i in range(len(headers)) if headers[i]}
        if any(v not in (None, "") for v in d.values()):
            d["__row"] = r[0].row
            rows.append(d)
    return rows


def update_cell(path: Path, row: int, col_name: str, value: Any, sheet_name="Sources"):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    if col_name not in headers:
        ws.cell(row=1, column=len(headers)+1, value=col_name)
        headers.append(col_name)
    col = headers.index(col_name) + 1
    ws.cell(row=row, column=col, value=value)
    wb.save(path)


def init_sheet(path: Path, headers: list[str], sample_rows: list[list[Any]] | None = None, sheet_name="Sources"):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        log(f"⚠️ Existe déjà : {path}")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(h)+4, 12), 42)
    for row in sample_rows or []:
        ws.append(row)
    wb.save(path)
    log(f"✅ Créé : {path}")


def izilife_post(path: str, payload: dict[str, Any], env_name: str, timeout=120) -> dict[str, Any]:
    load_env()
    env_name = normalize_env(env_name)
    env = IZILIFE_ENVS[env_name]
    token = os.getenv("IZILIFE_AGENT_TOKEN", "").strip()
    if not token:
        raise RuntimeError("IZILIFE_AGENT_TOKEN manquant")
    url = env["base_url"].rstrip("/") + path
    r = requests.post(url, data=payload, headers={"X-Agent-Token": token}, verify=env.get("verify_ssl", True), timeout=timeout)
    try:
        data = r.json()
    except Exception:
        data = {"success": False, "error": r.text[:1000], "status_code": r.status_code}
    if r.status_code >= 400:
        data.setdefault("success", False)
        data.setdefault("error", f"HTTP {r.status_code}")
    return data


def fetch_clean_text(url: str, timeout=30) -> str:
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent":"Mozilla/5.0 IzilifeAgent/1.0"})
        html = r.text
        html = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
        html = re.sub(r"<style[\s\S]*?</style>", " ", html, flags=re.I)
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text)
        return text[:50000]
    except Exception as e:
        return f"FETCH_ERROR: {e}"


def call_openai_json(system: str, user: str, model_env="OPENAI_MODEL") -> dict[str, Any]:
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    res = client.chat.completions.create(
        model=os.getenv(model_env, "gpt-4o"),
        messages=[{"role":"system","content":system},{"role":"user","content":user}],
        temperature=0.2,
        response_format={"type":"json_object"},
    )
    raw = res.choices[0].message.content or "{}"
    return json.loads(raw)
