"""
Agent liens Izilife

Usage:
  python .\scripts\izilife\objects\agent_links.py --zone=lille --env=local --init
  python .\scripts\izilife\objects\agent_links.py --zone=lille --env=local --dry-run
  python .\scripts\izilife\objects\agent_links.py --zone=lille --env=local

Sheet:
  G:/Mon Drive/agentic_workspace[_local|_staging]/izilife/links/lille-zone/links.xlsx
"""

from __future__ import annotations

import os, sys, re, argparse
from pathlib import Path
from datetime import datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts")); return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent)); return
_ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, ENV_GLOBAL, ENV_IZILIFE, normalize_zone, normalize_env, drive_workspace_root, IZILIFE_ENVS
except Exception:
    PROJECT_ROOT = Path.home() / "Documents" / "agentic_Workspace"
    ENV_GLOBAL = PROJECT_ROOT / ".env"
    ENV_IZILIFE = PROJECT_ROOT / "izilife" / ".env.izilife"
    def normalize_zone(z: str) -> str:
        z = str(z or "").strip().lower()
        return z if z.endswith("-zone") else f"{z}-zone"
    def normalize_env(e: str | None) -> str:
        e = str(e or "prod").strip().lower()
        return e if e in ("local","staging","prod") else "prod"
    def drive_workspace_root(env_name="prod"):
        roots = {"local":"agentic_workspace_local", "staging":"agentic_workspace_staging", "prod":"agentic_workspace"}
        return Path(os.getenv("AGENTIC_DRIVE_ROOT", "G:/Mon Drive")) / roots[normalize_env(env_name)]
    IZILIFE_ENVS = {
        "local": {"base_url": os.getenv("IZILIFE_LOCAL_URL", "https://localhost:4443/izilife-admin"), "verify_ssl": False},
        "staging": {"base_url": os.getenv("IZILIFE_STAGING_URL", "https://www.staging.izilife.co/izilife-admin"), "verify_ssl": True},
        "prod": {"base_url": os.getenv("IZILIFE_PROD_URL", "https://www.izilife.co/izilife-admin"), "verify_ssl": True},
    }

COLS = [
    "target_type", "target_string_id", "target_id", "url", "title",
    "is_official_website", "highlighted", "button_color", "button_shape", "font",
    "start_date", "end_date", "order", "statut", "last_result"
]
TARGET_TYPES = ["PLACE","SHOP","PAGE","PARTNER","EVENT","EXPERIENCE","EVENT_SERIE","ANNUAL_CELEBRATION"]
STATUSES = ["pending", "done", "error", "skip"]


def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def load_env():
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_IZILIFE.exists(): load_dotenv(ENV_IZILIFE, override=False)


def zone_dir(zone: str, env_name: str) -> Path:
    return drive_workspace_root(env_name) / "izilife" / "links" / normalize_zone(zone)


def sheet_path(zone: str, env_name: str) -> Path:
    return zone_dir(zone, env_name) / "links.xlsx"


def make_header(ws):
    thin = Side(style='thin', color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="1A1A2E")
    widths = [18,28,12,60,36,18,12,14,14,14,18,18,10,12,60]
    for i, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = fill; c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]
    ws.freeze_panes = "A2"


def init_sheet(zone: str, env_name: str):
    f = sheet_path(zone, env_name)
    f.parent.mkdir(parents=True, exist_ok=True)
    if f.exists():
        log(f"⚠️ existe déjà : {f}"); return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Links"
    make_header(ws)
    ws.add_data_validation(DataValidation(type="list", formula1='"' + ','.join(TARGET_TYPES) + '"', allow_blank=False))
    ws.data_validations.dataValidation[-1].add("A2:A500")
    ws.add_data_validation(DataValidation(type="list", formula1='"0,1"', allow_blank=True)); ws.data_validations.dataValidation[-1].add("F2:G500")
    ws.add_data_validation(DataValidation(type="list", formula1='"' + ','.join(STATUSES) + '"', allow_blank=True)); ws.data_validations.dataValidation[-1].add("N2:N500")
    ws.append(["PLACE", "grand-place-lille", "", "https://example.com", "Site officiel", 1, 0, "#EE5733", "", "Arial", "", "", 0, "skip", "exemple"])
    wb.save(f)
    log(f"✅ links.xlsx créé : {f}")


def row_to_payload(row) -> dict:
    values = {COLS[i]: row[i].value for i in range(len(COLS))}
    return {
        "target_type": str(values.get("target_type") or "").strip().upper(),
        "target_string_id": str(values.get("target_string_id") or "").strip(),
        "target_id": values.get("target_id"),
        "url": str(values.get("url") or "").strip(),
        "title": str(values.get("title") or "").strip(),
        "is_official_website": int(values.get("is_official_website") or 0),
        "highlighted": int(values.get("highlighted") or 0),
        "button_color": str(values.get("button_color") or "").strip(),
        "button_shape": str(values.get("button_shape") or "").strip(),
        "font": str(values.get("font") or "").strip(),
        "start_date": values.get("start_date"),
        "end_date": values.get("end_date"),
        "order": int(values.get("order") or 0),
    }


def post_json(env: dict, endpoint: str, data: dict) -> dict:
    token = os.getenv("IZILIFE_AGENT_TOKEN") or os.getenv("AGENT_TOKEN") or ""
    r = requests.post(env["base_url"] + endpoint, data=data, headers={"X-Agent-Token": token}, verify=env["verify_ssl"], timeout=60)
    try:
        return r.json()
    except Exception:
        return {"success": False, "error": f"HTTP {r.status_code}: {r.text[:500]}"}


def run(zone: str, env_name: str, dry_run=False, max_rows=100):
    load_env(); env_name = normalize_env(env_name); env = IZILIFE_ENVS[env_name]
    f = sheet_path(zone, env_name)
    if not f.exists():
        log(f"❌ Sheet introuvable. Lance --init : {f}"); return
    wb = openpyxl.load_workbook(f); ws = wb["Links"] if "Links" in wb.sheetnames else wb.active
    done = err = skip = 0
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, max_rows+1)):
        status = str(row[13].value or "pending").strip().lower()
        if status in ("done", "skip"): continue
        p = row_to_payload(row)
        if not p["target_type"] or not p["url"]:
            row[13].value = "error"; row[14].value = "target_type/url obligatoires"; err += 1; continue
        label = f"{p['target_type']} {p.get('target_string_id') or p.get('target_id')} → {p['url']}"
        if dry_run:
            log(f"[DRY RUN] {label}"); continue
        res = post_json(env, "/scraper/agentUpsertLink", p)
        if res.get("success"):
            row[13].value = "done"; row[14].value = res.get("message", "ok"); done += 1; log(f"✅ {label}")
        else:
            row[13].value = "error"; row[14].value = res.get("error", "Erreur inconnue"); err += 1; log(f"❌ {label}: {row[14].value}")
    if not dry_run: wb.save(f)
    log(f"Résultat: done={done} errors={err} skip={skip}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zone", required=True)
    ap.add_argument("--env", default="prod", choices=["local","staging","prod"])
    ap.add_argument("--init", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--max", type=int, default=100)
    args = ap.parse_args()
    if args.init: init_sheet(args.zone, args.env)
    else: run(args.zone, args.env, args.dry_run, args.max)

if __name__ == "__main__": main()
