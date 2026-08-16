from __future__ import annotations
import argparse,os,random,re,sys,time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from playwright.sync_api import sync_playwright
sys.path.insert(0,str(Path(__file__).resolve().parents[1]/"offers"))
from offer_agent_common import drive_workspace_root,normalize_zone

IGNORE={"toutes les publications","audio","outing ideas","actus","lectures"}
MERGE={"curate lieux à mettre":"Lieux Curate","lieux curate 2":"Lieux Curate"}
SUMMARY=["LISTE","URL","ONGLET","DERNIER_NUMERO","NB_LIENS","DERNIER_SCAN","ETAT"]
ITEMS=["NUMERO","NOM","URL","ETAT","AJOUTE_LE"]

def pause(a=.7,b=1.7): time.sleep(random.uniform(a,b))
def sheet_name(v): return (re.sub(r'[\\/*?:\[\]]',' ',v).strip() or "Collection")[:31]
def style(ws):
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="4472C4")
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(55,max(14,max(len(str(c.value or "")) for c in col)+2))
def book(path):
    if path.exists(): return openpyxl.load_workbook(path)
    path.parent.mkdir(parents=True,exist_ok=True); wb=openpyxl.Workbook(); wb.active.title="Résumé"; wb.active.append(SUMMARY); return wb
def scroll(page,rounds=120):
    old=-1; stable=0
    for _ in range(rounds):
        n=page.locator('a[href*="/p/"],a[href*="/reel/"],a[href*="/tv/"]').count(); stable=stable+1 if n==old else 0; old=n
        if stable>=5: break
        page.mouse.wheel(0,random.randint(900,1500)); pause(.8,1.5)
def saved_lists(page,root):
    page.goto(root,wait_until="domcontentloaded",timeout=90000); pause(2,4); scroll(page,80); out={}; root_path=re.sub(r'^https?://[^/]+','',root).rstrip('/')
    for a in page.locator('a[href*="/saved/"]').all():
        href=a.get_attribute("href") or ""; name=(a.inner_text() or "").strip()
        if not name or name.lower() in IGNORE or href.rstrip('/')==root_path: continue
        out[href if href.startswith("http") else "https://www.instagram.com"+href]=MERGE.get(name.lower(),name)
    return out
def scan(page,url):
    page.goto(url,wait_until="domcontentloaded",timeout=90000); pause(2,4)
    for _ in range(3):
        retry=page.get_by_text("Réessayer",exact=False)
        if retry.count(): retry.first.click(); pause(3,5)
        scroll(page); links=page.locator('a[href*="/p/"],a[href*="/reel/"],a[href*="/tv/"]')
        if links.count(): break
        page.reload(wait_until="domcontentloaded"); pause(3,5)
    out={}
    for a in links.all():
        href=(a.get_attribute("href") or "").split("?")[0]; url=href if href.startswith("http") else "https://www.instagram.com"+href
        out[url]=(a.get_attribute("aria-label") or a.inner_text() or "Publication Instagram").strip()[:250]
    return out
def add_items(wb,name,items):
    sn=sheet_name(name)
    if sn not in wb.sheetnames:
        ws=wb.create_sheet(sn); ws.append(ITEMS); dv=DataValidation(type="list",formula1='"pending,done,skip"'); ws.add_data_validation(dv); dv.add("D2:D10000")
    ws=wb[sn]; known={str(r[2].value) for r in ws.iter_rows(min_row=2) if r[2].value}; number=max([int(r[0].value) for r in ws.iter_rows(min_row=2) if str(r[0].value or '').isdigit()] or [0]); added=0
    for url,title in reversed(list(items.items())):
        if url in known: continue
        number+=1; added+=1; ws.append([number,title,url,"pending",datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    style(ws); return sn,number,added
def summary_set(ws,name,url,sn,number,count,state):
    row=next((i for i in range(2,ws.max_row+1) if str(ws.cell(i,1).value or '')==name),ws.max_row+1)
    for col,val in enumerate([name,url,sn,number,count,datetime.now().strftime("%Y-%m-%d %H:%M:%S"),state],1): ws.cell(row,col,val)

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--zone",required=True); ap.add_argument("--env",choices=["local","staging","prod"],default="prod"); ap.add_argument("--account",default="izilifehdf"); ap.add_argument("--init",action="store_true"); a=ap.parse_args()
    path=drive_workspace_root(a.env)/"izilife"/"instagram"/normalize_zone(a.zone)/"instagram_collections.xlsx"; wb=book(path); summary=wb["Résumé"]
    if a.init: style(summary); wb.save(path); print("Créé :",path); return
    sources={str(r[1].value):MERGE.get(str(r[0].value).strip().lower(),str(r[0].value).strip()) for r in summary.iter_rows(min_row=2) if r[0].value and r[1].value}
    profile=Path(os.getenv("AGENTIC_INSTAGRAM_PROFILE",Path.home()/".izilife"/"instagram-chrome-profile")); profile.mkdir(parents=True,exist_ok=True)
    with sync_playwright() as p:
        ctx=p.chromium.launch_persistent_context(str(profile),channel="chrome",headless=False,locale="fr-FR",timezone_id="Europe/Paris",slow_mo=150); page=ctx.pages[0] if ctx.pages else ctx.new_page(); root=f"https://www.instagram.com/{a.account}/saved/"; page.goto(root,wait_until="domcontentloaded",timeout=90000)
        if "/accounts/login" in page.url: input("Connecte-toi dans Chrome puis appuie sur ENTRÉE...")
        sources.update(saved_lists(page,root))
        for url,name in sources.items():
            if name.lower() in IGNORE: continue
            try: items=scan(page,url); sn,num,added=add_items(wb,name,items); summary_set(summary,name,url,sn,num,len(items),"OK" if items else "VIDE/ÉCHEC WEB"); print(f"{name}: {added} nouveau(x)")
            except Exception as e: summary_set(summary,name,url,sheet_name(name),0,0,"ERREUR: "+str(e)[:180]); print("ERREUR",name,e)
            style(summary); wb.save(path)
        ctx.close()

if __name__=="__main__":
    sys.path.insert(0,str(Path(__file__).resolve().parents[1])); from agent_excel_logger import run_logged; run_logged("instagram","saved-collections",main)
