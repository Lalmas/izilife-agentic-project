"""Agent Ambiance : Reels/liens vers ElementExternalMedia, sans Playwright."""
from __future__ import annotations
import argparse,json,re
from pathlib import Path
from urllib.parse import parse_qsl,urlencode,urlsplit,urlunsplit
import openpyxl
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from objects.object_agent_common import drive_workspace_root,izilife_post,load_env,normalize_env,normalize_zone

HEADERS=["TARGET_REF","TARGET_TYPE","SOURCE","LIEN_1","LIEN_2","LIEN_3","LIEN_4","LIEN_5","LIEN_6","LIEN_7","LIEN_8"]
TYPES=["PAGE","PLACE","SHOP","EVENT","EVENT_SERIE","ANNUAL_CELEBRATION","EXPERIENCE"]
SOURCES=["AUTO","instagram","tiktok","youtube","facebook","website"]
def zone_dir(zone,env):return drive_workspace_root(env)/"izilife"/"ambiance"/normalize_zone(zone)
def default_file(zone,env):return zone_dir(zone,env)/"ambiance.xlsx"
def init_sheet(zone,env):
    root=zone_dir(zone,normalize_env(env));root.mkdir(parents=True,exist_ok=True);path=root/"ambiance.xlsx"
    if path.exists():print(f"Existe déjà : {path}");return
    wb=openpyxl.Workbook();ws=wb.active;ws.title="Ambiances";ws.append(HEADERS);widths=[55,25,18]+[58]*8
    for i,(h,w) in enumerate(zip(HEADERS,widths),1):
        c=ws.cell(1,i);c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="7030A0");c.alignment=Alignment(horizontal="center");ws.column_dimensions[c.column_letter].width=w
    ws.freeze_panes="A2";ws.auto_filter.ref="A1:K500"
    dv=DataValidation(type="list",formula1='"'+','.join(TYPES)+'"',allow_blank=False);ws.add_data_validation(dv);dv.add("B2:B500")
    ds=DataValidation(type="list",formula1='"'+','.join(SOURCES)+'"');ws.add_data_validation(ds);ds.add("C2:C500")
    for r in range(2,501):
        if r%2==0:
            for c in ws[r]:c.fill=PatternFill("solid",fgColor="F3EAF8")
    wb.save(path);print(f"Créé : {path}")
def clean_url(v):
    v=str(v or "").strip()
    if not re.match(r"^https?://",v,re.I):return ""
    p=urlsplit(v);q=urlencode([(k,x) for k,x in parse_qsl(p.query) if not k.lower().startswith(("utm_","igsh"))]);return urlunsplit((p.scheme.lower(),p.netloc.lower(),p.path.rstrip("/") or "/",q,""))
def run(zone,env,file,dry_run,maximum):
    load_env();env=normalize_env(env);path=Path(file) if file else default_file(zone,env)
    if not path.exists():raise SystemExit(f"Fichier introuvable : {path}. Lance --init.")
    wb=openpyxl.load_workbook(path,data_only=False);total=created=skipped=errors=0
    for ws in wb.worksheets:
        headers={str(c.value or "").strip().upper():i for i,c in enumerate(ws[1])}
        for row in ws.iter_rows(min_row=2):
            def val(k):i=headers.get(k);return str(row[i].value or "").strip() if i is not None else ""
            ref=val("TARGET_REF");typ=val("TARGET_TYPE").upper()
            if not ref or not typ:continue
            links=[]
            for key,i in headers.items():
                if key.startswith("LIEN_"):
                    c=row[i];u=clean_url(c.hyperlink.target if c.hyperlink else c.value)
                    if u and u not in links:links.append(u)
            for order,url in enumerate(links):
                if total>=maximum:break
                payload={"target_ref":ref,"target_type":typ,"external_url":url,"display_order":order};total+=1
                if dry_run:print("[DRY RUN]",json.dumps(payload,ensure_ascii=False));continue
                result=izilife_post("/scraper/agentAddExternalMedia",{"payload":json.dumps(payload,ensure_ascii=False)},env)
                if result.get("success") and result.get("created"):created+=1;print(f"OK {typ} {ref} -> {url}")
                elif result.get("success") and result.get("skipped"):skipped+=1;print(f"SKIP déjà présent -> {url}")
                else:errors+=1;print(f"ERREUR {typ} {ref}: {result.get('error',result)}")
    print(f"Résultat: liens={total} ajoutés={created} ignorés={skipped} erreurs={errors}")
def main():
    ap=argparse.ArgumentParser();ap.add_argument("--zone",required=True);ap.add_argument("--env",default="prod",choices=["local","staging","prod"]);ap.add_argument("--file");ap.add_argument("--init",action="store_true");ap.add_argument("--dry-run",action="store_true");ap.add_argument("--max",type=int,default=1000);a=ap.parse_args();init_sheet(a.zone,a.env) if a.init else run(a.zone,a.env,a.file,a.dry_run,a.max)

if __name__=="__main__":
    from agent_excel_logger import run_logged
    run_logged("external_medias", "external_medias", main)
