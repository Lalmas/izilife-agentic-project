"""
createur_lieux.py
-----------------
Curation manuelle de lieux — lit curate_places.xlsx depuis Google Drive par zone
et appelle postAgentFetchAndStoreOnePlace pour chaque lieu pending.

Le PHP fait tout : recherche Google Places, vérifie existence, importe.
Ce script ne fait qu'orchestrer les appels et tenir le fichier à jour.

Usage :
    python createur_lieux.py --zone=lille  --env=local --init
    python createur_lieux.py --zone=lille --env=local --city=lille --dry-run
    python createur_lieux.py --zone=lille --env=local --city=lille

Prérequis :
    pip install requests openpyxl python-dotenv
    Variable : IZILIFE_AGENT_TOKEN

Fichier xlsx :
    G:/Mon Drive/agentic_workspace/izilife/places/{zone}-zone/curate_places.xlsx
    Colonnes :
        nom          : Nom du lieu (obligatoire)
        ville        : Slug ville izilife (ex: roubaix). Si vide → ville du --city
        type_impose  : SHOP ou PLACE (optionnel — laissé vide = Google décide)
        notes        : Tes notes perso
        statut       : pending → done | unmapped | error | skip
"""

import os
import sys
import time
import random
import argparse
import requests
from datetime import datetime, date
from pathlib import Path

# ─────────────────────────────────────────────
# CORE PATHS / CONFIG — migration safe
# ─────────────────────────────────────────────

def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return

_ensure_core_import_path()

try:
    from core.paths import IZILIFE_ENVS, ENV_GLOBAL, ENV_IZILIFE, izilife_places_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False

CURRENT_ENV = "prod"

def set_current_env(env_name: str):
    global CURRENT_ENV
    CURRENT_ENV = str(env_name or "prod").strip().lower()


try:
    from dotenv import load_dotenv
    if HAS_CORE_PATHS:
        for _env_file in (ENV_GLOBAL, ENV_IZILIFE, Path(__file__).parent / ".env"):
            if _env_file and _env_file.exists():
                load_dotenv(_env_file, override=False)
    else:
        load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────
# CHEMINS
# ─────────────────────────────────────────────

def get_drive_root() -> Path:
    if sys.platform == "win32":
        candidates = [Path("G:/Mon Drive"), Path("G:/My Drive"),
                      Path.home() / "Google Drive", Path.home() / "Mon Drive"]
    else:
        candidates = [Path.home() / "GoogleDrive", Path.home() / "Google Drive",
                      Path.home() / "gdrive", Path("/mnt/gdrive")]
    for p in candidates:
        if p.exists():
            return p
    fallback = Path(__file__).parent / "izilife-agent-workspace"
    fallback.mkdir(parents=True, exist_ok=True)
    print(f"⚠️  Google Drive non trouvé — fallback : {fallback}")
    return fallback


def get_curate_file(zone: str) -> Path:
    if HAS_CORE_PATHS:
        return izilife_places_zone(zone, CURRENT_ENV) / "curate_places.xlsx"
    return get_drive_root() / "agentic_workspace" / "izilife" / "places" / f"{zone}-zone" / "curate_places.xlsx"


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

ENVS = IZILIFE_ENVS if HAS_CORE_PATHS else {
    "local":   {"base_url": "https://localhost:4443/izilife-admin",          "verify_ssl": False},
    "staging": {"base_url": "https://www.staging.izilife.co/izilife-admin", "verify_ssl": True},
    "prod":    {"base_url": "https://www.izilife.co/izilife-admin",          "verify_ssl": True},
}

AGENT_TOKEN = os.environ.get("IZILIFE_AGENT_TOKEN", "METTRE_TOKEN_ICI")
DELAY       = (2, 4)

# ─────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def sleep_random(mn=2, mx=4):
    time.sleep(random.uniform(mn, mx))

def izilife_get(endpoint: str, env: dict) -> dict | None:
    try:
        r = requests.get(
            env["base_url"] + endpoint,
            headers={"X-Agent-Token": AGENT_TOKEN},
            verify=env["verify_ssl"], timeout=15,
        )
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ GET {endpoint} : {e}")
        return None

def izilife_post(endpoint: str, data: dict, env: dict) -> dict | None:
    try:
        r = requests.post(
            env["base_url"] + endpoint,
            data=data,
            headers={"X-Agent-Token": AGENT_TOKEN},
            verify=env["verify_ssl"], timeout=30,
        )
        if r.status_code == 200:
            return r.json()
        log(f"  ❌ HTTP {r.status_code} : {r.text[:200]}")
        return None
    except Exception as e:
        log(f"  ❌ POST {endpoint} : {e}")
        return None

def resolve_city_id(city_slug: str, env: dict) -> int:
    r = izilife_get(f"/scraper/cityByStringId/{city_slug}", env)
    if r and r.get("success") and r.get("city"):
        city_id = int(r["city"]["id"])
        log(f"Ville résolue : {city_slug} → city_id={city_id}")
        return city_id
    log(f"❌ Ville introuvable : {city_slug}")
    sys.exit(1)

# ─────────────────────────────────────────────
# LECTURE / ÉCRITURE XLSX
# ─────────────────────────────────────────────

def read_pending_rows(ws) -> list[dict]:
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)}
        if data.get("statut", "").lower() in ("done", "error", "skip", "unmapped"):
            continue
        if not data.get("nom", "").strip():
            continue
        data["row_idx"] = row_idx
        rows.append(data)
    return rows

def update_row(ws, row_idx: int, new_status: str):
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    try:
        col = headers.index("statut") + 1
        ws.cell(row=row_idx, column=col).value = new_status
    except ValueError:
        pass

# ─────────────────────────────────────────────
# TRAITEMENT D'UN LIEU
# ─────────────────────────────────────────────

def process_one(row: dict, default_city_slug: str, env: dict, dry_run: bool) -> str:
    nom         = row.get("nom", "").strip()
    ville       = row.get("ville", "").strip() or default_city_slug
    type_impose = row.get("type_impose", "").strip().upper()

    log(f"\n  → {nom} ({ville})")

    if dry_run:
        log(f"    [DRY RUN] → postAgentFetchAndStoreOnePlace/{ville} query={nom}")
        return "done"

    # Appel PHP — il cherche sur Google Places, vérifie existence, importe
    payload = {"query": nom}
    if type_impose in ("SHOP", "PLACE"):
        payload["destination"] = type_impose

    resp = izilife_post(f"/scraper/agentFetchAndStoreOnePlace/{ville}", payload, env)

    if not resp:
        log(f"    ❌ Pas de réponse du serveur")
        return "error"

    if resp.get("success"):
        result_type  = resp.get("result_type", "")
        name_created = resp.get("name", nom)
        if result_type == "unmapped":
            log(f"    ⚠️  Type non mappé → ScrapingUnmappedPoi : {name_created}")
            return "unmapped"
        log(f"    ✅ Créé : {name_created} [{result_type}] id={resp.get('entity_id')}")
        return "done"
    else:
        err = resp.get("error", "Erreur inconnue")
        if resp.get("already_exists"):
            log(f"    ℹ️  Déjà existant — skip")
            return "done"
        log(f"    ❌ {err}")
        return "error"

# ─────────────────────────────────────────────
# INIT — créer le template xlsx
# ─────────────────────────────────────────────

def create_template(zone: str):
    pfile = get_curate_file(zone)
    pfile.parent.mkdir(parents=True, exist_ok=True)

    if pfile.exists():
        log(f"⚠️  Fichier déjà existant : {pfile}")
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lieux"

    HDR_BG = "1F4E79"
    ALT_BG = "EBF3FB"
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["NOM", "VILLE", "TYPE_IMPOSE", "NOTES", "STATUT"]
    widths  = [42,    18,      14,             38,      14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    exemples = [
        ["Big Luck",                 "lille",  "SHOP",  "Bar cocktails Vieux-Lille", "pending"],
        ["Parc Jean-Baptiste Lebas", "lille",  "PLACE", "",                          "pending"],
        ["Le Barafine",              "roubaix","SHOP",  "",                          "pending"],
        ["",                         "",       "",      "",                          "pending"],
        ["",                         "",       "",      "",                          "pending"],
    ]
    for r, row_data in enumerate(exemples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor=ALT_BG if r % 2 == 0 else "FFFFFF")
        ws.row_dimensions[r].height = 18

    # Dropdown type_impose
    dv_type = DataValidation(type="list", formula1='"SHOP,PLACE,EQUIPMENT"',
                             allow_blank=True, showDropDown=False)
    dv_type.sqref = "C2:C500"
    ws.add_data_validation(dv_type)

    # Dropdown statut
    dv_status = DataValidation(type="list",
                               formula1='"pending,done,error,unmapped,skip"',
                               allow_blank=False, showDropDown=False)
    dv_status.sqref = "E2:E500"
    ws.add_data_validation(dv_status)

    # Mise en forme conditionnelle par statut
    for status, color in [("done","C6EFCE"),("error","FFC7CE"),
                          ("unmapped","FFEB9C"),("skip","D9D9D9")]:
        ws.conditional_formatting.add("A2:E500", FormulaRule(
            formula=[f'$E2="{status}"'],
            fill=PatternFill("solid", fgColor=color)
        ))

    ws.freeze_panes = "A2"

    wb.save(pfile)
    log(f"✅ Template créé : {pfile}")
    log(f"   Remplis le fichier puis lance :")
    log(f"   python createur_lieux.py --zone={zone} --env=local --city=<ville_par_defaut>")


# ─────────────────────────────────────────────
# ARCHIVE AUTO — déplacer les done vers logs/WXX/
# ─────────────────────────────────────────────

def get_log_file(zone: str, script: str = "createur_lieux") -> Path:
    today = date.today()
    week  = today.strftime("W%V")
    d = get_curate_file(zone).parent / "logs" / week
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{script}_{today.isoformat()}.xlsx"


def purge_done(zone: str) -> int:
    """Déplace automatiquement les lignes statut=done vers logs/WXX/createur_lieux_YYYY-MM-DD.xlsx.

    Convention : on n'ajoute pas d'onglet Log dans curate_places.xlsx.
    Le fichier principal garde uniquement les lignes à suivre : pending, error, unmapped, skip.
    Retourne le nombre de lignes archivées.
    """
    pfile = get_curate_file(zone)
    if not pfile.exists():
        log("Aucun fichier à purger.")
        return 0

    for _attempt in range(3):
        try:
            wb = openpyxl.load_workbook(pfile)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            if not headers:
                log("Aucun header dans le fichier principal.")
                return 0

            stat_col = (headers.index("statut") + 1) if "statut" in headers else 5
            done_rows = []
            keep_rows = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_values = list(row)
                if not any(v is not None and str(v).strip() != "" for v in row_values):
                    continue
                status = str(row_values[stat_col - 1] or "").strip().lower()
                if status == "done":
                    done_rows.append(row_values)
                else:
                    keep_rows.append(row_values)

            if not done_rows:
                log("Aucune ligne done à archiver.")
                return 0

            # Écrire le log externe
            lf = get_log_file(zone, "createur_lieux")
            lb = openpyxl.Workbook()
            lws = lb.active
            lws.title = "Done"
            for col, h in enumerate(headers, 1):
                lws.cell(row=1, column=col, value=str(h).upper())
            for r, row in enumerate(done_rows, 2):
                for c, val in enumerate(row, 1):
                    lws.cell(row=r, column=c, value=val)
            lb.save(lf)
            log(f"✅ {len(done_rows)} ligne(s) done → {lf}")

            # Réécrire le fichier principal sans les done
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            for r, row in enumerate(keep_rows, 2):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=val)
            wb.save(pfile)
            log(f"✅ Fichier principal nettoyé — {len(keep_rows)} ligne(s) restante(s)")
            return len(done_rows)

        except PermissionError:
            time.sleep(3)

    log("❌ Fichier verrouillé — archivage impossible")
    return 0

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Créateur de lieux izilife — curation manuelle")
    parser.add_argument("--zone",    type=str, required=True,
                        help="Zone cible : lille, valenciennes, paris ...")
    parser.add_argument("--env",     choices=ENVS.keys(), default="local")
    parser.add_argument("--city",    type=str, default=None,
                        help="Ville par défaut si colonne 'ville' vide dans le xlsx")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--init",    action="store_true",
                        help="Créer le fichier xlsx template pour la zone")
    args = parser.parse_args()
    set_current_env(args.env)

    if args.init:
        create_template(args.zone)
        return


    if not args.city:
        print("❌ --city obligatoire sauf avec --init")
        sys.exit(1)

    env     = ENVS[args.env]
    zone    = args.zone.lower().strip()
    dry_run = args.dry_run

    log(f"=== createur_lieux.py — zone={zone} env={args.env} city={args.city}" +
        (" [DRY RUN]" if dry_run else "") + " ===")

    if AGENT_TOKEN == "METTRE_TOKEN_ICI" and not dry_run:
        log("❌ IZILIFE_AGENT_TOKEN non défini.")
        sys.exit(1)

    curate_file = get_curate_file(zone)
    if not curate_file.exists():
        log(f"❌ Fichier introuvable : {curate_file}")
        log(f"   Lance : python createur_lieux.py --zone={zone} --init")
        sys.exit(1)

    wb = openpyxl.load_workbook(curate_file)
    ws = wb.active
    rows = read_pending_rows(ws)

    if not rows:
        log("✅ Aucun lieu pending.")
        return

    log(f"{len(rows)} lieu(x) à traiter")
    stats = {"done": 0, "unmapped": 0, "errors": 0}

    for row in rows:
        new_status = process_one(row, args.city, env, dry_run)
        if not dry_run:
            update_row(ws, row["row_idx"], new_status)
            wb.save(curate_file)
        stats[new_status if new_status in stats else "errors"] += 1
        sleep_random(*DELAY)

    # Archivage automatique des done vers logs/WXX/ après traitement
    if not dry_run:
        purge_done(zone)

    log(f"\n=== Résultat ===")
    log(f"  Créés      : {stats['done']}")
    log(f"  Unmapped   : {stats['unmapped']}  (à traiter dans le BO)")
    log(f"  Erreurs    : {stats['errors']}")


if __name__ == "__main__":
    main()