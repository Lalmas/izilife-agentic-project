"""
chasseur_lieux.py
-----------------
Agent chasseur de lieux — parcourt Google Maps par ville × catégorie
et stocke les lieux trouvés dans izilife_villes.xlsx pour review puis import.

Deux modes :
  --collect  : parcourt Google Maps, stocke les inconnus dans l'onglet "À créer"
  --insert   : prend les lignes "pending" de l'onglet "À créer" → appelle
               postAgentFetchAndStoreOnePlace pour chacun
  (défaut : les deux à la suite)

Usage :
    python chasseur_lieux.py --zone=lille --env=local --dry-run
    python chasseur_lieux.py --zone=lille --env=local --city=lille
    python chasseur_lieux.py --zone=lille --env=local --city=lille --collect-only
    python chasseur_lieux.py --zone=lille --env=local --city=lille --insert-only
    python chasseur_lieux.py --zone=lille --env=local --city=lille --max-per-cat=30
    python chasseur_lieux.py --zone=lille --env=local --init

Prérequis :
    pip install playwright requests playwright-stealth openpyxl python-dotenv
    python -m playwright install chromium
    Variable : IZILIFE_AGENT_TOKEN

Fichiers Drive (zone) :
    $AGENTIC_DRIVE_ROOT/agentic_workspace/izilife/places/{zone}-zone/
        izilife_villes.xlsx     ← Sheet principal (villes + statuts)
        categories_lieux.json   ← Liste des catégories à chasser (auto-créé si absent)

izilife_villes.xlsx — onglets :
    Villes       : ville_slug | nom_affichage | prio | tags | chasseur_statut |
                   chasseur_date | enrichisseur_statut | notes
    À créer      : nom | adresse | ville_slug | google_place_id | categorie_source |
                   google_types | statut | date_trouve
    Catégories   : categorie | label_fr | actif (copie de categories_lieux.json)
"""

from __future__ import annotations

import os
import sys
import re
import json
import time
import random
import argparse
import requests
from datetime import datetime, date
from pathlib import Path

# ─────────────────────────────────────────────
# CORE PATHS / CONFIG — migration safe
# ─────────────────────────────────────────────

def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return

_ensure_core_import_path()

try:
    from core.paths import IZILIFE_ENVS, ENV_GLOBAL, ENV_IZILIFE, izilife_places_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False

CURRENT_ENV = "prod"

def set_current_env(env_name: str):
    global CURRENT_ENV
    CURRENT_ENV = str(env_name or "prod").strip().lower()


try:
    from dotenv import load_dotenv
    if HAS_CORE_PATHS:
        for _env_file in (ENV_GLOBAL, ENV_IZILIFE, Path(__file__).parent / ".env"):
            if _env_file and _env_file.exists():
                load_dotenv(_env_file, override=False)
    else:
        load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    print("pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

# ─────────────────────────────────────────────
# CHEMINS
# ─────────────────────────────────────────────

def get_drive_root() -> Path:
    value = os.environ.get("AGENTIC_DRIVE_ROOT", "").strip()
    if not value: raise RuntimeError("AGENTIC_DRIVE_ROOT non défini.")
    return Path(value).expanduser()


def get_zone_dir(zone: str) -> Path:
    if HAS_CORE_PATHS:
        d = izilife_places_zone(zone, CURRENT_ENV)
    else:
        d = get_drive_root() / "agentic_workspace" / "izilife" / "places" / f"{zone}-zone"
    d.mkdir(parents=True, exist_ok=True)
    return d

def get_villes_file(zone: str) -> Path:
    return get_zone_dir(zone) / "izilife_villes.xlsx"

def get_categories_file(zone: str) -> Path:
    return get_zone_dir(zone) / "categories_lieux.json"

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

ENVS = IZILIFE_ENVS if HAS_CORE_PATHS else {
    "local":   {"base_url": "https://localhost:4443/izilife-admin",          "verify_ssl": False},
    "staging": {"base_url": "https://www.staging.izilife.co/izilife-admin", "verify_ssl": True},
    "prod":    {"base_url": "https://www.izilife.co/izilife-admin",          "verify_ssl": True},
}

AGENT_TOKEN = os.environ.get("IZILIFE_AGENT_TOKEN", "METTRE_TOKEN_ICI")

# Délais
DELAY_BETWEEN_RESULTS = (1, 2)
DELAY_BETWEEN_CATS    = (4, 8)
DELAY_SCROLL          = (2, 3)

# Catégories par défaut (créées si categories_lieux.json absent)
# Ces catégories sont des requêtes pour Google Maps (texte tapé dans la recherche)
DEFAULT_CATEGORIES = [
    # 1. NUIT / SOCIAL
    {"categorie": "bar", "label_fr": "Bars", "actif": True},
    {"categorie": "bar à cocktails", "label_fr": "Bars à cocktails", "actif": True},
    {"categorie": "bar à bière", "label_fr": "Bars à bière", "actif": True},
    {"categorie": "bar à vin", "label_fr": "Bars à vin", "actif": True},
    {"categorie": "pub", "label_fr": "Pubs", "actif": True},
    {"categorie": "bar à jeux", "label_fr": "Bars à jeux", "actif": True},
    {"categorie": "bar karaoké", "label_fr": "Bars karaoké", "actif": True},
    {"categorie": "bar lounge", "label_fr": "Bars lounge", "actif": True},
    {"categorie": "discothèque boîte de nuit club", "label_fr": "Clubs / Boîtes", "actif": True},

    # 2. RESTAURATION
    {"categorie": "restaurant", "label_fr": "Restaurants", "actif": True},
    {"categorie": "brasserie", "label_fr": "Brasseries", "actif": True},
    {"categorie": "bistro", "label_fr": "Bistros", "actif": True},
    {"categorie": "restaurant traditionnel", "label_fr": "Restos traditionnels", "actif": True},
    {"categorie": "restaurant gastronomique", "label_fr": "Restos gastronomiques", "actif": True},
    {"categorie": "restaurant italien", "label_fr": "Restos italiens", "actif": True},
    {"categorie": "restaurant japonais sushi", "label_fr": "Sushis", "actif": True},
    {"categorie": "pizzeria pizza", "label_fr": "Pizzas", "actif": True},
    {"categorie": "burger restaurant", "label_fr": "Burgers", "actif": True},
    {"categorie": "kebab", "label_fr": "Kebabs", "actif": True},
    {"categorie": "tacos", "label_fr": "Tacos", "actif": True},
    {"categorie": "friterie", "label_fr": "Friteries", "actif": True},
    {"categorie": "sandwicherie", "label_fr": "Sandwicheries", "actif": True},
    {"categorie": "food truck", "label_fr": "Food trucks", "actif": True},
    {"categorie": "food court", "label_fr": "Food courts", "actif": True},

    # 3. CAFÉ / DOUCEURS
    {"categorie": "café coffee shop", "label_fr": "Cafés", "actif": True},
    {"categorie": "salon de thé", "label_fr": "Salons de thé", "actif": True},
    {"categorie": "boulangerie", "label_fr": "Boulangeries", "actif": True},
    {"categorie": "pâtisserie", "label_fr": "Pâtisseries", "actif": True},
    {"categorie": "chocolaterie", "label_fr": "Chocolateries", "actif": True},
    {"categorie": "glacier marchand de glace", "label_fr": "Glaciers", "actif": True},
    {"categorie": "crêperie", "label_fr": "Crêperies", "actif": True},

    # 4. LOISIRS PAYANTS / RÉSERVABLES
    {"categorie": "bowling", "label_fr": "Bowlings", "actif": True},
    {"categorie": "billard salle de billard", "label_fr": "Billards", "actif": True},
    {"categorie": "escape game", "label_fr": "Escape games", "actif": True},
    {"categorie": "laser game", "label_fr": "Laser games", "actif": True},
    {"categorie": "karting", "label_fr": "Kartings", "actif": True},
    {"categorie": "paintball", "label_fr": "Paintballs", "actif": True},
    {"categorie": "réalité virtuelle VR", "label_fr": "Réalité virtuelle", "actif": True},
    {"categorie": "salle d'arcade", "label_fr": "Salles d'arcade", "actif": True},
    {"categorie": "karaoké", "label_fr": "Karaokés", "actif": True},
    {"categorie": "mini golf", "label_fr": "Mini-golfs", "actif": True},
    {"categorie": "salle d'escalade", "label_fr": "Escalade", "actif": True},

    # 5. CULTURE / SORTIES
    {"categorie": "cinéma", "label_fr": "Cinémas", "actif": True},
    {"categorie": "théâtre", "label_fr": "Théâtres", "actif": True},
    {"categorie": "café théâtre", "label_fr": "Cafés-théâtres", "actif": True},
    {"categorie": "salle de spectacle concert", "label_fr": "Salles de spectacle", "actif": True},
    {"categorie": "musée", "label_fr": "Musées", "actif": True},
    {"categorie": "galerie d'art", "label_fr": "Galeries d'art", "actif": True},
    {"categorie": "bibliothèque médiathèque", "label_fr": "Bibliothèques / Médiathèques", "actif": True},

    # 6. VISITES / TOURISME LOCAL
    {"categorie": "ferme pédagogique", "label_fr": "Fermes pédagogiques", "actif": True},
    {"categorie": "ferme à visiter", "label_fr": "Fermes à visiter", "actif": True},
    {"categorie": "zoo parc animalier", "label_fr": "Zoos / Parcs animaliers", "actif": True},
    {"categorie": "aquarium", "label_fr": "Aquariums", "actif": True},
    {"categorie": "château à visiter", "label_fr": "Châteaux", "actif": True},
    {"categorie": "monument historique à visiter", "label_fr": "Monuments historiques", "actif": True},
    {"categorie": "site touristique", "label_fr": "Sites touristiques", "actif": True},
    {"categorie": "jardin remarquable", "label_fr": "Jardins remarquables", "actif": True},
    {"categorie": "visite guidée", "label_fr": "Visites guidées", "actif": True},

    # 7. LIEUX STRUCTURANTS DE VILLE
    {"categorie": "marché", "label_fr": "Marchés", "actif": True},
    {"categorie": "halle alimentaire", "label_fr": "Halles", "actif": True},
    {"categorie": "centre commercial", "label_fr": "Centres commerciaux", "actif": True},
    {"categorie": "place célèbre", "label_fr": "Places célèbres", "actif": True},
    {"categorie": "parc jardin public", "label_fr": "Parcs / Jardins", "actif": True},

    # 8. SHOPPING UTILE
    {"categorie": "librairie", "label_fr": "Librairies", "actif": True},
    {"categorie": "magasin de jeux jeux de société", "label_fr": "Magasins de jeux", "actif": True},
    {"categorie": "magasin de jeux vidéos", "label_fr": "Jeux vidéo", "actif": True},
    {"categorie": "magasin de jouets", "label_fr": "Jouets", "actif": True},
    {"categorie": "magasin de vêtements", "label_fr": "Vêtements", "actif": True},
    {"categorie": "magasin de sport", "label_fr": "Articles de sport", "actif": True},
    {"categorie": "caviste cave à vin", "label_fr": "Cavistes", "actif": True},
    {"categorie": "épicerie fine", "label_fr": "Épiceries fines", "actif": True},

    # 9. SPORT / BIEN-ÊTRE
    {"categorie": "salle de sport fitness", "label_fr": "Salles de sport", "actif": True},
    {"categorie": "piscine", "label_fr": "Piscines", "actif": True},
    {"categorie": "spa bien-être", "label_fr": "Spas", "actif": True},
    {"categorie": "hammam", "label_fr": "Hammams", "actif": True},
    {"categorie": "institut de beauté", "label_fr": "Instituts de beauté", "actif": True},
    {"categorie": "barbier", "label_fr": "Barbiers", "actif": True},
    {"categorie": "salon de coiffure", "label_fr": "Coiffeurs", "actif": True},

    # 10. PLEIN AIR / FAMILLE
    {"categorie": "aire de jeux", "label_fr": "Aires de jeux", "actif": True},
    {"categorie": "parc de loisirs", "label_fr": "Parcs de loisirs", "actif": True},
    {"categorie": "parc d'attractions", "label_fr": "Parcs d'attractions", "actif": True},
    {"categorie": "base de loisirs", "label_fr": "Bases de loisirs", "actif": True},

    # 11. DERNIÈRE PASSE — à activer plus tard
    {"categorie": "parking", "label_fr": "Parkings", "actif": False},
    {"categorie": "toilettes publiques", "label_fr": "Toilettes publiques", "actif": False},
    {"categorie": "monument statue", "label_fr": "Monuments / Statues", "actif": False},
    {"categorie": "_autre", "label_fr": "Autre (non catégorisé)", "actif": False},
]

# ─────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def sleep_random(mn=2, mx=5):
    time.sleep(random.uniform(mn, mx))

def to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(str(value).replace(",", ".").strip()))
    except Exception:
        return default

def apply_stealth(page):
    if HAS_STEALTH:
        stealth_sync(page)
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['fr-FR', 'fr']});
        window.chrome = { runtime: {} };
    """)

def izilife_get(endpoint: str, env: dict) -> dict | None:
    try:
        r = requests.get(env["base_url"] + endpoint,
                         headers={"X-Agent-Token": AGENT_TOKEN},
                         verify=env["verify_ssl"], timeout=15)
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ GET {endpoint} : {e}")
        return None

def izilife_post(endpoint: str, data: dict, env: dict) -> dict | None:
    try:
        r = requests.post(env["base_url"] + endpoint, data=data,
                          headers={"X-Agent-Token": AGENT_TOKEN},
                          verify=env["verify_ssl"], timeout=30)
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ POST {endpoint} : {e}")
        return None

def resolve_city_id(city_slug: str, env: dict) -> int:
    r = izilife_get(f"/scraper/cityByStringId/{city_slug}", env)
    if r and r.get("success") and r.get("city"):
        return int(r["city"]["id"])
    log(f"❌ Ville introuvable : {city_slug}")
    sys.exit(1)

# ─────────────────────────────────────────────
# CATÉGORIES
# ─────────────────────────────────────────────

def load_categories(zone: str) -> list[dict]:
    cfile = get_categories_file(zone)
    if cfile.exists():
        with open(cfile, "r", encoding="utf-8") as f:
            cats = json.load(f)
        return [c for c in cats if c.get("actif") and c.get("categorie") != "_autre"]
    return [c for c in DEFAULT_CATEGORIES if c.get("actif") and c.get("categorie") != "_autre"]

# ─────────────────────────────────────────────
# PLAYWRIGHT — collecte Google Maps
# ─────────────────────────────────────────────

class GoogleMapsConsentError(RuntimeError):
    """Google Maps reste bloque sur l'ecran de consentement."""


def accept_google_cookies(page):
    """Refuse le consentement Google puis continue vers Maps."""
    selectors = [
        'button:has-text("Tout refuser")', 'button:has-text("Reject all")',
        'button:has-text("Refuser tout")', 'input[value="Tout refuser"]',
        'input[value="Reject all"]', '[aria-label*="Tout refuser" i]',
        '[aria-label*="Reject all" i]', 'form:has-text("Tout refuser") button',
        'button:has-text("Tout accepter")', 'button:has-text("Accept all")',
        'button:has-text("Accepter tout")', 'input[value="Tout accepter"]',
        'input[value="Accept all"]', '[aria-label*="Tout accepter" i]',
        '[aria-label*="Accept all" i]', 'form:has-text("Tout accepter") button',
        'form[action*="consent"] button[jsname]',
    ]
    for _ in range(3):
        for sel in selectors:
            try:
                btn = page.locator(sel).last
                if btn.is_visible(timeout=1000):
                    btn.click(force=True)
                    page.wait_for_timeout(1200)
                    if "consent.google." not in page.url:
                        log("  🍪 Consentement Google traité")
                        return True
            except Exception:
                pass
        page.wait_for_timeout(500)
    return "consent.google." not in page.url


def search_google_maps(page, query: str, max_results: int = 40) -> list[dict]:
    """
    Lance une recherche sur Google Maps et retourne la liste des lieux trouvés.
    Chaque lieu : {nom, adresse, google_place_id, rating, types_raw}
    """
    results = []
    search_url = f"https://www.google.com/maps/search/{requests.utils.quote(query)}"

    try:
        page.goto(search_url, timeout=30000)
        sleep_random(2, 3)

        # Accepter cookies si popup présente
        if not accept_google_cookies(page):
            raise GoogleMapsConsentError("Ecran de consentement Google non ferme")
        sleep_random(1, 2)

        # Fermer les éventuelles autres popups
        try:
            page.keyboard.press("Escape")
            sleep_random(0.5, 1)
        except:
            pass

        # Scroller la liste des résultats pour charger plus
        scrolled = 0
        prev_count = 0
        while len(results) < max_results and scrolled < 15:
            # Extraire les résultats visibles
            raw = page.evaluate("""
                () => {
                    const items = [];
                    // Les liens de fiches dans le panneau de résultats
                    document.querySelectorAll('a[href*="/maps/place/"]').forEach(a => {
                        const href = a.href || '';
                        // Extraire le place_id depuis l'URL si disponible
                        const placeMatch = href.match(/place\\/[^/]+\\/@[\\d.,]+\\/[^/]+\\/([^/?]+)/);
                        const cidMatch   = href.match(/0x[0-9a-fA-F]+:0x[0-9a-fA-F]+/);

                        // Nom : h3 dans le lien, ou aria-label
                        const h3 = a.querySelector('h3');
                        const nom = h3 ? h3.textContent.trim()
                                       : (a.getAttribute('aria-label') || '').trim();

                        // Adresse : span juste après le nom
                        const spans = a.querySelectorAll('span');
                        let adresse = '';
                        spans.forEach(s => {
                            const t = s.textContent.trim();
                            if (t.match(/\\d/) && t.length > 5) adresse = t;
                        });

                        if (nom && href.includes('/maps/place/')) {
                            items.push({
                                nom:       nom,
                                adresse:   adresse,
                                url:       href,
                                place_key: cidMatch ? cidMatch[0] : (placeMatch ? placeMatch[1] : ''),
                            });
                        }
                    });
                    // Dédupliquer par nom
                    const seen = new Set();
                    return items.filter(i => {
                        if (!i.nom || seen.has(i.nom)) return false;
                        seen.add(i.nom);
                        return true;
                    });
                }
            """)

            # Ajouter les nouveaux
            existing_names = {r["nom"] for r in results}
            for item in (raw or []):
                if item["nom"] and item["nom"] not in existing_names:
                    results.append({
                        "nom":             item["nom"],
                        "adresse":         item.get("adresse", ""),
                        "maps_url":        item.get("url", ""),
                        "place_key":       item.get("place_key", ""),
                        "google_place_id": "",  # sera résolu si besoin
                    })
                    existing_names.add(item["nom"])

            if len(results) == prev_count:
                # Rien de nouveau — on a probablement tout
                break
            prev_count = len(results)

            # Scroller le panneau gauche
            try:
                panel = page.locator('[role="feed"]').first
                if panel.is_visible(timeout=1000):
                    panel.evaluate("el => el.scrollTop += 800")
                else:
                    page.evaluate("window.scrollBy(0, 800)")
            except:
                page.evaluate("window.scrollBy(0, 800)")

            sleep_random(*DELAY_SCROLL)
            scrolled += 1

        log(f"    → {len(results)} lieux trouvés sur Maps pour '{query}'")

    except GoogleMapsConsentError:
        raise
    except Exception as e:
        log(f"    ❌ Erreur Maps : {e}")

    return results[:max_results]


# ─────────────────────────────────────────────
# EXCEL — lecture / écriture
# ─────────────────────────────────────────────

VILLES_COLS   = ["ville_slug", "nom_affichage", "prio", "tags",
                 "chasseur_statut", "chasseur_date", "chasseur_cat_offset",
                 "enrichisseur_statut", "notes"]
A_CREER_COLS  = ["nom", "adresse", "ville_slug", "google_place_id",
                 "maps_url", "categorie_source", "statut", "date_trouve"]
CATS_COLS     = ["categorie", "label_fr", "actif"]


def _header_row(ws, cols: list[str], widths: list[int] = None):
    from openpyxl.styles import Alignment, Border, Side
    HDR_BG = "1F4E79"
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col)
        cell.value     = h.upper()
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        if widths and col <= len(widths):
            ws.column_dimensions[cell.column_letter].width = widths[col-1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _style_data_row(ws, row_idx: int, ncols: int, alt: bool = False):
    from openpyxl.styles import Alignment, Border, Side
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bg = "EBF3FB" if alt else "FFFFFF"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        if cell.value is None:
            cell.value = ""
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = border
        cell.fill      = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[row_idx].height = 18


def _add_dropdown(ws, col_letter: str, formula: str, max_row: int = 500):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=formula,
                        allow_blank=True, showDropDown=False,
                        sqref=f"{col_letter}2:{col_letter}{max_row}")
    ws.data_validations.append(dv)


def _add_conditional_colors(ws, cell_range: str, status_col: str,
                             statuses: dict):
    from openpyxl.formatting.rule import FormulaRule
    for status, color in statuses.items():
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'${status_col}2="{status}"'],
                        fill=PatternFill("solid", fgColor=color))
        )


def create_villes_xlsx(zone: str):
    """Crée izilife_villes.xlsx avec ses 3 onglets si absent."""
    vfile = get_villes_file(zone)
    vfile.parent.mkdir(parents=True, exist_ok=True)

    if vfile.exists():
        log(f"⚠️  Fichier déjà existant : {vfile}")
        return vfile

    wb = openpyxl.Workbook()

    STATUS_COLORS = {"done":"C6EFCE","collecté":"C6EFCE","error":"FFC7CE",
                     "skip":"D9D9D9","pending":"FFFFFF"}
    A_CREER_COLORS = {"done":"C6EFCE","error":"FFC7CE","unmapped":"FFEB9C",
                      "skip":"D9D9D9"}

    # ── Onglet Villes ──────────────────────────────────────────────
    ws_v = wb.active
    ws_v.title = "Villes"
    _header_row(ws_v, VILLES_COLS, widths=[22, 22, 6, 28, 16, 14, 8, 18, 25])
    examples = [
        ["lille",           "Lille",             1, "capitale_regionale", "pending", "", 0, "", ""],
        ["roubaix",         "Roubaix",           2, "",                   "pending", "", 0, "", ""],
        ["tourcoing",       "Tourcoing",         3, "",                   "pending", "", 0, "", ""],
        ["villeneuve-ascq", "Villeneuve-d'Ascq", 4, "",                   "pending", "", 0, "", ""],
        ["marcq-en-baroeul","Marcq-en-Baroeul",  5, "",                   "pending", "", 0, "", ""],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            ws_v.cell(row=r, column=c).value = val
        _style_data_row(ws_v, r, len(VILLES_COLS), alt=(r % 2 == 0))
    # 5 lignes vides supplémentaires pour que l'utilisateur puisse ajouter des villes
    for r in range(len(examples) + 2, len(examples) + 7):
        _style_data_row(ws_v, r, len(VILLES_COLS), alt=(r % 2 == 0))
    # Dropdown chasseur_statut (col E=5) et enrichisseur_statut (col G=7)
    _add_dropdown(ws_v, "E", '"pending,done,collecté,skip"')
    _add_dropdown(ws_v, "G", '"pending,done,skip"')
    _add_conditional_colors(ws_v, f"A2:H500", "E", STATUS_COLORS)

    # ── Onglet À créer ─────────────────────────────────────────────
    ws_c = wb.create_sheet("À créer")
    _header_row(ws_c, A_CREER_COLS, widths=[38, 38, 18, 32, 55, 22, 12, 12])
    _add_dropdown(ws_c, "G", '"pending,done,error,unmapped,skip"', max_row=2000)
    _add_conditional_colors(ws_c, "A2:H2000", "G", A_CREER_COLORS)
    # Pas de lignes vides — append() ajoutera les lignes au fur et à mesure

    # ── Onglet Catégories ──────────────────────────────────────────
    ws_cats = wb.create_sheet("Catégories")
    _header_row(ws_cats, CATS_COLS, widths=[38, 28, 8])
    all_cats = DEFAULT_CATEGORIES
    for r, cat in enumerate(all_cats, 2):
        ws_cats.cell(row=r, column=1).value = cat["categorie"]
        ws_cats.cell(row=r, column=2).value = cat["label_fr"]
        ws_cats.cell(row=r, column=3).value = "oui" if cat.get("actif") else "non"
        _style_data_row(ws_cats, r, 3, alt=(r % 2 == 0))
    _add_dropdown(ws_cats, "C", '"oui,non"')

    wb.save(vfile)
    log(f"✅ izilife_villes.xlsx créé : {vfile}")

    # Sauvegarder aussi categories_lieux.json
    cfile = get_categories_file(zone)
    if not cfile.exists():
        with open(cfile, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CATEGORIES, f, ensure_ascii=False, indent=2)
        log(f"✅ categories_lieux.json créé : {cfile}")

    return vfile


def read_villes_pending(vfile: Path) -> list[dict]:
    wb = openpyxl.load_workbook(vfile)
    if "Villes" not in wb.sheetnames:
        return []
    ws = wb["Villes"]
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)}
        if data.get("chasseur_statut", "").lower() in ("done", "skip"):
            continue
        if not data.get("ville_slug"):
            continue
        data["row_idx"] = row_idx
        rows.append(data)
    return rows


def read_a_creer_pending(vfile: Path) -> list[dict]:
    wb = openpyxl.load_workbook(vfile)
    if "À créer" not in wb.sheetnames:
        return []
    ws = wb["À créer"]
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)}
        if data.get("statut", "").lower() in ("done", "error", "skip", "unmapped"):
            continue
        if not data.get("nom"):
            continue
        data["row_idx"] = row_idx
        rows.append(data)
    return rows


def _get_sheet(wb, name: str):
    """Trouve un onglet par nom insensible à la casse et aux variantes d'accent."""
    name_clean = name.lower().replace("à", "a").replace("é", "e").replace("è", "e")
    for sh in wb.sheetnames:
        sh_clean = sh.lower().replace("à", "a").replace("é", "e").replace("è", "e")
        if sh_clean == name_clean:
            return wb[sh]
    return None


def append_to_a_creer(vfile: Path, items: list[dict]):
    """Ajoute les nouveaux lieux à l'onglet À créer sans doublons sur nom+ville."""
    wb = openpyxl.load_workbook(vfile)

    # Chercher l'onglet — log les noms disponibles pour debug
    log(f"    [xlsx] Onglets disponibles : {wb.sheetnames}")
    ws = _get_sheet(wb, "À créer")

    if ws is None:
        # Onglet absent — on le crée
        log(f"    [xlsx] Onglet 'À créer' absent → création")
        ws = wb.create_sheet("À créer")
        for col, h in enumerate(A_CREER_COLS, 1):
            ws.cell(row=1, column=col).value = h.upper()
    else:
        log(f"    [xlsx] Onglet trouvé : '{ws.title}'")

    # Colonnes par position fixe (NOM=0, ADRESSE=1, VILLE_SLUG=2, ...)
    # On n'utilise pas headers.index() pour éviter les soucis d'encodage
    NOM_COL   = 0  # col A
    VILLE_COL = 2  # col C

    # Doublons existants
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(NOM_COL, VILLE_COL):
            continue
        nom   = str(row[NOM_COL]   or "").strip().lower()
        ville = str(row[VILLE_COL] or "").strip().lower()
        if nom:
            existing.add((nom, ville))

    log(f"    [xlsx] {len(existing)} doublons existants, {len(items)} items à ajouter")

    added = 0
    for item in items:
        nom   = item.get("nom", "").strip()
        ville = item.get("ville_slug", "").strip()
        key   = (nom.lower(), ville.lower())
        if not nom or key in existing:
            continue
        ws.append([
            nom,
            item.get("adresse", ""),
            ville,
            item.get("google_place_id", ""),
            item.get("maps_url", ""),
            item.get("categorie_source", ""),
            "pending",
            str(date.today()),
        ])
        existing.add(key)
        added += 1

    log(f"    [xlsx] Sauvegarde → {added} lignes ajoutées dans {vfile.name}")
    wb.save(vfile)
    return added


def update_a_creer_status(vfile: Path, row_idx: int, new_status: str):
    import time as _t
    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(vfile)
            ws = _get_sheet(wb, "À créer") or wb[wb.sheetnames[1]]
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            try:
                col = headers.index("statut") + 1
            except ValueError:
                col = 7  # position fixe colonne STATUT
            ws.cell(row=row_idx, column=col).value = new_status
            wb.save(vfile)
            return
        except PermissionError:
            if attempt < 4:
                log(f"    ⚠️  Fichier xlsx verrouillé (Drive sync) — retry {attempt+1}/5")
                _t.sleep(3)
            else:
                log(f"    ❌ Impossible de sauvegarder xlsx après 5 essais — statut non mis à jour")


def update_ville_statut(vfile: Path, row_idx: int, statut: str, cat_offset: int = None):
    wb = openpyxl.load_workbook(vfile)
    ws = _get_sheet(wb, "Villes") or wb["Villes"]
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    try:
        col_s = headers.index("chasseur_statut") + 1
        col_d = headers.index("chasseur_date")   + 1
        ws.cell(row=row_idx, column=col_s).value = statut
        ws.cell(row=row_idx, column=col_d).value = str(date.today())
        if cat_offset is not None and "chasseur_cat_offset" in headers:
            col_o = headers.index("chasseur_cat_offset") + 1
            ws.cell(row=row_idx, column=col_o).value = cat_offset
    except ValueError:
        pass
    wb.save(vfile)


# ─────────────────────────────────────────────
# PHASE 1 : COLLECTE
# ─────────────────────────────────────────────

def phase_collect(zone: str, city_filter: str | None, env: dict,
                  max_per_cat: int, dry_run: bool, max_duration_min: int = 0):
    """
    Parcourt Google Maps pour chaque ville × catégorie active.
    Stocke les inconnus dans l'onglet "À créer".
    max_duration_min : si > 0, arrête après ce nombre de minutes
                       (finit la catégorie en cours, reprend là où on s'est arrêté le lendemain)
    """
    import time as _time
    vfile = get_villes_file(zone)
    if not vfile.exists():
        log(f"❌ izilife_villes.xlsx introuvable. Lance : python chasseur_lieux.py --zone={zone} --init")
        sys.exit(1)

    categories = load_categories(zone)
    villes     = read_villes_pending(vfile)

    if city_filter:
        villes = [v for v in villes if v.get("ville_slug") == city_filter]

    if not villes:
        log("✅ Aucune ville à chasser.")
        return

    deadline = (_time.time() + max_duration_min * 60) if max_duration_min > 0 else None
    if deadline:
        log(f"⏱️  Timer : {max_duration_min} min — arrêt automatique à {datetime.fromtimestamp(deadline).strftime('%H:%M:%S')}")

    log(f"{len(villes)} ville(s) × {len(categories)} catégorie(s)")
    total_new = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled",
                  "--disable-notifications",
                  "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"]
        )
        page = browser.new_page()
        apply_stealth(page)

        for ville in villes:
            slug       = ville.get("ville_slug", "")
            nom_v      = ville.get("nom_affichage", slug)
            city_id    = resolve_city_id(slug, env) if not dry_run else 0
            # Reprendre à la catégorie suivante si offset stocké
            cat_offset = to_int(ville.get("chasseur_cat_offset"), 0)

            log(f"\n{'='*55}")
            log(f"🏙️  {nom_v} ({slug}) — reprise catégorie #{cat_offset}")
            log(f"{'='*55}")

            ville_new    = 0
            timed_out    = False
            cats_to_do   = categories[cat_offset:]

            for cat_idx_rel, cat in enumerate(cats_to_do):
                cat_idx_abs = cat_offset + cat_idx_rel
                cat_query   = cat["categorie"]
                cat_label   = cat["label_fr"]
                query       = f"{cat_query} à {nom_v}"

                log(f"\n  [{cat_idx_abs+1}/{len(categories)}] [{cat_label}] → '{query}'")

                if dry_run:
                    log(f"  [DRY RUN] Maps search : {query}")
                    continue

                # Chercher sur Google Maps (on finit toujours la catégorie en cours)
                try:
                    found = search_google_maps(page, query, max_results=max_per_cat)
                except GoogleMapsConsentError as exc:
                    log(f"    ERREUR BLOQUANTE : {exc}")
                    log(f"    Offset conserve a #{cat_idx_abs}; cette categorie sera retentee.")
                    raise

                if not found:
                    sleep_random(*DELAY_BETWEEN_CATS)
                else:
                    nouveaux = []
                    for lieu in found:
                        check = izilife_post("/scraper/agentCheckPlace", {
                            "name":    lieu["nom"],
                            "city_id": city_id,
                        }, env)
                        if check and check.get("exists"):
                            continue
                        lieu["ville_slug"]       = slug
                        lieu["categorie_source"] = cat_query
                        nouveaux.append(lieu)
                        sleep_random(0.2, 0.5)

                    if nouveaux:
                        added = append_to_a_creer(vfile, nouveaux)
                        log(f"    → {len(nouveaux)} inconnus, {added} nouveaux ajoutés au Sheet")
                        ville_new += added
                    else:
                        log(f"    → Tous déjà dans izilife")

                # Sauvegarder l'offset après chaque catégorie
                next_offset = cat_idx_abs + 1
                update_ville_statut(vfile, ville["row_idx"], "en_cours", cat_offset=next_offset)

                # Vérifier le timer APRÈS avoir fini la catégorie
                if deadline and _time.time() >= deadline:
                    log(f"\n⏱️  Timer atteint — arrêt après catégorie '{cat_label}'")
                    log(f"   Reprise demain à partir de la catégorie #{next_offset}")
                    timed_out = True
                    break

                sleep_random(*DELAY_BETWEEN_CATS)

            if not timed_out:
                # Ville terminée — reset offset
                update_ville_statut(vfile, ville["row_idx"], "collecté", cat_offset=0)
                log(f"\n  ✅ {nom_v} terminé — {ville_new} nouveaux lieux trouvés")
            else:
                total_new += ville_new
                break  # Arrêter aussi les autres villes

            total_new += ville_new

        browser.close()

    log(f"\n=== COLLECTE TERMINÉE ===")
    log(f"  Nouveaux lieux dans 'À créer' : {total_new}")
    log(f"  → Lance '--insert' pour les importer dans izilife")


# ─────────────────────────────────────────────
# PHASE 2 : INSERT
# ─────────────────────────────────────────────

def phase_insert(zone: str, city_filter: str | None, env: dict, dry_run: bool):
    """
    Prend les lignes 'pending' de l'onglet 'À créer' et appelle
    postAgentFetchAndStoreOnePlace pour chacun.
    """
    vfile = get_villes_file(zone)
    if not vfile.exists():
        log(f"❌ izilife_villes.xlsx introuvable.")
        sys.exit(1)

    rows = read_a_creer_pending(vfile)
    if city_filter:
        rows = [r for r in rows if r.get("ville_slug") == city_filter]

    if not rows:
        log("✅ Aucun lieu pending dans 'À créer'.")
        return

    log(f"{len(rows)} lieu(x) à importer")
    stats = {"done": 0, "unmapped": 0, "errors": 0}

    for row in rows:
        nom       = row.get("nom", "")
        slug      = row.get("ville_slug", "")
        maps_url  = row.get("maps_url", "")
        cat_src   = row.get("categorie_source", "")

        log(f"\n  → {nom} ({slug}) [{cat_src}]")

        if dry_run:
            log(f"  [DRY RUN] postAgentFetchAndStoreOnePlace({nom}, {slug})")
            continue

        payload = {"city": slug, "query": nom}
        resp    = izilife_post(f"/scraper/agentFetchAndStoreOnePlace/{slug}", payload, env)

        if not resp:
            log(f"    ❌ Pas de réponse")
            update_a_creer_status(vfile, row["row_idx"], "error")
            stats["errors"] += 1
            continue

        if resp.get("success"):
            result_type = resp.get("result_type", "")
            if result_type == "unmapped":
                log(f"    ⚠️  Unmapped → ScrapingUnmappedPoi")
                update_a_creer_status(vfile, row["row_idx"], "unmapped")
                stats["unmapped"] += 1
            else:
                log(f"    ✅ {resp.get('name')} [{result_type}] id={resp.get('entity_id')}")
                update_a_creer_status(vfile, row["row_idx"], "done")
                stats["done"] += 1
        elif "déjà existant" in str(resp.get("error", "")).lower():
            log(f"    ℹ️  Déjà existant — skip")
            update_a_creer_status(vfile, row["row_idx"], "done")
            stats["done"] += 1
        else:
            log(f"    ❌ {resp.get('error', '?')}")
            update_a_creer_status(vfile, row["row_idx"], "error")
            stats["errors"] += 1

        sleep_random(2, 4)

    # Purge automatique des done
    if not dry_run:
        purge_done_chasseur(zone)

    log(f"\n=== INSERT TERMINÉ ===")
    log(f"  Créés    : {stats['done']}")
    log(f"  Unmapped : {stats['unmapped']}  (traiter dans le BO)")
    log(f"  Erreurs  : {stats['errors']}")



# ─────────────────────────────────────────────
# PURGE — déplacer les done vers logs/WXX/
# ─────────────────────────────────────────────

def get_log_file_chasseur(zone: str, sheet_name: str) -> Path:
    from datetime import date
    today = date.today()
    week  = today.strftime("W%V")
    d = get_zone_dir(zone) / "logs" / week
    d.mkdir(parents=True, exist_ok=True)
    return d / f"chasseur_lieux_{sheet_name}_{today.isoformat()}.xlsx"


def purge_done_chasseur(zone: str):
    vfile = get_villes_file(zone)
    if not vfile.exists():
        log("Aucun fichier à purger."); return

    sheets_purged = 0
    for sheet_name in ["À créer", "A_creer", "Villes"]:
        for attempt in range(3):
            try:
                wb = openpyxl.load_workbook(vfile)
                if sheet_name not in wb.sheetnames: break
                ws = wb[sheet_name]
                headers = [str(c.value or "").strip().lower() for c in ws[1]]
                stat_col_name = "statut" if "statut" in headers else "chasseur_statut"
                if stat_col_name not in headers: break
                stat_col = headers.index(stat_col_name) + 1

                done_rows = []
                keep_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        status = str(row[stat_col-1] or "").strip().lower()
                        if status == "done":
                            done_rows.append(list(row))
                        else:
                            keep_rows.append(list(row))

                if not done_rows: break

                lf = get_log_file_chasseur(zone, sheet_name.replace("À","A").replace(" ","_"))
                lb = openpyxl.Workbook()
                lws = lb.active
                lws.title = "Done"
                for col, h in enumerate(headers, 1):
                    lws.cell(row=1, column=col, value=str(h).upper())
                for r, row in enumerate(done_rows, 2):
                    for c, val in enumerate(row, 1):
                        lws.cell(row=r, column=c, value=val)
                lb.save(lf)
                log(f"✅ [{sheet_name}] {len(done_rows)} lignes → {lf.name}")

                for row in ws.iter_rows(min_row=2):
                    for cell in row: cell.value = None
                for r, row in enumerate(keep_rows, 2):
                    for c, val in enumerate(row, 1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(vfile)
                sheets_purged += 1
                log(f"✅ [{sheet_name}] {len(keep_rows)} lignes restantes")
                break
            except PermissionError:
                import time; time.sleep(3)

    if sheets_purged == 0:
        log("Aucune ligne done trouvée.")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Chasseur de lieux izilife")
    parser.add_argument("--zone",         type=str, required=True,
                        help="Zone : lille, valenciennes, paris ...")
    parser.add_argument("--env",          choices=ENVS.keys(), default="local")
    parser.add_argument("--city",         type=str, default=None,
                        help="Filtrer sur une ville uniquement (slug)")
    parser.add_argument("--max-per-cat",  type=int, default=40,
                        help="Nombre max de résultats par catégorie (défaut: 40)")
    parser.add_argument("--collect-only",  action="store_true",
                        help="Phase 1 uniquement (collecte Maps → Sheet)")
    parser.add_argument("--insert-only",   action="store_true",
                        help="Phase 2 uniquement (Sheet → izilife)")
    parser.add_argument("--max-duration",  type=int, default=0,
                        help="Durée max en minutes (ex: 90). Finit la catégorie "
                             "en cours puis s'arrête. Reprend là où on s'est arrêté "
                             "le lendemain. 0 = pas de limite.")
    parser.add_argument("--dry-run",       action="store_true")
    parser.add_argument("--init",          action="store_true",
                        help="Créer izilife_villes.xlsx + categories_lieux.json")
    args = parser.parse_args()
    set_current_env(args.env)

    if args.init:
        create_villes_xlsx(args.zone)
        log(f"\nProchain step :")
        log(f"  1. Ouvre izilife_villes.xlsx et ajoute tes villes dans l'onglet 'Villes'")
        log(f"  2. Active/désactive les catégories dans l'onglet 'Catégories'")
        log(f"  3. Lance : python chasseur_lieux.py --zone={args.zone} --env=local --city=<ville>")
        return

    env     = ENVS[args.env]
    zone    = args.zone.lower().strip()
    dry_run = args.dry_run

    log(f"=== chasseur_lieux.py — zone={zone} env={args.env}" +
        (f" city={args.city}" if args.city else "") +
        (" [DRY RUN]" if dry_run else "") + " ===")

    if AGENT_TOKEN == "METTRE_TOKEN_ICI" and not dry_run:
        log("❌ IZILIFE_AGENT_TOKEN non défini.")
        sys.exit(1)

    do_collect = not args.insert_only
    do_insert  = not args.collect_only

    if do_collect:
        log("\n── PHASE 1 : COLLECTE ──────────────────────────────────")
        phase_collect(zone, args.city, env, args.max_per_cat, dry_run, args.max_duration)

    if do_insert:
        log("\n── PHASE 2 : INSERT ────────────────────────────────────")
        phase_insert(zone, args.city, env, dry_run)


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from agent_excel_logger import run_logged
    run_logged("chasseur_lieux", "places", main)
