"""
enrichisseur.py
---------------
Agent d'enrichissement des lieux izilife.
Pour chaque lieu avec un completion_score bas :
  1. Lit le JSON Google Places en cache (0 appel API)
  2. Si absent → 1 appel API Details + stockage cache
  3. Playwright → site web du lieu
  4. Playwright → Google Maps (avis, équipements, ambiance)
  5. Playwright → TripAdvisor (restaurants/bars)
  6. Construit un diff JSON et l'envoie via postAgentSubmitEnrichment

Priorité : si un fichier priority_places.xlsx existe dans la zone,
les URLs izilife listées dedans sont traitées en premier.

Usage :
    python enrichisseur.py --env=local --city=lille --dry-run
    python enrichisseur.py --env=local --city=lille --max=20 --score-max=60
    python enrichisseur.py --env=local --city=lille --type=SHOP
    python enrichisseur.py --env=local --city=lille --max=10 --no-playwright
    python enrichisseur.py --env=local --city=lille --zone=lille --priority-only

Prérequis :
    pip install playwright requests playwright-stealth openpyxl python-dotenv
    python -m playwright install chromium
    Variable : IZILIFE_AGENT_TOKEN, GOOGLE_PLACES_API_KEY

Fichier priorité (optionnel) :
    G:/Mon Drive/agentic_workspace/izilife/places/{zone}-zone/priority_places.xlsx
    Colonnes : url_izilife | notes | statut
    url_izilife : https://izilife.co/place/le-baron-club
                  ou https://izilife.co/shop/le-bon-marche
                  ou juste le string_id : le-baron-club (type déduit de --type ou PLACE par défaut)
    statut : pending → done | error | skip
"""

import os
import sys
import re
import time
import json
import random
import argparse
import requests
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
except ImportError:
    print("pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

ENVS = {
    "local":   {"base_url": "https://localhost:4443/izilife-admin", "verify_ssl": False},
    "staging": {"base_url": "https://www.staging.izilife.co/izilife-admin", "verify_ssl": True},
    "prod":    {"base_url": "https://www.izilife.co/izilife-admin",  "verify_ssl": True},
}

AGENT_TOKEN    = os.environ.get("IZILIFE_AGENT_TOKEN", "")
GOOGLE_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")
AGENT_NAME     = "enrichisseur"

DELAY_BETWEEN_PLACES = (3, 6)
DELAY_PLAYWRIGHT     = (2, 4)

# ─────────────────────────────────────────────
# CHEMINS DRIVE
# ─────────────────────────────────────────────

def get_drive_root() -> Path:
    if sys.platform == "win32":
        candidates = [Path("G:/Mon Drive"), Path("G:/My Drive"),
                      Path.home() / "Google Drive", Path.home() / "Mon Drive"]
    else:
        candidates = [Path.home() / "GoogleDrive", Path.home() / "Google Drive",
                      Path.home() / "gdrive", Path("/mnt/gdrive")]
    for p in candidates:
        if p.exists():
            return p
    fallback = Path(__file__).parent.parent.parent.parent / "izilife-agent-workspace"
    fallback.mkdir(parents=True, exist_ok=True)
    print(f"⚠️  Google Drive non trouvé — fallback : {fallback}")
    return fallback


def get_priority_file(zone: str) -> Path:
    return get_drive_root() / "agentic_workspace" / "izilife" / "places" / f"{zone}-zone" / "priority_places.xlsx"

# ─────────────────────────────────────────────
# MAPPING : données Google Places → payload izilife
# ─────────────────────────────────────────────

GOOGLE_TYPE_TO_CHARCT_TAG = {
    "meal_takeaway":          "a-emporter",
    "delivery":               "livraison",
    "outdoor_seating":        "terasse-agreable",
    "good_for_children":      "adapte-aux-familles",
    "good_for_groups":        "bien-pour-les-groupes",
    "serves_vegetarian_food": "vegetarien",
    "live_music":             "bien-pour-concerts",
    "rooftop":                "roof-top",
    "reservable":             None,
}

GOOGLE_SERVES_TO_CHARACTS = {
    "serves_beer":                    {"can_drink_alcool": 1, "sell_alcool": 1},
    "serves_wine":                    {"can_drink_alcool": 1, "sell_alcool": 1},
    "serves_cocktails":               {"can_drink_alcool": 1, "sell_alcool": 1},
    "serves_brunch":                  {},
    "serves_breakfast":               {},
    "dine_in":                        {},
    "takeout":                        {},
    "curbside_pickup":                {},
    "wheelchair_accessible_entrance": {"adapted_to_handicap": 1},
    "good_for_children":              {"adapted_to_children": 1, "adapted_to_family": 1},
    "good_for_groups":                {"adapted_to_groups": 1},
    "live_music":                     {"play_music": 1},
    "outdoor_seating":                {"have_terrace": 1},
}

GOOGLE_TYPE_TO_MEAL_TYPE = {
    "serves_brunch":    "brunch",
    "serves_breakfast": "petit-dejeuner",
    "serves_lunch":     "dejeuner",
    "serves_dinner":    "diner",
}

ABOUT_AMBIANCE_TO_ATMOSPHERE = {
    "chaleureux":  "chaleureux",
    "décontracté": "decontracte",
    "chic":        "chic",
    "cosy":        "cosy",
    "romantique":  "romantique",
    "animé":       None,
}

ABOUT_TO_EQUIPMENT = {
    "terrasse":              None,
    "terrain de pétanque":   "terrain-de-petanque",
    "billard":               "billard",
    "baby-foot":             "baby-foot",
    "babyfoot":              "baby-foot",
    "fléchettes":            "jeu-de-flechettes",
    "ping-pong":             "ping-pong",
    "tennis de table":       "ping-pong",
    "cabine photo":          "cabine-photo",
    "borne électrique":      "borne-electrique",
    "wifi":                  None,
}

# ─────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def sleep_random(mn, mx):
    time.sleep(random.uniform(mn, mx))

def apply_stealth(page):
    if HAS_STEALTH:
        stealth_sync(page)
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['fr-FR', 'fr']});
        window.chrome = { runtime: {} };
    """)

def izilife_get(endpoint: str, env: dict) -> dict | None:
    try:
        r = requests.get(
            env["base_url"] + endpoint,
            headers={"X-Agent-Token": AGENT_TOKEN},
            verify=env["verify_ssl"], timeout=30,
        )
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ GET {endpoint} : {e}")
        return None

def izilife_post(endpoint: str, data: dict, env: dict, json_body=False) -> dict | None:
    try:
        kwargs = dict(headers={"X-Agent-Token": AGENT_TOKEN},
                      verify=env["verify_ssl"], timeout=30)
        if json_body:
            kwargs["json"] = data
            kwargs["headers"]["Content-Type"] = "application/json"
        else:
            kwargs["data"] = data
        r = requests.post(env["base_url"] + endpoint, **kwargs)
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ POST {endpoint} : {e}")
        return None

def resolve_city_id(city_slug: str, env: dict) -> int:
    r = izilife_get(f"/scraper/cityByStringId/{city_slug}", env)
    if r and r.get("success") and r.get("city"):
        city_id = int(r["city"]["id"])
        log(f"Ville résolue : {city_slug} → city_id={city_id}")
        return city_id
    log(f"❌ Ville introuvable : {city_slug}")
    sys.exit(1)

def google_place_details(place_id: str) -> dict | None:
    if not GOOGLE_API_KEY:
        return None
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields":   "name,rating,user_ratings_total,formatted_phone_number,"
                            "international_phone_number,website,opening_hours,"
                            "editorial_summary,photos,types,business_status,"
                            "serves_beer,serves_wine,serves_brunch,serves_breakfast,"
                            "serves_lunch,serves_dinner,serves_vegetarian_food,"
                            "dine_in,takeout,outdoor_seating,delivery,"
                            "wheelchair_accessible_entrance,reservable,"
                            "good_for_children,good_for_groups,live_music",
                "key":      GOOGLE_API_KEY,
                "language": "fr",
            },
            timeout=15,
        )
        data = r.json()
        return data.get("result") if data.get("status") == "OK" else None
    except Exception as e:
        log(f"  ❌ Google API : {e}")
        return None

# ─────────────────────────────────────────────
# PRIORITÉ — lecture du Sheet priority_places.xlsx
# ─────────────────────────────────────────────

def read_priority_xlsx(zone: str, env: dict) -> list[dict]:
    """
    Lit priority_places.xlsx et retourne une liste de lieux à enrichir en priorité.
    Chaque entrée : {id, type, name, string_id, google_place_id, has_cached_json,
                     google_place_json, website, completion_score, city_id, _row_idx, _wb, _ws, _file}
    """
    if not HAS_OPENPYXL:
        log("⚠️  openpyxl non installé — priorité désactivée")
        return []

    pfile = get_priority_file(zone)
    if not pfile.exists():
        return []

    log(f"📋 Fichier priorité trouvé : {pfile.name}")
    wb = openpyxl.load_workbook(pfile)
    ws = wb.active

    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    try:
        col_url    = headers.index("url_izilife")
        col_status = headers.index("statut")
    except ValueError:
        log("⚠️  Colonnes 'url_izilife' ou 'statut' introuvables dans priority_places.xlsx")
        return []

    lieux = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        status = str(row[col_status] or "").strip().lower()
        if status in ("done", "error", "skip"):
            continue
        url = str(row[col_url] or "").strip()
        if not url:
            continue

        # Résoudre string_id et type depuis l'URL ou valeur brute
        entity_type = "PLACE"
        string_id   = url

        if "/place/" in url:
            entity_type = "PLACE"
            string_id   = url.split("/place/")[-1].strip("/").split("?")[0]
        elif "/shop/" in url:
            entity_type = "SHOP"
            string_id   = url.split("/shop/")[-1].strip("/").split("?")[0]

        # Appeler l'API pour récupérer les données du lieu
        resp = izilife_post(f"/scraper/agentGetPlaceByStringId", {
            "string_id":   string_id,
            "entity_type": entity_type,
        }, env)

        if not resp or not resp.get("success") or not resp.get("item"):
            log(f"  ⚠️  Lieu introuvable pour string_id='{string_id}' — ignoré")
            # Marquer error dans le xlsx
            ws.cell(row=row_idx, column=col_status + 1).value = "error"
            wb.save(pfile)
            continue

        item = resp["item"]
        item["_row_idx"] = row_idx
        item["_wb"]      = wb
        item["_ws"]      = ws
        item["_file"]    = pfile
        item["_col_status"] = col_status + 1  # openpyxl 1-indexed
        lieux.append(item)

    log(f"  → {len(lieux)} lieux prioritaires à traiter")
    return lieux

def mark_priority_done(lieu: dict, status: str = "done"):
    """Met à jour le statut dans le fichier xlsx priorité."""
    wb    = lieu.get("_wb")
    ws    = lieu.get("_ws")
    pfile = lieu.get("_file")
    row   = lieu.get("_row_idx")
    col   = lieu.get("_col_status")
    if wb and ws and pfile and row and col:
        ws.cell(row=row, column=col).value = status
        wb.save(pfile)

# ─────────────────────────────────────────────
# EXTRACTEURS depuis le JSON Google Places
# ─────────────────────────────────────────────

def extract_from_google_json(g: dict) -> dict:
    payload = {"scalar": {}, "principal_characteristics": {}, "satellites": {}, "media": {}}

    desc = g.get("editorial_summary", {}).get("overview")
    if desc:
        payload["scalar"]["short_description"] = desc

    phone = g.get("international_phone_number") or g.get("formatted_phone_number")
    if phone:
        payload["scalar"]["phone_number"] = re.sub(r"[^\d+]", "", phone)

    website = g.get("website")
    if website:
        payload["scalar"]["website"] = website

    rating = g.get("rating")
    if rating:
        payload["scalar"]["google_rating"]       = rating
        payload["scalar"]["google_rating_count"] = g.get("user_ratings_total", 0)

    characts = {}
    for gfield, mapping in GOOGLE_SERVES_TO_CHARACTS.items():
        if g.get(gfield) is True:
            characts.update(mapping)
    if characts:
        payload["principal_characteristics"] = characts

    meal_types = []
    for gfield, string_id in GOOGLE_TYPE_TO_MEAL_TYPE.items():
        if g.get(gfield) is True:
            meal_types.append({"string_id": string_id})
    if meal_types:
        payload["satellites"]["MealType"] = meal_types

    charct_tags = []
    for gtype, string_id in GOOGLE_TYPE_TO_CHARCT_TAG.items():
        if string_id and g.get(gtype) is True:
            charct_tags.append({"string_id": string_id})
    if charct_tags:
        payload["satellites"]["ShopAndPlaceCharactTag"] = charct_tags

    photos = g.get("photos", [])
    refs = [p["photo_reference"] for p in photos[:5] if "photo_reference" in p]
    if refs:
        payload["media"]["photo_references"] = refs

    if g.get("outdoor_seating") is True:
        payload["satellites"]["Terrace"] = [{
            "waterfront": False, "roadside": False,
            "has_parasol": False, "has_heater": False,
            "is_rooftop": False, "suggest_only": True,
        }]

    return _clean_payload(payload)


def extract_from_opening_hours(g: dict) -> dict:
    hours = g.get("opening_hours", {}).get("weekday_text", [])
    if hours:
        return {"scalar": {"opening_hours_text": " | ".join(hours)}}
    return {}


# ─────────────────────────────────────────────
# EXTRACTEURS Playwright
# ─────────────────────────────────────────────

def scrape_website(page, url: str) -> dict:
    payload = {"scalar": {}, "satellites": {}}
    try:
        page.goto(url, timeout=20000)
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except:
            pass

        meta_desc = page.evaluate("""
            () => {
                const m = document.querySelector('meta[name="description"]');
                return m ? m.getAttribute('content') : null;
            }
        """)
        if meta_desc and len(meta_desc) > 30:
            payload["scalar"]["website_description"] = meta_desc[:500]

        page_text = page.evaluate("() => document.body.innerText").lower()
        equipment_found = []
        for keyword, equip_id in ABOUT_TO_EQUIPMENT.items():
            if keyword in page_text and equip_id:
                equipment_found.append({
                    "equipment_category_string_id": equip_id,
                    "detectable": "website",
                })
        if equipment_found:
            payload["satellites"]["Equipment"] = equipment_found

    except Exception as e:
        log(f"    ⚠️  Website scraping : {e}")

    return _clean_payload(payload)


def scrape_google_maps(page, place_name: str, city: str = "") -> dict:
    payload = {"scalar": {}, "principal_characteristics": {}, "satellites": {}}
    search_q = f"{place_name} {city}".strip()

    try:
        url = f"https://www.google.com/maps/search/{requests.utils.quote(search_q)}"
        page.goto(url, timeout=25000)
        sleep_random(2, 4)

        try:
            first_result = page.locator('a[href*="/maps/place/"]').first
            if first_result.is_visible(timeout=3000):
                first_result.click()
                sleep_random(2, 3)
        except:
            pass

        try:
            page.wait_for_selector('h1', timeout=10000)
        except:
            pass

        content = page.content().lower()

        equipment_found = []
        for keyword, equip_id in ABOUT_TO_EQUIPMENT.items():
            if keyword in content:
                if equip_id:
                    equipment_found.append({
                        "equipment_category_string_id": equip_id,
                        "detectable": "google_maps",
                    })
                elif keyword == "terrasse":
                    payload["satellites"]["Terrace"] = [{
                        "waterfront": False, "roadside": False,
                        "has_parasol": False, "has_heater": False,
                        "is_rooftop": "rooftop" in content,
                        "suggest_only": True,
                    }]
                elif keyword == "wifi":
                    payload["principal_characteristics"]["wifi"] = 1

        if equipment_found:
            payload["satellites"]["Equipment"] = equipment_found

        atmosphere_found = []
        for keyword, string_id in ABOUT_AMBIANCE_TO_ATMOSPHERE.items():
            if keyword in content and string_id:
                atmosphere_found.append({"string_id": string_id})
        if atmosphere_found:
            payload["satellites"]["Atmosphere"] = atmosphere_found

        charct_found = []
        if "terrasse en bord de mer" in content or "vue sur mer" in content:
            charct_found.append({"string_id": "vue-sur-mer"})
        if "rooftop" in content or "toit" in content:
            charct_found.append({"string_id": "roof-top"})
        if "parking" in content:
            charct_found.append({"string_id": "stationnement-facile"})
        if charct_found:
            payload["satellites"]["ShopAndPlaceCharactTag"] = charct_found

    except Exception as e:
        log(f"    ⚠️  Google Maps scraping : {e}")

    return _clean_payload(payload)


def scrape_tripadvisor(page, place_name: str, city: str = "") -> dict:
    payload = {"scalar": {}, "satellites": {}}
    search_q = f"{place_name} {city}".strip()

    try:
        url = f"https://www.tripadvisor.fr/Search?q={requests.utils.quote(search_q)}&searchSessionId=&sid="
        page.goto(url, timeout=25000)
        sleep_random(2, 4)

        try:
            result = page.locator('a[href*="Restaurant_Review"]').first
            if result.is_visible(timeout=5000):
                result.click()
                sleep_random(2, 4)
                page.wait_for_selector('h1', timeout=10000)
        except:
            return {}

        content = page.content().lower()

        meta_desc = page.evaluate("""
            () => {
                const m = document.querySelector('meta[name="description"]');
                return m ? m.getAttribute('content') : null;
            }
        """)
        if meta_desc and len(meta_desc) > 30:
            payload["scalar"]["tripadvisor_description"] = meta_desc[:500]

        equipment_found = []
        for keyword, equip_id in ABOUT_TO_EQUIPMENT.items():
            if keyword in content and equip_id:
                equipment_found.append({
                    "equipment_category_string_id": equip_id,
                    "detectable": "tripadvisor",
                })
        if equipment_found:
            payload["satellites"]["Equipment"] = equipment_found

        charct = []
        if "terrasse" in content:
            charct.append({"string_id": "terasse-agreable"})
        if "rooftop" in content:
            charct.append({"string_id": "roof-top"})
        if charct:
            payload["satellites"]["ShopAndPlaceCharactTag"] = charct

    except Exception as e:
        log(f"    ⚠️  TripAdvisor scraping : {e}")

    return _clean_payload(payload)


# ─────────────────────────────────────────────
# FUSION DES PAYLOADS
# ─────────────────────────────────────────────

def merge_payloads(*payloads) -> dict:
    result = {"scalar": {}, "principal_characteristics": {}, "satellites": {}, "media": {}}

    for p in payloads:
        if not p:
            continue
        for k, v in p.get("scalar", {}).items():
            if k not in result["scalar"]:
                result["scalar"][k] = v
        for k, v in p.get("principal_characteristics", {}).items():
            if v == 1:
                result["principal_characteristics"][k] = 1
        for sat_key, items in p.get("satellites", {}).items():
            if sat_key not in result["satellites"]:
                result["satellites"][sat_key] = []
            existing_ids = set()
            for existing in result["satellites"][sat_key]:
                existing_ids.add(
                    existing.get("string_id") or
                    existing.get("equipment_category_string_id") or ""
                )
            for item in items:
                item_id = (
                    item.get("string_id") or
                    item.get("equipment_category_string_id") or ""
                )
                if item_id not in existing_ids:
                    result["satellites"][sat_key].append(item)
                    existing_ids.add(item_id)
        for k, v in p.get("media", {}).items():
            if k not in result["media"]:
                result["media"][k] = v
            elif isinstance(v, list):
                result["media"][k] = list(set(result["media"][k] + v))

    return _clean_payload(result)


def _clean_payload(p: dict) -> dict:
    return {k: v for k, v in p.items() if v}


def payload_is_meaningful(p: dict) -> bool:
    if p.get("scalar"):
        return True
    if p.get("principal_characteristics"):
        return True
    for sat_key, items in p.get("satellites", {}).items():
        if items:
            return True
    if p.get("media", {}).get("photo_references"):
        return True
    return False


# ─────────────────────────────────────────────
# WORKFLOW PRINCIPAL PAR LIEU
# ─────────────────────────────────────────────

def enrich_one(lieu: dict, page, env: dict, dry_run: bool,
               no_playwright: bool, stats: dict):
    entity_id   = lieu["id"]
    entity_type = lieu["type"]
    name        = lieu["name"]
    google_pid  = lieu.get("google_place_id")
    website     = lieu.get("website")
    score       = lieu.get("completion_score", 0)

    log(f"\n  [{entity_type}#{entity_id}] {name} (score={score})")

    partial_payloads = []

    # ── Source 1 : JSON Google Places en cache ────────────────────────────
    cached_json = lieu.get("google_place_json")
    if cached_json:
        log("    → JSON cache Google Places disponible")
        partial_payloads.append(extract_from_google_json(cached_json))
        partial_payloads.append(extract_from_opening_hours(cached_json))

    # ── Source 2 : Google Places API si pas de cache ──────────────────────
    elif google_pid and GOOGLE_API_KEY:
        log(f"    → Appel Google Places API ({google_pid})")
        g = google_place_details(google_pid)
        if g:
            partial_payloads.append(extract_from_google_json(g))
            partial_payloads.append(extract_from_opening_hours(g))
        sleep_random(1, 2)

    # ── Source 3 : Site web du lieu ───────────────────────────────────────
    if not no_playwright and website:
        log(f"    → Scraping site web : {website[:60]}")
        try:
            partial_payloads.append(scrape_website(page, website))
            sleep_random(*DELAY_PLAYWRIGHT)
        except Exception as e:
            log(f"    ⚠️  {e}")

    # ── Source 4 : Google Maps ────────────────────────────────────────────
    if not no_playwright:
        log(f"    → Scraping Google Maps : {name}")
        try:
            partial_payloads.append(scrape_google_maps(page, name))
            sleep_random(*DELAY_PLAYWRIGHT)
        except Exception as e:
            log(f"    ⚠️  {e}")

    # ── Source 5 : TripAdvisor (SHOP uniquement — restaurants/bars) ───────
    if not no_playwright and entity_type == "SHOP":
        log(f"    → Scraping TripAdvisor : {name}")
        try:
            partial_payloads.append(scrape_tripadvisor(page, name))
            sleep_random(*DELAY_PLAYWRIGHT)
        except Exception as e:
            log(f"    ⚠️  {e}")

    # ── Fusion ────────────────────────────────────────────────────────────
    final_payload = merge_payloads(*partial_payloads)

    if not payload_is_meaningful(final_payload):
        log("    → Rien à enrichir")
        stats["skipped"] += 1
        return

    # ── Envoi ─────────────────────────────────────────────────────────────
    if dry_run:
        log(f"    [DRY RUN] Payload prêt :")
        for section, content in final_payload.items():
            if content:
                log(f"      {section}: {json.dumps(content, ensure_ascii=False)[:120]}")
        stats["sent"] += 1
        return

    resp = izilife_post(
        "/scraper/agentSubmitEnrichment",
        {
            "entity_type":      entity_type,
            "entity_id":        entity_id,
            "agent_name":       AGENT_NAME,
            "proposed_payload": final_payload,
            "source_url":       None,
            "confidence_score": 0.7,
        },
        env,
        json_body=True,
    )

    if resp and resp.get("success"):
        vr_id = resp.get("validation_request_id")
        log(f"    ✅ validation_request #{vr_id} créée (risk={resp.get('risk_level')})")
        stats["sent"] += 1
    else:
        log(f"    ❌ Erreur envoi : {resp}")
        stats["errors"] += 1


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--env",           choices=ENVS.keys(), default="local")
    ap.add_argument("--city",          type=str, required=True,
                    help="Slug de la ville : lille, roubaix, valenciennes ...")
    ap.add_argument("--zone",          type=str, required=True,
                    help="Zone Drive : lille, valenciennes, paris ...")
    ap.add_argument("--max",           type=int, default=20)
    ap.add_argument("--score-max",     type=int, default=60)
    ap.add_argument("--type",          choices=["PLACE", "SHOP", ""], default="")
    ap.add_argument("--dry-run",       action="store_true")
    ap.add_argument("--no-playwright", action="store_true",
                    help="Sauter le scraping Playwright (JSON/API uniquement)")
    ap.add_argument("--priority-only", action="store_true",
                    help="Traiter uniquement le fichier priorité, pas la queue normale")
    args = ap.parse_args()

    env     = ENVS[args.env]
    dry_run = args.dry_run
    no_pw   = args.no_playwright

    log(f"=== enrichisseur.py — env={args.env} city={args.city} "
        f"max={args.max} score<{args.score_max}" +
        (f" zone={args.zone}" if args.zone else "") +
        (" [DRY RUN]" if dry_run else "") + " ===")

    if not AGENT_TOKEN and not dry_run:
        log("❌ IZILIFE_AGENT_TOKEN non défini.")
        sys.exit(1)
    if not GOOGLE_API_KEY:
        log("⚠️  GOOGLE_PLACES_API_KEY absent — source API désactivée, cache uniquement")

    city_id = resolve_city_id(args.city, env) if not dry_run else 0

    # ── Lieux prioritaires (fichier xlsx) ────────────────────────────────
    priority_lieux = []
    if args.zone:
        priority_lieux = read_priority_xlsx(args.zone, env)

    # ── File normale via API ──────────────────────────────────────────────
    normal_lieux = []
    if not args.priority_only:
        resp = izilife_post("/scraper/agentGetPlacesToEnrich", {
            "city_id":   city_id,
            "score_max": args.score_max,
            "limit":     args.max,
            "type":      args.type,
        }, env) if not dry_run else None

        normal_lieux = resp["items"] if resp and resp.get("success") else []

        if dry_run and not normal_lieux:
            log("[DRY RUN] Simulation avec 2 lieux fictifs")
            normal_lieux = [
                {"id": 1, "type": "SHOP", "name": "Le Bar du Nord", "google_place_id": "",
                 "has_cached_json": False, "website": None, "completion_score": 35, "city_id": city_id},
                {"id": 2, "type": "PLACE", "name": "Parc de la Citadelle", "google_place_id": "",
                 "has_cached_json": False, "website": None, "completion_score": 42, "city_id": city_id},
            ]

    # Prioritaires d'abord, puis normaux
    # Dédupliquer : retirer de normal_lieux les ids déjà dans priority_lieux
    priority_ids = {(l["id"], l["type"]) for l in priority_lieux}
    normal_lieux = [l for l in normal_lieux if (l["id"], l["type"]) not in priority_ids]

    lieux = priority_lieux + normal_lieux
    log(f"{len(lieux)} lieux à traiter "
        f"({len(priority_lieux)} prioritaires + {len(normal_lieux)} queue normale)")

    if not lieux:
        log("✅ Aucun lieu à enrichir.")
        return

    stats = {"found": len(lieux), "sent": 0, "skipped": 0, "errors": 0}

    # ── Lancer Playwright si besoin ───────────────────────────────────────
    if no_pw:
        for lieu in lieux:
            enrich_one(lieu, None, env, dry_run, no_playwright=True, stats=stats)
            if lieu in priority_lieux and not dry_run:
                mark_priority_done(lieu, "done" if stats["errors"] == 0 else "error")
            sleep_random(*DELAY_BETWEEN_PLACES)
    else:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                args=["--disable-blink-features=AutomationControlled",
                      "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"]
            )
            page = browser.new_page()
            apply_stealth(page)

            for lieu in lieux:
                errors_before = stats["errors"]
                enrich_one(lieu, page, env, dry_run, no_playwright=False, stats=stats)
                if lieu in priority_lieux and not dry_run:
                    status = "done" if stats["errors"] == errors_before else "error"
                    mark_priority_done(lieu, status)
                sleep_random(*DELAY_BETWEEN_PLACES)

            browser.close()

    log(f"\n=== RÉSULTAT ===")
    log(f"  Trouvés  : {stats['found']}")
    log(f"  Envoyés  : {stats['sent']}")
    log(f"  Skippés  : {stats['skipped']}")
    log(f"  Erreurs  : {stats['errors']}")


if __name__ == "__main__":
    main()