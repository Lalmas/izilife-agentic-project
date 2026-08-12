"""
enrichisseur.py
---------------
Agent d'enrichissement des lieux izilife.
Playwright collecte, LLM interprète, on soumet un diff via postAgentSubmitEnrichment.

Sources (dans l'ordre) :
  1. JSON Google Places en cache BDD
  2. postAgentImproveFromGoogle si pas de google_place_id
  3. Playwright → Google Maps
  4. Playwright → site web du lieu
  5. LLM via postAgentAnalyzeContent (PHP, clé côté serveur)

Usage :
    python enrichisseur.py --zone=lille --env=local --city=lille --dry-run
    python enrichisseur.py --zone=lille --env=local --city=lille --max=20
    python enrichisseur.py --zone=lille --env=local --city=lille --max=10 --priority-only
    python enrichisseur.py --zone=lille --env=local --city=lille --purge

Prérequis :
    pip install playwright requests playwright-stealth openpyxl python-dotenv
    python -m playwright install chromium
    Variable : IZILIFE_AGENT_TOKEN

Fichiers Drive (zone) :
    ameliorateur.xlsx   → Onglet Prioritaires : string_id | type | notes | statut
    logs/WXX/enrichisseur_YYYY-MM-DD.xlsx  → lignes done déplacées par --purge
"""

from __future__ import annotations

import os, sys, re, time, json, random, argparse, requests
from datetime import datetime, date
from pathlib import Path

# ─────────────────────────────────────────────
# CORE PATHS / CONFIG — migration safe
# ─────────────────────────────────────────────

def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts"))
            return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent))
            return

_ensure_core_import_path()

try:
    from core.paths import IZILIFE_ENVS, ENV_GLOBAL, ENV_IZILIFE, izilife_places_zone
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False

CURRENT_ENV = "prod"

def set_current_env(env_name: str):
    global CURRENT_ENV
    CURRENT_ENV = str(env_name or "prod").strip().lower()


try:
    from dotenv import load_dotenv
    if HAS_CORE_PATHS:
        for _env_file in (ENV_GLOBAL, ENV_IZILIFE, Path(__file__).parent / ".env"):
            if _env_file and _env_file.exists():
                load_dotenv(_env_file, override=False)
    else:
        load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

ENVS = IZILIFE_ENVS if HAS_CORE_PATHS else {
    "local":   {"base_url": "https://localhost:4443/izilife-admin", "verify_ssl": False},
    "staging": {"base_url": "https://www.staging.izilife.co/izilife-admin", "verify_ssl": True},
    "prod":    {"base_url": "https://www.izilife.co/izilife-admin", "verify_ssl": True},
}

AGENT_TOKEN = os.environ.get("IZILIFE_AGENT_TOKEN", "")
DELAY = (3, 6)

# ─────────────────────────────────────────────
# CHEMINS
# ─────────────────────────────────────────────

def get_drive_root() -> Path:
    value = os.environ.get("AGENTIC_DRIVE_ROOT", "").strip()
    if not value: raise RuntimeError("AGENTIC_DRIVE_ROOT non défini.")
    return Path(value).expanduser()

def zone_dir(zone: str) -> Path:
    if HAS_CORE_PATHS:
        d = izilife_places_zone(zone, CURRENT_ENV)
        d.mkdir(parents=True, exist_ok=True)
        return d
    return get_drive_root() / "agentic_workspace" / "izilife" / "places" / f"{zone}-zone"

def ameliorateur_file(zone: str) -> Path:
    return zone_dir(zone) / "ameliorateur.xlsx"

def log_file(zone: str, script: str = "enrichisseur") -> Path:
    from datetime import date
    today = date.today()
    week  = today.strftime("W%V")
    d     = zone_dir(zone) / "logs" / week
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{script}_{today.isoformat()}.xlsx"

# ─────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def sleep_rnd(mn=2, mx=5):
    time.sleep(random.uniform(mn, mx))

def apply_stealth(page):
    if HAS_STEALTH:
        stealth_sync(page)
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['fr-FR', 'fr']});
        window.chrome = { runtime: {} };
    """)

def izilife_get(ep: str, env: dict) -> dict | None:
    try:
        r = requests.get(env["base_url"]+ep, headers={"X-Agent-Token": AGENT_TOKEN},
                         verify=env["verify_ssl"], timeout=15)
        return r.json() if r.status_code == 200 else None
    except Exception as e:
        log(f"  ❌ GET {ep}: {e}"); return None

def izilife_post(ep: str, data: dict, env: dict, json_body=False) -> dict | None:
    try:
        kw = dict(headers={"X-Agent-Token": AGENT_TOKEN}, verify=env["verify_ssl"], timeout=60)
        if json_body:
            kw["json"] = data; kw["headers"]["Content-Type"] = "application/json"
        else:
            kw["data"] = data
        r = requests.post(env["base_url"]+ep, **kw)
        if r.status_code == 200:
            try:
                return r.json()
            except ValueError:
                log(f"  ❌ Réponse non JSON sur POST {ep}")
                log(f"  ↳ HTTP {r.status_code}")
                log(f"  ↳ Content-Type: {r.headers.get('Content-Type')}")
                log(f"  ↳ Body: {r.text[:1000]}")
                return None
        log(f"  ❌ HTTP {r.status_code}: {r.text[:500]}"); return None
    except Exception as e:
        log(f"  ❌ POST {ep}: {e}"); return None

def resolve_city_id(slug: str, env: dict) -> int:
    r = izilife_get(f"/scraper/cityByStringId/{slug}", env)
    if r and r.get("success"):
        cid = int(r["city"]["id"])
        log(f"Ville résolue : {slug} → city_id={cid}")
        return cid
    log(f"❌ Ville introuvable : {slug}"); sys.exit(1)

# ─────────────────────────────────────────────
# XLSX — ameliorateur.xlsx
# ─────────────────────────────────────────────

HDR_BG = "1F4E79"
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _make_header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h.upper())
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def _style_row(ws, r, ncols, alt=False):
    bg = "EBF3FB" if alt else "FFFFFF"
    for c in range(1, ncols+1):
        cell = ws.cell(row=r, column=c)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = BORDER
        cell.fill      = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[r].height = 18

def _add_dv(ws, col, formula, max_row=2000):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False,
                        sqref=f"{col}2:{col}{max_row}")
    ws.data_validations.append(dv)

def _add_cf(ws, rng, col, colors):
    for status, color in colors.items():
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'${col}2="{status}"'],
            fill=PatternFill("solid", fgColor=color)
        ))

PRIO_COLS = ["string_id", "type", "ferme_definitivement", "notes", "statut"]


def create_ameliorateur(zone: str):
    """
    Crée ameliorateur.xlsx pour les priorités manuelles.

    IMPORTANT :
    - pas de lignes exemple actives ;
    - pas de précréation de centaines de lignes vides ;
    - l'utilisateur renseigne seulement les lignes utiles.

    Colonnes :
      string_id : slug izilife, ex: big-luck
      type      : SHOP ou PLACE
      notes     : notes libres
      statut    : pending | done | error | skip
    """
    f = ameliorateur_file(zone)
    f.parent.mkdir(parents=True, exist_ok=True)
    if f.exists():
        log(f"⚠️  Fichier déjà existant : {f}"); return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prioritaires"
    _make_header(ws, PRIO_COLS, [42, 12, 22, 50, 14])

    # Validations sans écrire de lignes fantômes.
    _add_dv(ws, "B", '"SHOP,PLACE"', max_row=500)
    _add_dv(ws, "C", '"NON,OUI"', max_row=500)
    _add_dv(ws, "E", '"pending,done,error,skip"', max_row=500)
    _add_cf(ws, "A2:E500", "E", {"done":"C6EFCE","error":"FFC7CE","skip":"D9D9D9"})

    wb.save(f)
    log(f"✅ ameliorateur.xlsx créé : {f}")

def read_prioritaires(zone: str, env: dict) -> list[dict]:
    """Lit les lignes manuelles prioritaires de ameliorateur.xlsx.

    Format attendu : string_id | type | notes | statut
    Les URL anciennes /place/... ou /shop/... restent tolérées pour compatibilité.
    """
    f = ameliorateur_file(zone)
    if not f.exists():
        return []

    wb = openpyxl.load_workbook(f)
    ws = wb["Prioritaires"] if "Prioritaires" in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    if "ferme_definitivement" not in headers:
        insert_at = headers.index("notes") + 1 if "notes" in headers else len(headers) + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at, "FERME_DEFINITIVEMENT")
        _make_header(ws, [str(c.value or "") for c in ws[1]], [42, 12, 22, 50, 14])
        _add_dv(ws, ws.cell(1, insert_at).column_letter, '"NON,OUI"', max_row=500)
        wb.save(f)
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
    rows = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        data = {headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)}
        status = data.get("statut", "").lower() or "pending"
        if status in ("done", "error", "skip"):
            continue

        raw_string_id = data.get("string_id", "").strip()
        if not raw_string_id:
            continue

        entity_type = data.get("type", "").strip().upper()
        string_id = raw_string_id

        # Compat ancien format URL : https://izilife.co/place/xxx ou /shop/xxx
        if "/place/" in raw_string_id:
            entity_type = "PLACE"
            string_id = raw_string_id.split("/place/")[-1].strip("/").split("?")[0]
        elif "/shop/" in raw_string_id:
            entity_type = "SHOP"
            string_id = raw_string_id.split("/shop/")[-1].strip("/").split("?")[0]

        if entity_type not in ("SHOP", "PLACE"):
            log(f"  ⚠️  Type manquant/invalide pour {string_id} — renseigne SHOP ou PLACE")
            continue

        resp = izilife_post("/scraper/agentGetPlaceByStringId",
                            {"string_id": string_id, "entity_type": entity_type}, env)
        if not resp or not resp.get("success"):
            log(f"  ⚠️  Lieu introuvable : {entity_type}/{string_id}")
            continue

        item = resp["item"]
        closed_value = data.get("ferme_definitivement", data.get("ferme", ""))
        item["_close_permanently"] = str(closed_value).strip().lower() in {
            "1", "oui", "true", "yes", "ferme", "fermé", "ferme-definitivement", "closed"
        }
        item["_row_idx"] = r_idx
        item["_file"]    = f
        rows.append(item)

    log(f"📋 {len(rows)} lieux prioritaires")
    return rows

def mark_prioritaire_done(lieu: dict, status="done"):
    f   = lieu.get("_file")
    idx = lieu.get("_row_idx")
    if not f or not idx: return
    for attempt in range(3):
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb["Prioritaires"] if "Prioritaires" in wb.sheetnames else wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            col = (headers.index("statut")+1) if "statut" in headers else 3
            ws.cell(row=idx, column=col).value = status
            wb.save(f)
            return
        except PermissionError:
            time.sleep(2)

def close_permanently(lieu: dict, env: dict, dry_run: bool, stats: dict):
    label = f"{lieu['type']}#{lieu['id']} {lieu['name']}"
    if dry_run:
        log(f"  [DRY RUN] fermeture définitive : {label}")
        stats["closed"] += 1
        return True
    resp = izilife_post("/scraper/agentClosePlace", {
        "entity_type": lieu["type"],
        "entity_id": lieu["id"],
    }, env)
    if resp and resp.get("success"):
        log(f"  ✅ Fermé définitivement : {label}")
        stats["closed"] += 1
        return True
    log(f"  ❌ Fermeture impossible : {label} — {resp}")
    stats["errors"] += 1
    return False

# ─────────────────────────────────────────────
# PURGE — déplacer les done vers logs/WXX/
# ─────────────────────────────────────────────

def purge_done(zone: str, script: str = "enrichisseur"):
    f = ameliorateur_file(zone)
    if not f.exists():
        log("Aucun fichier à purger."); return
    wb = openpyxl.load_workbook(f)
    ws = wb["Prioritaires"] if "Prioritaires" in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    stat_col = (headers.index("statut")+1) if "statut" in headers else 3

    done_rows = []
    keep_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[stat_col-1] or "").strip().lower()
        if status == "done":
            done_rows.append(list(row))
        else:
            keep_rows.append(list(row))

    if not done_rows:
        log("Aucune ligne done à purger."); return

    # Écrire le log
    lf = log_file(zone, script)
    lb = openpyxl.Workbook()
    lws = lb.active
    lws.title = "Done"
    width_by_header = {"string_id": 42, "type": 12, "ferme_definitivement": 22, "notes": 50, "statut": 14}
    _make_header(lws, headers, [width_by_header.get(h, 20) for h in headers])
    for r, row in enumerate(done_rows, 2):
        for c, val in enumerate(row, 1):
            lws.cell(row=r, column=c, value=val)
        _style_row(lws, r, len(headers), alt=(r%2==0))
    lb.save(lf)
    log(f"✅ {len(done_rows)} lignes déplacées vers {lf.name}")

    # Réécrire le fichier principal sans les done
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    for r, row in enumerate(keep_rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        _style_row(ws, r, len(headers), alt=(r%2==0))
    wb.save(f)
    log(f"✅ Fichier principal purgé — {len(keep_rows)} lignes restantes")

# ─────────────────────────────────────────────
# PLAYWRIGHT — collecteurs
# ─────────────────────────────────────────────

def accept_cookies(page):
    selectors = [
        'button:has-text("Tout refuser")', 'button:has-text("Reject all")',
        '[role="button"]:has-text("Tout refuser")', '[role="button"]:has-text("Reject all")',
        'button:has-text("Tout accepter")', 'button:has-text("Accept all")',
        'button:has-text("Accepter tout")', 'input[value="Tout accepter"]',
        'input[value="Tout refuser"]', 'input[value="Reject all"]',
        'input[value="Accept all"]', '[aria-label*="Tout accepter" i]',
        '[aria-label*="Tout refuser" i]', '[aria-label*="Reject all" i]',
        '[aria-label*="Accept all" i]', '[role="button"]:has-text("Tout accepter")',
        '[role="button"]:has-text("Accept all")',
    ]
    for _ in range(5):
        targets = [page] + list(page.frames)
        for target in targets:
            for sel in selectors:
                try:
                    btn = target.locator(sel).last
                    if btn.is_visible(timeout=500):
                        btn.click(force=True, timeout=3000)
                        page.wait_for_timeout(1500)
                        if "consent.google." not in page.url:
                            return True
                except Exception:
                    pass
        page.wait_for_timeout(750)
    return "consent.google." not in page.url

def scrape_google_maps(page, name: str, city: str = "") -> str:
    query = f"{name} {city}".strip()
    url = f"https://www.google.com/maps/search/{requests.utils.quote(query)}"

    try:
        page.goto(url, timeout=30000)
        sleep_rnd(2, 3)
        if not accept_cookies(page):
            raise RuntimeError("Ecran de consentement Google non ferme")
        sleep_rnd(1, 2)

        # Si liste de résultats : cliquer le premier résultat pertinent.
        try:
            first = page.locator('a[href*="/maps/place/"]').first
            if first.is_visible(timeout=3000):
                first.click()
                sleep_rnd(2, 3)
        except Exception:
            pass

        try:
            page.wait_for_selector("h1", timeout=10000)
        except Exception:
            pass

        # Onglet "À propos" / "About"
        clicked_about = False
        about_selectors = [
            'button:has-text("À propos")',
            'button:has-text("About")',
            'div[role="tab"]:has-text("À propos")',
            'div[role="tab"]:has-text("About")',
        ]

        for sel in about_selectors:
            try:
                tab = page.locator(sel).first
                if tab.is_visible(timeout=2000):
                    tab.click()
                    clicked_about = True
                    sleep_rnd(1, 2)
                    break
            except Exception:
                pass

        # Scroll interne du panneau Google Maps, pas seulement window.
        for _ in range(4):
            try:
                panel = page.locator('div[role="main"]').first
                if panel.is_visible(timeout=1000):
                    panel.evaluate("el => el.scrollTop += 700")
                else:
                    page.evaluate("window.scrollBy(0, 700)")
            except Exception:
                try:
                    page.evaluate("window.scrollBy(0, 700)")
                except Exception:
                    pass
            sleep_rnd(0.8, 1.3)

        text = page.evaluate("() => document.body.innerText")

        marker = "=== GOOGLE MAPS"
        if clicked_about:
            marker += " / À PROPOS"
        marker += " ==="

        return f"{marker}\n{text[:10000]}"

    except Exception as e:
        log(f"    ⚠️  Google Maps: {e}")
        return ""

def scrape_website(page, url: str) -> str:
    try:
        page.goto(url, timeout=20000)
        try: page.wait_for_load_state("networkidle", timeout=6000)
        except: pass
        text = page.evaluate("() => document.body.innerText")
        return f"=== SITE WEB ({url[:60]}) ===\n{text[:5000]}"
    except Exception as e:
        log(f"    ⚠️  Site web: {e}"); return ""

# ─────────────────────────────────────────────
# WORKFLOW PAR LIEU
# ─────────────────────────────────────────────

def enrich_one(lieu: dict, page, env: dict, dry_run: bool, stats: dict):
    eid    = lieu["id"]
    etype  = lieu["type"]
    name   = lieu["name"]
    gid    = lieu.get("google_place_id","")
    website= lieu.get("website","") or ""
    score  = lieu.get("completion_score", 0)
    city_id= lieu.get("city_id", 0)

    log(f"\n  [{etype}#{eid}] {name} (score={score})")

    # ── Step 1 : S'assurer qu'on a google_place_id ────────────────────
    if not gid:
        if dry_run:
            log("    [DRY RUN] Pas de google_place_id — n'appelle pas agentImproveFromGoogle")
        else:
            log("    → Pas de google_place_id — appel agentImproveFromGoogle")
            resp = izilife_post("/scraper/agentImproveFromGoogle",
                                {"entity_type": etype, "entity_id": eid}, env)
            if resp and resp.get("success"):
                gid = resp.get("google_place_id","")
                lieu["google_place_json"] = None  # sera rechargé
                log(f"    → google_place_id obtenu : {gid}")
            else:
                log("    ⚠️  Impossible d'obtenir google_place_id")

    # ── Step 2 : Collecter le texte depuis toutes les sources ──────────
    content_parts = []

    # Cache JSON BDD
    cached = lieu.get("google_place_json")
    if cached:
        import json as _json
        brief = {k: cached.get(k) for k in
                 ["name","rating","user_ratings_total","editorial_summary",
                  "website","opening_hours","types","wheelchair_accessible_entrance",
                  "serves_beer","serves_wine","outdoor_seating","dine_in"]}
        content_parts.append(f"=== GOOGLE PLACES CACHE ===\n{_json.dumps(brief, ensure_ascii=False)[:3000]}")

    # Playwright (si dispo)
    if page:
        # Google Maps
        gm = scrape_google_maps(page, name)
        if gm: content_parts.append(gm); sleep_rnd(2,3)

        # Site web
        if website:
            sw = scrape_website(page, website)
            if sw: content_parts.append(sw); sleep_rnd(1,2)

    full_content = "\n\n".join(filter(None, content_parts))
    if not full_content.strip():
        log("    → Aucun contenu collecté — skip")
        stats["skipped"] += 1; return

    if dry_run:
        log(f"    [DRY RUN] {len(full_content)} chars collectés")
        log(f"    [DRY RUN] → postAgentAnalyzeContent → postAgentSubmitEnrichment")
        stats["sent"] += 1; return

    # ── Step 3 : LLM via postAgentAnalyzeContent ──────────────────────
    resp = izilife_post("/scraper/agentAnalyzeContent", {
        "entity_type": etype,
        "entity_id":   eid,
        "content":     full_content,
    }, env)

    if not resp or not resp.get("success"):
        log(f"    ❌ LLM erreur : {resp}")
        stats["errors"] += 1; return

    payload = resp.get("payload", {})
    if not payload:
        log("    → Payload LLM vide"); stats["skipped"] += 1; return

    # ── Step 4 : Soumettre l'enrichissement ───────────────────────────
    submit = izilife_post("/scraper/agentSubmitEnrichment", {
        "entity_type":      etype,
        "entity_id":        eid,
        "agent_name":       "enrichisseur",
        "confidence_score": 0.75,
        "proposed_payload": payload,
    }, env, json_body=True)

    if submit and submit.get("success"):
        vr = submit.get("validation_request_id")
        log(f"    ✅ validation_request #{vr} créée (risk={submit.get('risk_level')})")
        stats["sent"] += 1
    else:
        log(f"    ❌ Submit erreur : {submit}")
        stats["errors"] += 1

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zone",         type=str, required=True)
    ap.add_argument("--env",          choices=ENVS.keys(), default="local")
    ap.add_argument("--city",         type=str, required=True)
    ap.add_argument("--max",          type=int, default=20)
    ap.add_argument("--score-max",    type=int, default=60)
    ap.add_argument("--type",         choices=["PLACE","SHOP",""], default="")
    ap.add_argument("--dry-run",      action="store_true")
    ap.add_argument("--priority-only",action="store_true")
    ap.add_argument("--init",         action="store_true", help="Créer ameliorateur.xlsx")
    ap.add_argument("--purge",        action="store_true", help="Déplacer les done vers logs/WXX/")
    args = ap.parse_args()
    set_current_env(args.env)

    if args.init:
        create_ameliorateur(args.zone); return

    if args.purge:
        purge_done(args.zone, "enrichisseur"); return

    env     = ENVS[args.env]
    dry_run = args.dry_run

    log(f"=== enrichisseur.py — zone={args.zone} env={args.env} city={args.city} max={args.max}" +
        (" [DRY RUN]" if dry_run else "") + " ===")

    if not AGENT_TOKEN and not dry_run:
        log("❌ IZILIFE_AGENT_TOKEN non défini."); sys.exit(1)

    city_id = resolve_city_id(args.city, env)

    # Prioritaires
    priority = read_prioritaires(args.zone, env)

    # Queue normale : appel lecture côté app, même en dry-run.
    # Le dry-run interdit seulement les écritures/mutations.
    normal = []
    if not args.priority_only:
        resp = izilife_post("/scraper/agentGetPlacesToEnrich", {
            "city_id":   city_id,
            "score_max": args.score_max,
            "limit":     args.max,
            "type":      args.type,
        }, env)
        normal = (resp or {}).get("items", []) if resp else []

    priority_ids = {(l["id"],l["type"]) for l in priority}
    normal = [l for l in normal if (l["id"],l["type"]) not in priority_ids]
    lieux  = priority + normal

    log(f"{len(lieux)} lieux ({len(priority)} prioritaires + {len(normal)} queue)")
    if not lieux:
        log("✅ Aucun lieu à enrichir."); return

    stats = {"found": len(lieux), "sent": 0, "closed": 0, "skipped": 0, "errors": 0}

    to_close = [l for l in priority if l.get("_close_permanently")]
    for lieu in to_close:
        err_before = stats["errors"]
        close_permanently(lieu, env, dry_run, stats)
        if not dry_run:
            mark_prioritaire_done(lieu, "done" if stats["errors"] == err_before else "error")
    lieux = [l for l in lieux if not l.get("_close_permanently")]

    if not lieux:
        pass
    elif not HAS_PLAYWRIGHT:
        for lieu in lieux:
            enrich_one(lieu, None, env, dry_run, stats)
            if lieu in priority: mark_prioritaire_done(lieu, "done" if not stats["errors"] else "error")
            sleep_rnd(*DELAY)
    else:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False, args=[
                "--disable-blink-features=AutomationControlled",
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            ])
            page = browser.new_page()
            apply_stealth(page)
            for lieu in lieux:
                err_before = stats["errors"]
                enrich_one(lieu, page, env, dry_run, stats)
                if lieu in priority:
                    mark_prioritaire_done(lieu, "done" if stats["errors"]==err_before else "error")
                sleep_rnd(*DELAY)
            browser.close()

    # Purge automatique des done
    if not dry_run:
        purge_done(args.zone, "enrichisseur")

    log(f"\n=== RÉSULTAT ===")
    log(f"  Trouvés  : {stats['found']}")
    log(f"  Envoyés  : {stats['sent']}")
    log(f"  Fermés    : {stats['closed']}")
    log(f"  Skippés  : {stats['skipped']}")
    log(f"  Erreurs  : {stats['errors']}")

if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from agent_excel_logger import run_logged
    run_logged("enrichisseur_lieux", "places", main)
