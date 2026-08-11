from __future__ import annotations

import json, os, re, sys
from datetime import date, datetime, time as datetime_time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts")); return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent)); return


_ensure_core_import_path()
try:
    from core.paths import ENV_GLOBAL, ENV_IZILIFE, IZILIFE_ENVS, drive_workspace_root, normalize_env, normalize_zone
except Exception:
    ROOT = Path.home() / "Documents" / "agentic_Workspace"
    ENV_GLOBAL, ENV_IZILIFE = ROOT / ".env", ROOT / "izilife" / ".env.izilife"
    IZILIFE_ENVS = {
        "local":{"base_url":os.getenv("IZILIFE_LOCAL_URL","https://localhost:4443/izilife-admin"),"verify_ssl":False},
        "staging":{"base_url":os.getenv("IZILIFE_STAGING_URL","https://www.staging.izilife.co/izilife-admin"),"verify_ssl":True},
        "prod":{"base_url":os.getenv("IZILIFE_PROD_URL","https://www.izilife.co/izilife-admin"),"verify_ssl":True},
    }
    def normalize_env(v): return str(v or "prod").strip().lower()
    def normalize_zone(v):
        v = str(v or "").strip().lower(); return v if v.endswith("-zone") else v + "-zone"
    def drive_workspace_root(env_name="prod"):
        suffix={"local":"agentic_workspace_local","staging":"agentic_workspace_staging","prod":"agentic_workspace"}
        return Path(os.environ["AGENTIC_DRIVE_ROOT"]).expanduser() / suffix[normalize_env(env_name)]


DAY = {"lun":1,"mar":2,"mer":3,"jeu":4,"ven":5,"sam":6,"dim":7}


def load_env():
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_IZILIFE.exists(): load_dotenv(ENV_IZILIFE, override=False)


def workbook_path(env_name: str, zone: str, filename: str) -> Path:
    return drive_workspace_root(env_name) / "izilife" / "offers" / normalize_zone(zone) / filename


def normalize_time(value: str) -> str:
    if isinstance(value, datetime_time):
        return value.strftime("%H:%M")
    value = str(value or "").strip().lower().replace("h", ":")
    if re.fullmatch(r"(?:[01]?\d|2[0-3]):[0-5]\d:[0-5]\d", value):
        value = value[:5]
    if re.fullmatch(r"\d{1,2}", value): value += ":00"
    if re.fullmatch(r"\d{1,2}:\d", value): value += "0"
    if not re.fullmatch(r"(?:[01]?\d|2[0-3]):[0-5]\d", value):
        raise ValueError(f"Heure invalide: {value}")
    h, m = value.split(":")
    return f"{int(h):02d}:{m}"


def expand_days(expr: str) -> list[int]:
    out = []
    for part in re.split(r"[,/+ ]+", str(expr).strip().lower()):
        if not part: continue
        if "-" in part:
            a, b = part.split("-", 1)
            if a not in DAY or b not in DAY: raise ValueError(f"Jours invalides: {part}")
            start, end = DAY[a], DAY[b]
            out.extend(range(start, end + 1) if start <= end else list(range(start, 8)) + list(range(1, end + 1)))
        else:
            if part not in DAY: raise ValueError(f"Jour invalide: {part}")
            out.append(DAY[part])
    return sorted(set(out))


def parse_schedules(text: str) -> list[dict]:
    text = str(text or "").strip()
    if not text: return []
    result = []
    for group in re.split(r"\s*&\s*", text):
        match = re.fullmatch(r"\s*([^:]+)\s*:\s*([^\s]+)\s*-\s*([^\s]+)\s*", group)
        if not match: raise ValueError(f"Créneau invalide: {group}")
        start, end = normalize_time(match.group(2)), normalize_time(match.group(3))
        result.extend({"day_of_week": day, "start_time": start, "end_time": end} for day in expand_days(match.group(1)))
    return result


def parse_time_windows(text: str) -> list[dict]:
    text = str(text or "").strip()
    if not text: return []
    result=[]
    for part in re.split(r"\s*&\s*", text):
        match=re.fullmatch(r"\s*([^\s]+)\s*-\s*([^\s]+)\s*", part)
        if not match: raise ValueError(f"Plage horaire invalide: {part}")
        result.append({"start":normalize_time(match.group(1)),"end":normalize_time(match.group(2))})
    return result


def constraints(row: dict, prefix: str = "") -> dict:
    result = {}
    days = str(row.get(prefix + "usable_days") or "").strip()
    if days: result["usable_days"] = expand_days(days)
    windows = parse_time_windows(row.get(prefix + "access_hours") or "")
    schedules = parse_schedules(row.get(prefix + "access_schedule") or "")
    windows.extend({"start": item["start_time"], "end": item["end_time"], "days": [item["day_of_week"]]} for item in schedules)
    if windows: result["time_windows"] = windows
    blackouts = parse_time_windows(row.get(prefix + "blackout_hours") or "")
    if blackouts: result["blackout_windows"] = blackouts
    last = str(row.get(prefix + "last_use_time") or "").strip()
    if last: result["last_use_time"] = normalize_time(last)
    if row.get(prefix + "opening_hours_required") not in (None, ""):
        result["opening_hours_required"] = as_bool(row.get(prefix + "opening_hours_required"))
    if row.get(prefix + "allow_if_opening_hours_unknown") not in (None, ""):
        result["allow_if_opening_hours_unknown"] = as_bool(row.get(prefix + "allow_if_opening_hours_unknown"))
    available = str(row.get(prefix + "available_for") or "").strip()
    if available: result["available_for"] = sorted(set(x.strip() for x in re.split(r"[,|]+", available) if x.strip()))
    return result


def as_bool(value) -> int:
    return 1 if str(value or "").strip().lower() in {"1","true","yes","oui","y"} else 0


def scalar(value):
    if isinstance(value, (datetime, date)): return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def rows(path: Path, sheet_name: str) -> tuple[openpyxl.Workbook, object, list[dict]]:
    wb=openpyxl.load_workbook(path); ws=wb[sheet_name]
    headers=[str(c.value or "").strip() for c in ws[1]]
    data=[]
    for cells in ws.iter_rows(min_row=2):
        item={headers[i]:scalar(cells[i].value) for i in range(len(headers)) if headers[i]}
        if any(v not in (None, "") for k,v in item.items() if k not in {"status","last_result"}):
            item["__row"] = cells[0].row; data.append(item)
    return wb, ws, data


def create_workbook(path: Path, sheet_name: str, columns: list[tuple[str,str]], example: dict, validations: dict[str,list[str]]):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists(): print(f"Existe déjà : {path}"); return
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=sheet_name
    colors={"scope":"4472C4","offer":"70AD47","access":"FFC000","bpr":"A64D79","result":"7F8C8D"}
    for idx,(name,group) in enumerate(columns,1):
        cell=ws.cell(1,idx,name); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=colors[group]); cell.alignment=Alignment(wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width=max(14,min(36,len(name)+3))
        if name in validations:
            dv=DataValidation(type="list",formula1='"'+','.join(validations[name])+'"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"{cell.column_letter}2:{cell.column_letter}500")
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{openpyxl.utils.get_column_letter(len(columns))}500"
    ws.append([example.get(name,"") for name,_ in columns])
    ws.row_dimensions[2].height=44
    wb.save(path); print(f"Créé : {path}")


def post(env_name: str, endpoint: str, payload: dict) -> dict:
    load_env(); token=os.getenv("IZILIFE_AGENT_TOKEN") or os.getenv("AGENT_SECRET_TOKEN") or ""
    if not token: raise RuntimeError("IZILIFE_AGENT_TOKEN manquant.")
    env=IZILIFE_ENVS[normalize_env(env_name)]
    response=requests.post(env["base_url"].rstrip("/")+endpoint,data={"payload":json.dumps(payload,ensure_ascii=False)},headers={"X-Agent-Token":token},verify=env.get("verify_ssl",True),timeout=120)
    try: result=response.json()
    except Exception: result={"success":False,"error":f"HTTP {response.status_code}: {response.text[:500]}"}
    if response.status_code >= 400: result.setdefault("success",False)
    return result


def process(path: Path, sheet_name: str, env_name: str, endpoint: str, payload_builder, dry_run: bool, max_rows: int):
    if not path.exists(): raise FileNotFoundError(f"Fichier absent, lance --init : {path}")
    wb,ws,data=rows(path,sheet_name); headers=[str(c.value or "").strip() for c in ws[1]]
    status_col=headers.index("status")+1; result_col=headers.index("last_result")+1
    handled=0
    for row in data:
        if handled >= max_rows: break
        if str(row.get("status") or "pending").lower() in {"done","skip"}: continue
        try:
            payload=payload_builder(row)
            if dry_run: print("[DRY RUN]",json.dumps(payload,ensure_ascii=False)); handled+=1; continue
            res=post(env_name,endpoint,payload)
            ws.cell(row["__row"],status_col,"done" if res.get("success") else "error")
            ws.cell(row["__row"],result_col,json.dumps(res,ensure_ascii=False)[:1000])
            print(("OK" if res.get("success") else "ERREUR"),res)
        except Exception as exc:
            if not dry_run:
                ws.cell(row["__row"],status_col,"error"); ws.cell(row["__row"],result_col,str(exc)[:1000])
            print("ERREUR",exc)
        handled+=1
    if not dry_run: wb.save(path)
