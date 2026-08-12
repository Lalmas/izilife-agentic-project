"""Agent Pages iziLife : Excel + médias locaux, par zone/environnement, sans Playwright."""
from __future__ import annotations
import argparse, html, json, os, re
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
import requests, openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from objects.object_agent_common import drive_workspace_root, izilife_post, load_env, normalize_env, normalize_zone, slugify, IZILIFE_ENVS

HEADERS=["NOM","CATEGORIE","EMAIL","TELEPHONE","DESCRIPTION","LIEN_1","LIEN_2","LIEN_3","LIEN_4","LIEN_5","LIEN_6","LIEN_7","LIEN_8","MEDIA_DIRECTORY_NAME"]
CATEGORIES=["Influenceur/Media","Artiste/DJ","Assos/Collectif","Communautés","Coach/Prof Danse","Création digitale","Média"]
MEDIA_EXT={".jpg",".jpeg",".png",".webp",".gif",".mp4",".mov",".webm"}

def zone_dir(zone,env): return drive_workspace_root(env)/"izilife"/"pages"/normalize_zone(zone)
def default_file(zone,env): return zone_dir(zone,env)/"pages.xlsx"

def init_sheet(zone,env):
    root=zone_dir(zone,normalize_env(env)); root.mkdir(parents=True,exist_ok=True); (root/"medias").mkdir(exist_ok=True)
    path=root/"pages.xlsx"
    if path.exists(): print(f"Existe déjà : {path}"); return
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Pages"; ws.append(HEADERS)
    widths=[34,24,30,18,65,48,48,48,48,48,48,48,48,28]
    for i,(h,w) in enumerate(zip(HEADERS,widths),1):
        c=ws.cell(1,i); c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78"); c.alignment=Alignment(horizontal="center"); ws.column_dimensions[c.column_letter].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:N500"
    dv=DataValidation(type="list",formula1='"'+','.join(CATEGORIES)+'"',allow_blank=False); ws.add_data_validation(dv); dv.add("B2:B500")
    for r in range(2,501):
        if r%2==0:
            for c in ws[r]: c.fill=PatternFill("solid",fgColor="EAF2F8")
    wb.save(path); print(f"Créé : {path}\nMédias : {root/'medias'}")

def clean_url(value):
    value=str(value or "").strip()
    if not re.match(r"^https?://",value,re.I): return ""
    p=urlsplit(value); q=urlencode([(k,v) for k,v in parse_qsl(p.query) if not k.lower().startswith(("utm_","igsh"))])
    return urlunsplit((p.scheme.lower(),p.netloc.lower(),p.path.rstrip("/") or "/",q,""))

def description_from(urls):
    for url in urls:
        try:
            text=requests.get(url,timeout=12,headers={"User-Agent":"Mozilla/5.0"}).text[:300000]
            for pat in (r'<meta[^>]+(?:property|name)=["\'](?:og:description|description)["\'][^>]+content=["\']([^"\']+)',r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:property|name)=["\'](?:og:description|description)["\']'):
                m=re.search(pat,text,re.I)
                if m:return re.sub(r"\s+"," ",html.unescape(m.group(1))).strip()[:1000]
        except requests.RequestException: pass
    return ""

def upload_media(page_id,path,env):
    cfg=IZILIFE_ENVS[env]; token=os.getenv("IZILIFE_AGENT_TOKEN","").strip()
    with path.open("rb") as fh:
        r=requests.post(cfg["base_url"].rstrip("/")+"/scraper/agentAddPageMedia",data={"page_id":page_id},files={"media":(path.name,fh)},headers={"X-Agent-Token":token},verify=cfg.get("verify_ssl",True),timeout=180)
    try:return r.json()
    except Exception:return {"success":False,"error":f"HTTP {r.status_code}: {r.text[:400]}"}

def run(zone,env,file,dry_run,maximum):
    load_env(); env=normalize_env(env); path=Path(file) if file else default_file(zone,env)
    if not path.exists(): raise SystemExit(f"Fichier introuvable : {path}. Lance --init.")
    wb=openpyxl.load_workbook(path,data_only=False); total=created=skipped=errors=media_added=0
    for ws in wb.worksheets:
        headers={str(c.value or "").strip().upper():i for i,c in enumerate(ws[1])}
        for row in ws.iter_rows(min_row=2):
            def val(name):
                i=headers.get(name); return str(row[i].value or "").strip() if i is not None and i<len(row) else ""
            name=val("NOM")
            if not name or total>=maximum: continue
            category=val("CATEGORIE") or ws.title; links=[]
            for key,i in headers.items():
                if key.startswith("LIEN_"):
                    cell=row[i]; u=clean_url(cell.hyperlink.target if cell.hyperlink else cell.value)
                    if u and u not in links: links.append(u)
            description=val("DESCRIPTION") or description_from(links)
            payload={"name":name,"username":slugify(name),"category":category,"email":val("EMAIL"),"phone":val("TELEPHONE"),"description":description,"links":links}
            media_name=val("MEDIA_DIRECTORY_NAME"); media_dir=zone_dir(zone,env)/"medias"/media_name if media_name else None
            total+=1
            if dry_run:
                print("[DRY RUN]",json.dumps(payload,ensure_ascii=False),f"medias={media_dir}"); continue
            result=izilife_post("/scraper/agentAddPage",{"payload":json.dumps(payload,ensure_ascii=False)},env)
            if not result.get("success"): errors+=1; print(f"ERREUR {name}: {result.get('error',result)}"); continue
            page_id=result.get("page_id")
            if result.get("created"): created+=1; print(f"OK Page inactive {name} -> {result.get('username')}")
            else: skipped+=1; print(f"SKIP Page existante {name}")
            if media_dir and media_dir.is_dir():
                for media in sorted(p for p in media_dir.iterdir() if p.is_file() and p.suffix.lower() in MEDIA_EXT):
                    mr=upload_media(page_id,media,env)
                    if mr.get("success"):
                        if mr.get("created"): media_added+=1; print(f"  + média {media.name}")
                        else: print(f"  = média déjà présent {media.name}")
                    else: errors+=1; print(f"  ERREUR média {media.name}: {mr.get('error',mr)}")
    print(f"Résultat: pages={total} créées={created} existantes={skipped} médias ajoutés={media_added} erreurs={errors}")

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--zone",required=True); ap.add_argument("--env",default="prod",choices=["local","staging","prod"]); ap.add_argument("--file"); ap.add_argument("--init",action="store_true"); ap.add_argument("--dry-run",action="store_true"); ap.add_argument("--max",type=int,default=500); a=ap.parse_args()
    init_sheet(a.zone,a.env) if a.init else run(a.zone,a.env,a.file,a.dry_run,a.max)

if __name__=="__main__":
    from agent_excel_logger import run_logged
    run_logged("pages", "pages", main)
