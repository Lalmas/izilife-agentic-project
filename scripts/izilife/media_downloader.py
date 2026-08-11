#!/usr/bin/env python3
"""
Téléchargement en masse de médias Instagram via iGram.

Le script reprend la mécanique iGram de ingest_manual_curate.py :
- un seul navigateur ;
- un seul contexte et une seule page pour tout le fichier ;
- attente du bandeau de consentement différé ;
- clic sur Refuser ou Accepter ;
- navigation vers Photo, Story ou Reels ;
- téléchargement de tous les résultats, notamment les carrousels.

Fichier attendu :
    $AGENTIC_DRIVE_ROOT/agentic_workspace_local/izilife/medias/medias_igram.xlsx

Onglet Medias :
    URL | Type | Etat

Onglet Logs :
    URL | Type | Etat | Traité le | Nb fichiers | Fichiers | Message

Dossiers :
    medias/images/
    medias/reels/

Types acceptés :
    photo | carousel | story | reel

Usage :
    python media_downloader.py --env=local
    python media_downloader.py --env=local --max=10
    python media_downloader.py --env=local --dry-run

Prérequis :
    pip install playwright playwright-stealth openpyxl
    python -m playwright install chromium
"""

from __future__ import annotations

import argparse
import hashlib
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from playwright_stealth import stealth_sync
    HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False


# ─────────────────────────────────────────────
# ARCHITECTURE IZILIFE
# ─────────────────────────────────────────────

CORE_ROOT = Path(__file__).resolve().parents[2]
if str(CORE_ROOT) not in sys.path:
    sys.path.append(str(CORE_ROOT))

try:
    from core.paths import workspace_root
except Exception:
    def workspace_root(env_name: str = "prod") -> Path:
        folder = {
            "local": "agentic_workspace_local",
            "staging": "agentic_workspace_staging",
            "prod": "agentic_workspace",
        }.get(str(env_name).lower(), "agentic_workspace")
        return Path(os.environ["AGENTIC_DRIVE_ROOT"]).expanduser() / folder


IGRAM_HOME = "https://igram.world/fr/"
SHEET_QUEUE = "Medias"
SHEET_LOGS = "Logs"

VALID_TYPES = {"photo", "carousel", "story", "reel"}
PENDING_STATES = {"", "pending", "retry", "error", "failed"}

TYPE_ALIASES = {
    "photos": "photo",
    "image": "photo",
    "images": "photo",
    "post": "photo",
    "carrousel": "carousel",
    "caroussel": "carousel",
    "stories": "story",
    "reels": "reel",
}


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def sleep_random(mn: float = 1.0, mx: float = 2.0) -> None:
    time.sleep(random.uniform(mn, mx))


def apply_stealth(page) -> None:
    if HAS_STEALTH:
        stealth_sync(page)
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['fr-FR', 'fr']});
        window.chrome = { runtime: {} };
    """)


def media_root(env_name: str, root_override: str | None = None) -> Path:
    if root_override:
        return Path(root_override)
    return workspace_root(env_name) / "izilife" / "medias"


def workbook_file(env_name: str, root_override: str | None = None) -> Path:
    return media_root(env_name, root_override) / "medias_igram.xlsx"


def images_dir(env_name: str, root_override: str | None = None) -> Path:
    folder = media_root(env_name, root_override) / "images"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def reels_dir(env_name: str, root_override: str | None = None) -> Path:
    folder = media_root(env_name, root_override) / "reels"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# ─────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────

def normalized_headers(sheet) -> list[str]:
    return [
        str(cell.value or "").strip().lower()
        for cell in sheet[1]
    ]


def ensure_logs_sheet(workbook):
    if SHEET_LOGS in workbook.sheetnames:
        return workbook[SHEET_LOGS]

    sheet = workbook.create_sheet(SHEET_LOGS)
    sheet.append([
        "URL",
        "Type",
        "Etat",
        "Traité le",
        "Nb fichiers",
        "Fichiers",
        "Message",
    ])
    return sheet


def read_pending_rows(sheet, maximum: int) -> list[dict]:
    headers = normalized_headers(sheet)

    required = {"url", "type", "etat"}
    missing = required.difference(headers)
    if missing:
        raise RuntimeError(
            "Colonnes manquantes dans Medias : " + ", ".join(sorted(missing))
        )

    rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {
            headers[index]: (cell.value or "")
            for index, cell in enumerate(row)
            if index < len(headers)
        }

        url = str(values.get("url", "")).strip()
        state = str(values.get("etat", "")).strip().lower()

        if not url.startswith("http") or state not in PENDING_STATES:
            continue

        media_type = normalize_media_type(values.get("type"), url)
        rows.append({
            "row_index": row_index,
            "url": url,
            "type": media_type,
        })

        if maximum > 0 and len(rows) >= maximum:
            break

    return rows


def set_queue_state(sheet, row_index: int, state: str) -> None:
    headers = normalized_headers(sheet)
    state_column = headers.index("etat") + 1
    sheet.cell(row=row_index, column=state_column).value = state


def move_to_logs(workbook, queue_sheet, row: dict, files: list[Path]) -> None:
    logs = ensure_logs_sheet(workbook)
    logs.append([
        row["url"],
        row["type"],
        "done",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        len(files),
        "\n".join(path.name for path in files),
        f"{len(files)} fichier(s) téléchargé(s)",
    ])
    queue_sheet.delete_rows(row["row_index"], 1)


# ─────────────────────────────────────────────
# TYPES ET NOMS DE FICHIERS
# ─────────────────────────────────────────────

def normalize_media_type(raw_type, url: str) -> str:
    value = str(raw_type or "").strip().lower()
    value = TYPE_ALIASES.get(value, value)

    if value in VALID_TYPES:
        return value

    if "/stories/" in url or "/s/" in url or "story_media_id=" in url:
        return "story"
    if "/reel/" in url or "/reels/" in url:
        return "reel"
    return "photo"


def igram_tab(media_type: str) -> tuple[str, str]:
    if media_type in {"photo", "carousel"}:
        return "Photo", "https://igram.world/fr/photo"
    if media_type == "story":
        return "Story", "https://igram.world/fr/story"
    return "Reels", "https://igram.world/fr/reels"


def source_identifier(url: str) -> str:
    story_match = re.search(r"story_media_id=(\d+)", url)
    if story_match:
        return story_match.group(1)

    parts = [part for part in urlparse(url).path.split("/") if part]
    if parts:
        candidate = parts[-1]
        if candidate in {"p", "reel", "reels", "stories", "s"} and len(parts) > 1:
            candidate = parts[-2]
        return re.sub(r"[^A-Za-z0-9_-]", "_", candidate)[:80]

    return hashlib.sha1(url.encode("utf-8")).hexdigest()[:16]


def suffix_from_content_type(content_type: str, media_type: str) -> str:
    content_type = (content_type or "").lower()

    if "video/mp4" in content_type:
        return ".mp4"
    if "video/" in content_type:
        return ".mp4"
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    if "jpeg" in content_type or "jpg" in content_type:
        return ".jpg"

    return ".mp4" if media_type == "reel" else ".jpg"


def unique_path(folder: Path, stem: str, suffix: str) -> Path:
    path = folder / f"{stem}{suffix}"
    number = 1
    while path.exists():
        path = folder / f"{stem}_{number}{suffix}"
        number += 1
    return path


# ─────────────────────────────────────────────
# IGRAM — REPRISE DE INGEST_MANUAL_CURATE
# ─────────────────────────────────────────────

def handle_igram_consent(page, appearance_timeout: int = 15000) -> bool:
    """
    Attend le CMP différé de iGram, puis clique sur Refuser ou Accepter.
    Le même contexte Playwright est ensuite conservé pour tout le fichier.
    """
    consent_root = page.locator(".fc-consent-root").first

    try:
        log("  ⏳ Attente éventuelle du consentement iGram...")
        consent_root.wait_for(state="visible", timeout=appearance_timeout)
    except Exception:
        log("  ℹ️ Aucun consentement iGram affiché")
        return False

    log("  🍪 Fenêtre de consentement iGram détectée")
    buttons = consent_root.locator("button")
    deadline = time.time() + 12

    reject_words = (
        "refuser",
        "tout refuser",
        "refuser tout",
        "continuer sans accepter",
    )
    accept_words = (
        "accepter",
        "tout accepter",
        "accepter tout",
        "consentir",
    )

    # Priorité au refus, puis acceptation si le refus n'est pas proposé.
    for expected_words in (reject_words, accept_words):
        while time.time() < deadline:
            try:
                count = buttons.count()
            except Exception:
                count = 0

            for index in range(count):
                button = buttons.nth(index)
                try:
                    if not button.is_visible():
                        continue
                    label = " ".join(filter(None, [
                        button.inner_text(timeout=1000).strip(),
                        (button.get_attribute("aria-label") or "").strip(),
                        (button.get_attribute("title") or "").strip(),
                    ])).lower()
                except Exception:
                    continue

                if any(word in label for word in expected_words):
                    button.scroll_into_view_if_needed()
                    button.click(force=True, timeout=5000)
                    try:
                        consent_root.wait_for(state="hidden", timeout=10000)
                    except Exception:
                        page.wait_for_timeout(1000)

                    action = "refusé" if expected_words is reject_words else "accepté"
                    log(f"  🍪 Consentement iGram {action}")
                    return True

            page.wait_for_timeout(400)

    log("  ⚠️ Aucun bouton de consentement exploitable trouvé")
    return False


def open_igram_session(page) -> None:
    page.goto(
        IGRAM_HOME,
        wait_until="domcontentloaded",
        timeout=30000,
    )

    handle_igram_consent(page, appearance_timeout=15000)

    consent_root = page.locator(".fc-consent-root").first
    try:
        consent_root.wait_for(state="hidden", timeout=3000)
    except Exception:
        try:
            if consent_root.is_visible():
                raise RuntimeError(
                    "Le bandeau de consentement iGram bloque toujours la page"
                )
        except Exception:
            pass

    log("✅ Session iGram prête")


def open_type_tool(page, media_type: str) -> None:
    label, fallback_url = igram_tab(media_type)

    # Retour sur l'accueil dans la même session pour repartir d'un état propre.
    page.goto(
        IGRAM_HOME,
        wait_until="domcontentloaded",
        timeout=30000,
    )

    # Le CMP ne devrait pas revenir, mais on conserve le contrôle du script source.
    handle_igram_consent(page, appearance_timeout=1500)

    tool_link = page.locator(
        f"a:has-text('{label}'), "
        f"button:has-text('{label}')"
    ).first

    try:
        tool_link.wait_for(state="visible", timeout=10000)
        tool_link.click()
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        page.goto(
            fallback_url,
            wait_until="domcontentloaded",
            timeout=30000,
        )


def result_links(page):
    return page.locator(
        ".output-list a.button__download[href], "
        ".output-component a.button__download[href], "
        ".search-result a.button__download[href], "
        "a.button__download[href]"
    )


def close_igram_ad_popup(page) -> bool:
    """Ferme l'interstitiel publicitaire sans cliquer sur son bouton Ouvrir."""
    selectors = [
        "text=Fermer", "[aria-label='Fermer' i]", "[aria-label='Close' i]",
        "[title='Fermer' i]", "[title='Close' i]",
    ]
    for scope in [page, *page.frames]:
        for selector in selectors:
            try:
                close_button = scope.locator(selector).first
                if close_button.is_visible(timeout=250):
                    close_button.click(force=True, timeout=1500)
                    page.wait_for_timeout(500)
                    log("  ✅ Publicité interstitielle fermée")
                    return True
            except Exception:
                pass
    if "google_vignette" in page.url:
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(300)
        except Exception:
            pass
    return False


def direct_result_urls(page) -> list[str]:
    """Récupère les grands médias du résultat quand le bouton final est absent."""
    candidates = page.locator("img, video").evaluate_all("""
        elements => elements.map(el => ({
            src: el.currentSrc || el.src || '',
            width: el.naturalWidth || el.videoWidth || 0,
            height: el.naturalHeight || el.videoHeight || 0,
            context: `${el.alt || ''} ${el.className || ''} ${el.closest('a,article,section,div')?.className || ''}`.toLowerCase()
        })).filter(item =>
            /^https?:\/\//i.test(item.src)
            && item.width >= 300
            && item.height >= 300
            && !/(logo|avatar|icon|favicon|advert|publicit|banner|doubleclick|googleads)/i.test(`${item.src} ${item.context}`)
        ).sort((a, b) => (b.width * b.height) - (a.width * a.height))
    """)
    return list(dict.fromkeys(item["src"] for item in candidates if item.get("src")))


def download_one_url(
    page,
    instagram_url: str,
    media_type: str,
    image_output: Path,
    reel_output: Path,
) -> list[Path]:
    open_type_tool(page, media_type)

    input_box = page.locator(
        "form.search-form input[type='text'], "
        "input[type='text']"
    ).first
    input_box.wait_for(state="visible", timeout=15000)
    input_box.fill(instagram_url)

    sleep_random(0.7, 1.3)

    submit_button = page.locator(
        "form.search-form button.search-form__button[type='submit'], "
        "form.search-form button[type='submit']"
    ).first
    submit_button.wait_for(state="visible", timeout=15000)
    submit_button.click()

    log("  ⏳ URL envoyée, attente des résultats...")

    links = result_links(page)
    for _ in range(20):
        close_igram_ad_popup(page)
        try:
            if links.count() > 0 and links.first.is_visible(timeout=250):
                break
        except Exception:
            pass
        page.wait_for_timeout(1000)

    # Un carrousel expose plusieurs boutons. On attend brièvement que la liste
    # complète se stabilise avant de la parcourir.
    previous_count = -1
    stable_rounds = 0

    for _ in range(12):
        count = links.count()
        if count == previous_count and count > 0:
            stable_rounds += 1
        else:
            stable_rounds = 0
        if stable_rounds >= 2:
            break
        previous_count = count
        page.wait_for_timeout(500)

    hrefs = []
    for index in range(links.count()):
        href = links.nth(index).get_attribute("href")
        if href and href.startswith("http"):
            hrefs.append(href)
    if not hrefs:
        log("  ⚠️ Bouton final absent : détection directe des médias")
        close_igram_ad_popup(page)
        hrefs = direct_result_urls(page)
    hrefs = list(dict.fromkeys(hrefs))
    if not hrefs:
        raise RuntimeError("Aucun média de résultat iGram exploitable")
    log(f"  ✅ {len(hrefs)} média(s) trouvé(s)")

    folder = reel_output if media_type == "reel" else image_output
    identifier = source_identifier(instagram_url)
    saved: list[Path] = []
    hashes: set[str] = set()

    for index, href in enumerate(hrefs):
        response = page.request.get(
            href,
            headers={"Referer": page.url},
            timeout=90000,
        )

        if not response.ok:
            log(f"  ⚠️ Résultat #{index + 1} : HTTP {response.status}")
            continue

        body = response.body()
        if not body:
            continue

        digest = hashlib.sha256(body).hexdigest()
        if digest in hashes:
            log(f"  ℹ️ Résultat #{index + 1} ignoré : doublon")
            continue
        hashes.add(digest)

        content_type = response.headers.get("content-type") or ""
        suffix = suffix_from_content_type(content_type, media_type)
        stem = f"ig_{identifier}_{index}"
        local_path = unique_path(folder, stem, suffix)
        local_path.write_bytes(body)

        if local_path.stat().st_size < 5_000:
            log(
                f"  ⚠️ Fichier suspect : {local_path.name} "
                f"({local_path.stat().st_size} octets)"
            )
            local_path.unlink(missing_ok=True)
            continue

        saved.append(local_path)
        log(
            f"  📥 {local_path.name} "
            f"({local_path.stat().st_size // 1024} Ko)"
        )

    if not saved:
        raise RuntimeError("Aucun fichier valide n'a été enregistré")

    return saved


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Téléchargement en masse de médias Instagram via iGram"
    )
    parser.add_argument(
        "--env",
        choices=["local", "staging", "prod"],
        default="local",
    )
    parser.add_argument(
        "--max",
        type=int,
        default=0,
        help="Nombre maximum de lignes ; 0 = toutes",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--root",
        default=None,
        help="Dossier medias personnalisé",
    )
    args = parser.parse_args()

    xlsx = workbook_file(args.env, args.root)
    image_output = images_dir(args.env, args.root)
    reel_output = reels_dir(args.env, args.root)

    if not xlsx.exists():
        log(f"❌ Fichier introuvable : {xlsx}")
        sys.exit(1)

    workbook = openpyxl.load_workbook(xlsx)

    if SHEET_QUEUE not in workbook.sheetnames:
        log(f"❌ Onglet {SHEET_QUEUE} introuvable")
        sys.exit(1)

    queue = workbook[SHEET_QUEUE]
    ensure_logs_sheet(workbook)

    rows = read_pending_rows(queue, args.max)

    if not rows:
        log("✅ Aucune ligne à traiter")
        return

    log(f"Fichier : {xlsx}")
    log(f"À traiter : {len(rows)} ligne(s)")

    if args.dry_run:
        for row in rows:
            log(f"[DRY RUN] {row['type']} | {row['url']}")
        return

    stats = {"done": 0, "errors": 0, "files": 0}

    # Ordre décroissant pour pouvoir supprimer les lignes réussies sans
    # décaler les lignes qui restent à traiter.
    rows.sort(key=lambda item: item["row_index"], reverse=True)

    with sync_playwright() as playwright:
        launch_args = {
            "headless": False,
            "args": [
                "--disable-blink-features=AutomationControlled",
                "--disable-notifications",
            ],
        }

        # Même choix que ingest_manual_curate pour iGram.
        try:
            browser = playwright.chromium.launch(
                channel="chrome",
                **launch_args,
            )
        except Exception:
            browser = playwright.chromium.launch(**launch_args)

        context = browser.new_context(locale="fr-FR")
        page = context.new_page()
        apply_stealth(page)

        # Une seule session iGram pour tout le classeur.
        open_igram_session(page)

        for row in rows:
            log("")
            log(f"→ {row['type']} | {row['url']}")

            try:
                files = download_one_url(
                    page=page,
                    instagram_url=row["url"],
                    media_type=row["type"],
                    image_output=image_output,
                    reel_output=reel_output,
                )

                move_to_logs(
                    workbook,
                    queue,
                    row,
                    files,
                )
                workbook.save(xlsx)

                stats["done"] += 1
                stats["files"] += len(files)
                log(f"  ✅ done — {len(files)} fichier(s)")

            except Exception as exc:
                set_queue_state(
                    queue,
                    row["row_index"],
                    "error",
                )
                workbook.save(xlsx)

                stats["errors"] += 1
                log(f"  ❌ {type(exc).__name__} : {exc}")

            sleep_random(1.5, 3.0)

        browser.close()

    log("")
    log("=== Résultat ===")
    log(f"  Lignes done : {stats['done']}")
    log(f"  Fichiers    : {stats['files']}")
    log(f"  Erreurs     : {stats['errors']}")


if __name__ == "__main__":
    main()
