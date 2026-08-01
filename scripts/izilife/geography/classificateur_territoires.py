#!/usr/bin/env python3
"""Classe uniquement le littoral et la montagne, à partir de sources géographiques."""
from __future__ import annotations
import argparse, csv, io, json, os, time, zipfile
from pathlib import Path
import requests
from dotenv import load_dotenv

ROOT=Path(__file__).resolve().parents[1]
for p in (ROOT/'.env', ROOT.parent/'.env', ROOT.parent/'.env.izilife'):
    if p.exists(): load_dotenv(p, override=False)
URLS={'local':os.getenv('IZILIFE_LOCAL_URL','https://localhost:4443/izilife-admin'),'staging':os.getenv('IZILIFE_STAGING_URL','https://www.staging.izilife.co/izilife-admin'),'prod':os.getenv('IZILIFE_PROD_URL','https://www.izilife.co/izilife-admin')}
INSEE_DENSITY_URL='https://www.insee.fr/fr/statistiques/fichier/8571524/fichier_diffusion_2026.xlsx'

def api(env,path,payload=None):
    token=os.getenv('AGENT_SECRET_TOKEN','')
    if not token: raise RuntimeError('AGENT_SECRET_TOKEN manquant')
    r=requests.post(URLS[env].rstrip('/')+path, data=payload or {}, headers={'X-Agent-Token':token}, timeout=90, verify=env!='local')
    r.raise_for_status(); return r.json()

def geonames(country,cache):
    cache.mkdir(parents=True,exist_ok=True); z=cache/f'{country}.zip'
    if not z.exists():
        r=requests.get(f'https://download.geonames.org/export/dump/{country}.zip',timeout=180); r.raise_for_status(); z.write_bytes(r.content)
    with zipfile.ZipFile(z) as f:
        rows=[]
        with f.open(f'{country}.txt') as raw:
            for a in csv.reader(io.TextIOWrapper(raw,encoding='utf-8'),delimiter='\t'):
                if len(a)>16 and a[6]=='P': rows.append({'name':a[1],'ascii':a[2],'lat':float(a[4]),'lng':float(a[5]),'code':a[7],'population':int(a[14] or 0),'elevation':int(a[15] or a[16] or 0)})
        return rows

def dist2(a,b): return (float(a['lat'])-float(b['lat']))**2+(float(a['lng'])-float(b['lng']))**2
def match(city,rows):
    if city.get('latitude') is None or city.get('longitude') is None:return None
    near=sorted(rows,key=lambda r:dist2({'lat':city['latitude'],'lng':city['longitude']},r))[:8]
    name=str(city.get('name','')).casefold()
    named=[r for r in near if r['name'].casefold()==name or r['ascii'].casefold()==name]
    return (named or near)[0] if near else None

def coast(lat,lng,radius,session):
    q=f'[out:json][timeout:20];way(around:{radius*1000},{lat},{lng})[natural=coastline];out ids 1;'
    try:
        r=session.post('https://overpass-api.de/api/interpreter',data=q,timeout=35); r.raise_for_status()
        return bool(r.json().get('elements'))
    except requests.RequestException:return None

def insee_rural(cache):
    """Retourne {code_commune: 0|1} depuis la grille officielle INSEE 2026."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl requis pour lire la grille rurale INSEE (pip install openpyxl)') from exc
    cache.mkdir(parents=True,exist_ok=True); xlsx=cache/'insee_grille_densite_2026.xlsx'
    if not xlsx.exists():
        r=requests.get(INSEE_DENSITY_URL,timeout=180); r.raise_for_status(); xlsx.write_bytes(r.content)
    ws=load_workbook(xlsx,read_only=True,data_only=True)['Maille communale']
    return {str(row[0]).zfill(5):int(row[2]==3) for row in ws.iter_rows(min_row=6,values_only=True) if row[0] and row[2] in (1,2,3)}

def main():
    p=argparse.ArgumentParser(); p.add_argument('--country',required=True,help='ISO alpha-2, ex: FR'); p.add_argument('--env',choices=URLS,default='staging'); p.add_argument('--dry-run',action='store_true'); p.add_argument('--force',action='store_true'); p.add_argument('--limit',type=int,default=0); p.add_argument('--coast-km',type=float,default=10); p.add_argument('--mountain-elevation',type=int,default=500); p.add_argument('--pause',type=float,default=.3); a=p.parse_args()
    country=a.country.upper(); data=api(a.env,'/scraper/agentGetTerritories',{'country':country})
    cities=data.get('cities',[]); cities=cities[:a.limit or None]; cache=Path(__file__).resolve().parent/'.cache'; rows=geonames(country,cache/'geonames'); rural_by_code=insee_rural(cache) if country=='FR' else {}; session=requests.Session(); out=[]
    for i,c in enumerate(cities,1):
        g=match(c,rows)
        if not g: print(f'[{i}/{len(cities)}] {c["name"]}: non résolu'); continue
        coastal=coast(c['latitude'],c['longitude'],a.coast_km,session)
        confidence=.95 if g['name'].casefold()==str(c['name']).casefold() else .75
        city_code=str(c.get('city_code') or '').zfill(5); rural=rural_by_code.get(city_code)
        item={'scope_level':'CITY','scope_id':c['id'],'is_litoral':int(bool(coastal)) if coastal is not None else None,'is_mountain':int(g['elevation']>=a.mountain_elevation),'confidence':confidence,'source':'INSEE+GeoNames+OpenStreetMap','evidence':json.dumps({'geonames':g,'insee_city_code':city_code or None},ensure_ascii=False)}
        if rural is not None:item['is_rural']=rural
        out.append(item); print(f'[{i}/{len(cities)}] {c["name"]}: littoral={coastal} montagne={item["is_mountain"]} rural_INSEE={rural} village=inchangé')
        time.sleep(a.pause)
    if a.dry_run: print(json.dumps({'count':len(out),'sample':out[:5]},ensure_ascii=False,indent=2)); return
    for start in range(0,len(out),100): api(a.env,'/scraper/agentUpdateTerritoryGeography',{'items':json.dumps(out[start:start+100],ensure_ascii=False),'force':int(a.force)})
    print(f'{len(out)} ville(s) mises à jour. Les divisions administratives ont été propagées par le serveur.')
if __name__=='__main__': main()
