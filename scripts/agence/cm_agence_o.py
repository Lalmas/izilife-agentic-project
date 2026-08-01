"""
Agent Community Manager — Clients Agence
Usage :
  python cm_agence.py --client soultrain_lille
  python cm_agence.py --init   soultrain_lille
  python cm_agence.py --list
"""

import os, re, argparse, requests, anthropic, openpyxl, base64
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

# ══════════════════════════════════════════════════════
# CHEMINS FIXES
# ══════════════════════════════════════════════════════

DRIVE_ROOT   = Path("G:/Mon Drive/agentic_workspace/agence/clients")
CONTEXT_ROOT = Path("C:/Users/alcamara/Documents/agentic_Workspace/agence/context")

# Colonnes Excel — même structure qu'izilife
COL_DATE=1; COL_RESEAU=2; COL_TYPE=3; COL_VILLE=4; COL_LIEU=5
COL_THEME=6; COL_SERIE=7; COL_ARTISTE=8; COL_STYLE=9; COL_DATA=10
COL_DATE_EVENT=11; COL_HEURE=12; COL_INPUT_ID=13
COL_IMAGE_PROMPT=14; COL_IMAGE_PROVIDER=15
COL_POST_OUTPUT=16; COL_STATUT=17

STATUT_A_FAIRE="À faire"; STATUT_RELANCER="Relancer"; STATUT_GENERE="Généré"

ENV_GLOBAL  = Path("C:/Users/alcamara/Documents/agentic_Workspace/.env")
ENV_AGENCE  = Path("C:/Users/alcamara/Documents/agentic_Workspace/agence/.env.agence")

TYPE_CONTENT_DEFAULT = {
    "POST_STANDARD":"post","POST_PROMO":"post","POST_EVENT":"post",
    "POST_HUMOUR":"post","POST_HISTOIRE":"post","POST_PRODUIT":"post",
    "POST_TEMOIGNAGE":"post","EVENT_PHYSIQUE":"post","NOUVEAUTE":"post",
    "CARROUSEL":"carrousel","STORY":"story","REEL":"reel",
}

# ══════════════════════════════════════════════════════
# CONFIG CLIENT
# ══════════════════════════════════════════════════════

def get_client_config(slug:str) -> dict:
    client_drive = DRIVE_ROOT / slug
    return {
        "slug":      slug,
        "drive":     client_drive,
        "excel":     client_drive / f"planning_{slug}.xlsx",
        "outputs":   client_drive / "outputs",
        "inputs":    client_drive / "inputs",
        "templates": client_drive / "templates",

        "context":   CONTEXT_ROOT / f"community-manager-{slug}.md",
    }

def load_client_env(cfg:dict):
    """Charge les .env en cascade — JAMAIS de clés sur le Drive."""
    # 1. Clés API globales (local)
    if ENV_GLOBAL.exists():
        load_dotenv(ENV_GLOBAL, override=False)
    # 2. Config moteurs agence (local)
    if ENV_AGENCE.exists():
        load_dotenv(ENV_AGENCE, override=False)
    # 3. Override client si fichier local existe
    client_override = Path(f"C:/Users/alcamara/Documents/agentic_Workspace/agence/.env.{cfg['slug']}")
    if client_override.exists():
        load_dotenv(client_override, override=True)
        print(f"   .env override client : {client_override}")

def list_clients():
    print("\n👥 Clients :\n")
    if not DRIVE_ROOT.exists():
        print(f"  ❌ Drive introuvable : {DRIVE_ROOT}"); return
    for d in sorted(DRIVE_ROOT.iterdir()):
        if d.is_dir():
            cfg = get_client_config(d.name)
            print(f"  • {d.name}  "
                  f"{'✅' if cfg['excel'].exists() else '❌'} planning  "
                  f"{'✅' if cfg['env'].exists() else '⚠️ '} .env  "
                  f"{'✅' if cfg['context'].exists() else '❌'} contexte")

# ══════════════════════════════════════════════════════
# INIT CLIENT
# ══════════════════════════════════════════════════════

def init_client(slug:str):
    cfg = get_client_config(slug)

    # Dossiers Drive
    for d in ["outputs","inputs",
              "templates/top","templates/humour_local","templates/event_serie",
              "templates/event_animateur","templates/histoire_lieu","templates/pepite_niche"]:
        (cfg["drive"] / d).mkdir(parents=True, exist_ok=True)
    print(f"✅ Drive : {cfg['drive']}")
    print(f"   → Templates : {cfg['templates']}")

    print(f"ℹ️  Clés API  : {ENV_GLOBAL}")
    print(f"ℹ️  Config   : {ENV_AGENCE}")
    # Override client optionnel (local, pas sur Drive)
    client_override = Path(f"C:/Users/alcamara/Documents/agentic_Workspace/agence/.env.{slug}")
    print(f"ℹ️  Override : {client_override} (optionnel)")

    # Contexte client
    CONTEXT_ROOT.mkdir(parents=True, exist_ok=True)
    if not cfg["context"].exists():
        cfg["context"].write_text(f"""# Agent CM — {slug}

## Identité
- Nom :
- Secteur :
- Ville :
- Compte Instagram :
- Compte Facebook :

## Ton & style

## Thèmes éditoriaux

## Ce qu'on ne fait PAS

## Hashtags du client

## Exemples de posts passés

## Types de posts
POST_STANDARD | post | — | — | Post classique
POST_EVENT    | post | templates/event_serie/ | — | Événement
CARROUSEL     | carrousel | templates/top/ | — | Carrousel
STORY         | story | — | — | Story
""", encoding="utf-8")
        print(f"✅ Contexte : {cfg['context']}")
        print(f"   → Remplis ce fichier !")

    _create_client_excel(cfg)
    print(f"✅ Planning : {cfg['excel']}")
    print(f"\n🎉 Client '{slug}' prêt !")
    print(f"   1. Remplis : {cfg['context']}")
    print(f"   2. Remplis : {cfg['env']}")
    print(f"   3. Lance   : python cm_agence.py --client {slug}")

def _create_client_excel(cfg:dict):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning"
    ws.freeze_panes = "D3"

    thin = Side(style='thin', color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(r, c, v, bg="1A1A2E", ft="FFFFFF", sz=9):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(name="Arial", size=sz, bold=True, color=ft)
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = brd

    groups = [
        (1,3,"PLANIFICATION","1A1A2E"),(4,7,"CONTENU","0F3460"),
        (8,10,"DONNEES","533483"),(11,12,"EVENT","1B4332"),
        (13,13,"INPUT","784212"),(14,15,"IMAGE","0D6986"),
        (16,16,"OUTPUT","1A5276"),(17,17,"STATUT","6C3483"),
    ]
    for s,e,label,bg in groups:
        if s<e: ws.merge_cells(start_row=1,start_column=s,end_row=1,end_column=e)
        hdr(1,s,label,bg=bg,sz=10)

    headers = [
        ("DATE",12),("RESEAU",11),("TYPE",22),
        ("VILLE",12),("LIEU",16),("THEME / SUJET",22),("SERIE",13),
        ("ARTISTE / DJ / ASSO",26),("STYLE",14),("DATA",36),
        ("DATE EVENT",12),("HEURE / DUREE",11),
        ("INPUT_ID",16),
        ("IMAGE_PROMPT",36),("IMAGE_PROVIDER\ngpt / canva",14),
        ("POST OUTPUT",60),("STATUT",13),
    ]
    for i,(label,width) in enumerate(headers,1):
        hdr(2,i,label)
        ws.column_dimensions[get_column_letter(i)].width=width

    ws.row_dimensions[1].height=16; ws.row_dimensions[2].height=36

    all_types='","'.join(TYPE_CONTENT_DEFAULT.keys())
    dv_res  = DataValidation(type="list",formula1='"Instagram,Facebook,Les deux"',allow_blank=True)
    dv_type = DataValidation(type="list",formula1=f'"{all_types}"',allow_blank=True)
    dv_img  = DataValidation(type="list",formula1='"gpt,canva"',allow_blank=True)
    dv_stat = DataValidation(type="list",
        formula1='"À faire,Généré,Relancer,Validé,Publié,Skip"',allow_blank=True)

    for dv in [dv_res,dv_type,dv_img,dv_stat]: ws.add_data_validation(dv)
    for ri in range(3,103):
        ws.row_dimensions[ri].height=34
        for dv,col in [(dv_res,2),(dv_type,3),(dv_img,15),(dv_stat,17)]:
            dv.add(ws.cell(row=ri,column=col))

    cfg["excel"].parent.mkdir(parents=True,exist_ok=True)
    wb.save(cfg["excel"])

# ══════════════════════════════════════════════════════
# GÉNÉRATION
# ══════════════════════════════════════════════════════

def slugify(t:str) -> str:
    t = t.encode('ascii','ignore').decode()
    return re.sub(r"[^a-z0-9_]","",t.lower().replace(" ","_"))[:40]

def read_input(input_id:str, inputs_dir:Path) -> tuple[str, list[Path]]:
    if not input_id: return "",[]
    folder = inputs_dir / input_id
    if not folder.exists(): return "",[]
    desc=""
    f = folder/"description.txt"
    if f.exists(): desc=f.read_text(encoding="utf-8").strip()
    images=sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg",".jpeg",".png")])
    return desc,images

def get_content_type(post_type:str, context_text:str) -> str:
    """Cherche le CONTENT_TYPE dans le contexte client, sinon défaut."""
    for line in context_text.split("\n"):
        if post_type in line and "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 2: return parts[1]
    return TYPE_CONTENT_DEFAULT.get(post_type,"post")

def get_client_template(post_type:str, templates_dir:Path) -> Path | None:
    for sub in [post_type.lower(), "top", "event_serie"]:
        folder = templates_dir / sub
        if folder.exists():
            for ext in ["*.png","*.jpg"]:
                files = sorted(folder.glob(ext))
                if files: return files[0]
    return None

def build_prompt(row, content_type:str, input_desc:str, compte:str) -> str:
    def v(col): val=row[col-1].value; return str(val).strip() if val else ""
    desc = f"\nDESCRIPTION :\n{input_desc}" if input_desc else ""
    slides = "\n=== SLIDES ===\nSlide 1 : ...\nSlide 2 : ..." if content_type=="carrousel" else ""
    return f"""Tu es l'agent Community Manager de ce client. Compte : {compte}
CONTENT_TYPE: {content_type}

TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
THEME: {v(COL_THEME)} | ARTISTE: {v(COL_ARTISTE)}
DATA: {v(COL_DATA)} | DATE: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
RESEAU: {v(COL_RESEAU)}{desc}

Format :
=== TEXTE DU POST ===
[caption + hashtags]{slides}
"""

def call_llm(system:str, user:str) -> str:
    provider = os.getenv("LLM_PROVIDER","openai")
    if provider == "claude":
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        msg = client.messages.create(
            model=os.getenv("CLAUDE_MODEL","claude-sonnet-4-6"),
            max_tokens=1500, system=system,
            messages=[{"role":"user","content":user}])
        return msg.content[0].text
    else:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL","gpt-4o"),
            messages=[{"role":"system","content":system},{"role":"user","content":user}],
            max_tokens=1500)
        return r.choices[0].message.content

def generate_image_gpt(prompt:str, folder:Path, filename:str) -> Path | None:
    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        result = client.images.generate(
            model=os.getenv("IMAGE_MODEL","gpt-image-1"),
            prompt=f"{prompt}. Style photographique Instagram, format portrait.",
            n=1, size="1024x1536")
        img_b64 = result.data[0].b64_json
        if img_b64:
            out = folder/f"{filename}_gpt.png"
            out.write_bytes(base64.b64decode(img_b64))
            print(f"    🖼  GPT → {out.name}"); return out
        url = result.data[0].url
        if url:
            out = folder/f"{filename}_gpt.jpg"
            out.write_bytes(requests.get(url).content)
            print(f"    🖼  GPT → {out.name}"); return out
    except Exception as e:
        print(f"    ❌ GPT Image : {e}")
    return None

def compose_pillow(bg:Path, text:str, folder:Path, filename:str) -> Path | None:
    try:
        from PIL import Image, ImageDraw, ImageFont
        img = Image.open(bg).convert("RGBA")
        W,H = img.size
        overlay = Image.new("RGBA",(W,H),(0,0,0,130))
        img = Image.alpha_composite(img,overlay)
        draw = ImageDraw.Draw(img)
        try: font = ImageFont.truetype("arial.ttf",max(36,W//20))
        except: font = ImageFont.load_default()
        lines = [l for l in text.split("\n") if l.strip()][:8]
        y = H//4
        for line in lines:
            bb = draw.textbbox((0,0),line,font=font)
            x = (W-(bb[2]-bb[0]))//2
            draw.text((x+2,y+2),line,font=font,fill=(0,0,0,180))
            draw.text((x,y),line,font=font,fill=(255,255,255,255))
            y += bb[3]-bb[1]+14
        out = folder/f"{filename}_pillow.jpg"
        img.convert("RGB").save(out,quality=95)
        print(f"    🖼  Pillow → {out.name}"); return out
    except Exception as e:
        print(f"    ❌ Pillow : {e}")
    return None

def process_images(post_type,img_prompt,img_provider,input_images,
                   text,templates_dir,folder,filename):
    generated=[]
    engine=os.getenv("IMAGE_ENGINE","gpt")
    provider=img_provider or os.getenv("IMAGE_LLM","gpt")
    tpl=get_client_template(post_type,templates_dir)
    if tpl: print(f"    📁 Template : {tpl.name}")
    bg=input_images[0] if input_images else tpl
    if bg and engine in ("pillow","both"):
        short="\n".join([l for l in text.split("\n") if l.strip() and not l.startswith("===")][:6])
        out=compose_pillow(bg,short,folder,filename)
        if out: generated.append(out)
    if img_prompt and engine in ("gpt","both") and provider=="gpt":
        out=generate_image_gpt(img_prompt,folder,filename)
        if out: generated.append(out)
    return generated

def save_post(text,input_images,generated,content_type,folder):
    lines=[text.strip(),"",f"=== CONTENT_TYPE ===\n{content_type}"]
    if input_images: lines+=["","=== IMAGES SOURCE ==="]+[f"  {i.name}" for i in input_images]
    if generated:    lines+=["","=== IMAGES GENEREES ==="]+[f"  {i.name}" for i in generated]
    (folder/"post.txt").write_text("\n".join(lines),encoding="utf-8")
    print(f"    📄 post.txt")

# ══════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════

def run_client(slug:str):
    cfg = get_client_config(slug)
    load_client_env(cfg)

    print(f"\n🤖 Agent CM Agence — {slug} — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"   LLM    : {os.getenv('LLM_PROVIDER','?').upper()}")
    print(f"   Image  : {os.getenv('IMAGE_ENGINE','?').upper()}")
    print(f"   Excel  : {cfg['excel']}\n")

    if not cfg["context"].exists():
        print(f"❌ Contexte introuvable : {cfg['context']}")
        print(f"   Lance : python cm_agence.py --init {slug}"); return

    agence_cm = CONTEXT_ROOT / "community-manager-agence.md"
    system = ""
    if agence_cm.exists():
        system += agence_cm.read_text(encoding="utf-8") + "\n\n---\n\n"
    system += cfg["context"].read_text(encoding="utf-8")
    compte = f"@{slug}"
    for line in system.split("\n"):
        if "instagram" in line.lower() and "@" in line:
            compte = "@" + line.split("@")[-1].split()[0]; break

    if not cfg["excel"].exists():
        print(f"❌ Planning introuvable. Lance : python cm_agence.py --init {slug}"); return

    wb = openpyxl.load_workbook(cfg["excel"])
    ws = wb.active
    to_process = [r for r in ws.iter_rows(min_row=3)
                  if r[COL_STATUT-1].value in (STATUT_A_FAIRE,STATUT_RELANCER)]
    if not to_process: print("✅ Aucune ligne."); return
    print(f"📝 {len(to_process)} ligne(s)...\n")

    cfg["outputs"].mkdir(parents=True,exist_ok=True)
    cfg["inputs"].mkdir(parents=True,exist_ok=True)

    for i,row in enumerate(to_process,1):
        def v(col): val=row[col-1].value; return str(val).strip() if val else ""
        post_type=v(COL_TYPE); theme=v(COL_THEME); ville=v(COL_VILLE)
        input_id=v(COL_INPUT_ID); img_prompt=v(COL_IMAGE_PROMPT); img_provider=v(COL_IMAGE_PROVIDER)
        content_type=get_content_type(post_type,system)

        print(f"  [{i}/{len(to_process)}] {post_type} [{content_type}] — {ville} — {theme}")
        folder=cfg["outputs"]/f"{datetime.now().strftime('%Y%m%d')}_{slugify(post_type)}_{slugify(theme or ville)}"
        folder.mkdir(parents=True,exist_ok=True)

        input_desc,input_images=read_input(input_id,cfg["inputs"])
        if input_images: print(f"    📸 {len(input_images)} image(s)")
        if input_desc:   print(f"    📝 description.txt")

        try:
            text=call_llm(system,build_prompt(row,content_type,input_desc,compte))
            row[COL_POST_OUTPUT-1].value=text[:400]+"..." if len(text)>400 else text
            row[COL_STATUT-1].value=STATUT_GENERE
            print(f"    ✅ Texte ({os.getenv('LLM_PROVIDER','?').upper()})")
        except Exception as e:
            print(f"    ❌ {e}"); row[COL_STATUT-1].value="Erreur"; continue

        generated=process_images(post_type,img_prompt,img_provider,
                                  input_images,text,cfg["templates"],folder,
                                  slugify(f"{post_type}_{theme or ville}"))
        save_post(text,input_images,generated,content_type,folder)
        print(f"    📁 {folder.name}")

    wb.save(cfg["excel"])
    print(f"\n💾 Sauvegardé.\n✅ Terminé.\n")

def main():
    parser=argparse.ArgumentParser(description="Agent CM Agence")
    parser.add_argument("--client",type=str)
    parser.add_argument("--init",  type=str)
    parser.add_argument("--list",  action="store_true")
    args=parser.parse_args()
    if args.list:       list_clients()
    elif args.init:     init_client(args.init)
    elif args.client:   run_client(args.client)
    else:
        print("Usage:")
        print("  python cm_agence.py --client soultrain_lille")
        print("  python cm_agence.py --init   soultrain_lille")
        print("  python cm_agence.py --list")

if __name__=="__main__":
    main()