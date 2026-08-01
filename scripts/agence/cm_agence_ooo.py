"""
Agent Community Manager — Agence v3
Usage :
  python cm_agence.py --client soultrain_lille --init
  python cm_agence.py --client soultrain_lille
  python cm_agence.py --list

V3 :
- ajoute TEMPLATE_SOURCE : owned | inspiration | none
- ajoute TEMPLATE_REF : chemin / Canva ID / URL / note
- init crée/met à jour planning sans écraser
- gestion par client
- image GPT/Pillow branchée
"""

import os, sys, re, argparse, requests, anthropic, openpyxl, base64
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts")); return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent)); return
_ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, ENV_GLOBAL, ENV_AGENCE, agence_client
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False
    PROJECT_ROOT = Path(os.getenv("AGENTIC_WORKSPACE_ROOT", Path.home() / "Documents" / "agentic_Workspace"))
    ENV_GLOBAL = PROJECT_ROOT / ".env"; ENV_AGENCE = PROJECT_ROOT / "agence" / ".env.agence"
    def agence_client(slug: str, env_name: str = "prod") -> Path:
        return Path("G:/Mon Drive/agentic_workspace/agence/clients") / str(slug).strip()

DEFAULT_ENV = os.getenv("AGENCE_ENV", "prod")
CONTEXT_ROOT = PROJECT_ROOT / "agence" / "context"
TYPE_CONTENT_DEFAULT = {
    "POST_STANDARD":"post","POST_PROMO":"post","POST_EVENT":"post","POST_HUMOUR":"post","POST_HISTOIRE":"post",
    "POST_PRODUIT":"post","POST_TEMOIGNAGE":"post","EVENT_PHYSIQUE":"post","NOUVEAUTE":"post",
    "CARROUSEL":"carrousel","STORY":"story","REEL":"reel",
}
COL_DATE=1; COL_RESEAU=2; COL_TYPE=3; COL_VILLE=4; COL_LIEU=5
COL_THEME=6; COL_SERIE=7; COL_ARTISTE=8; COL_STYLE=9; COL_DATA=10
COL_DATE_EVENT=11; COL_HEURE=12; COL_INPUT_ID=13
COL_TEMPLATE_SOURCE=14; COL_TEMPLATE_REF=15; COL_IMAGE_PROMPT=16; COL_IMAGE_PROVIDER=17
COL_POST_OUTPUT=18; COL_IMAGE_OUTPUT=19; COL_IMAGE_STATUS=20; COL_STATUT=21
STATUT_A_FAIRE="À faire"; STATUT_RELANCER="Relancer"; STATUT_GENERE="Généré"


def slugify(t: str) -> str:
    t = str(t or "").encode("ascii","ignore").decode()
    return re.sub(r"[^a-z0-9_]", "", t.lower().replace(" ","_").replace("-","_"))[:45] or "post"


def get_client_config(slug: str, env_name: str = DEFAULT_ENV) -> dict:
    drive = agence_client(slug, env_name)
    return {"slug": slug, "drive": drive, "excel": drive / f"planning_{slug}.xlsx", "outputs": drive / "outputs",
            "inputs": drive / "inputs", "templates": drive / "templates",
            "context": CONTEXT_ROOT / f"community-manager-{slug}.md", "context_common": CONTEXT_ROOT / "community-manager-agence.md"}


def load_client_env(cfg: dict):
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_AGENCE.exists(): load_dotenv(ENV_AGENCE, override=False)
    override=PROJECT_ROOT/"agence"/f".env.{cfg['slug']}"
    if override.exists(): load_dotenv(override, override=True)


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin=Side(style='thin',color="CCCCCC"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    def hdr(r,c,v,bg="1A1A2E",ft="FFFFFF",sz=9):
        x=ws.cell(row=r,column=c,value=v); x.font=Font(name="Arial",size=sz,bold=True,color=ft)
        x.fill=PatternFill("solid",fgColor=bg); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=brd
    groups=[(1,3,"PLANIFICATION","1A1A2E"),(4,7,"CONTENU","0F3460"),(8,10,"DONNEES","533483"),(11,12,"EVENT","1B4332"),(13,13,"INPUT","784212"),(14,15,"TEMPLATE","8E44AD"),(16,17,"IMAGE","0D6986"),(18,20,"OUTPUT","1A5276"),(21,21,"STATUT","6C3483")]
    for s,e,label,bg in groups:
        if s<e: ws.merge_cells(start_row=1,start_column=s,end_row=1,end_column=e)
        hdr(1,s,label,bg=bg,sz=10)
    headers=[("DATE",12),("RESEAU",11),("TYPE",22),("VILLE",12),("LIEU",18),("THEME / SUJET",24),("SERIE",16),("ARTISTE / DJ / ASSO",26),("STYLE",16),("DATA",38),("DATE EVENT",12),("HEURE / DUREE",14),("INPUT_ID",16),("TEMPLATE_SOURCE\nowned/inspiration/none",22),("TEMPLATE_REF\nchemin / Canva ID / URL",32),("IMAGE_PROMPT",38),("IMAGE_PROVIDER\ngpt/canva/pillow",16),("POST OUTPUT",60),("IMAGE_OUTPUT",38),("IMAGE_STATUS",16),("STATUT",13)]
    for i,(label,width) in enumerate(headers,1): hdr(2,i,label); ws.column_dimensions[get_column_letter(i)].width=width
    ws.freeze_panes="D3"; ws.row_dimensions[2].height=42


def _add_validations(ws):
    from openpyxl.worksheet.datavalidation import DataValidation
    pairs=[
        (DataValidation(type="list",formula1='"Instagram,Facebook,Les deux"',allow_blank=True), COL_RESEAU),
        (DataValidation(type="list",formula1='"'+','.join(TYPE_CONTENT_DEFAULT.keys())+'"',allow_blank=True), COL_TYPE),
        (DataValidation(type="list",formula1='"owned,inspiration,none"',allow_blank=True), COL_TEMPLATE_SOURCE),
        (DataValidation(type="list",formula1='"gpt,canva,pillow"',allow_blank=True), COL_IMAGE_PROVIDER),
        (DataValidation(type="list",formula1='"non demandé,généré,erreur,source fournie,template utilisé"',allow_blank=True), COL_IMAGE_STATUS),
        (DataValidation(type="list",formula1='"À faire,Généré,Relancer,Validé,Publié,Skip"',allow_blank=True), COL_STATUT),
    ]
    for dv,col in pairs:
        ws.add_data_validation(dv)
        for r in range(3,503): dv.add(ws.cell(row=r,column=col))


def ensure_excel(path: Path):
    wb=openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    ws=wb["Planning"] if "Planning" in wb.sheetnames else wb.active; ws.title="Planning"
    _style_header(ws); _add_validations(ws)
    for r in range(3, max(ws.max_row,3)+1):
        if not ws.cell(r,COL_TEMPLATE_SOURCE).value: ws.cell(r,COL_TEMPLATE_SOURCE).value="none"
        if not ws.cell(r,COL_IMAGE_STATUS).value: ws.cell(r,COL_IMAGE_STATUS).value="non demandé"
    if "Post Types" in wb.sheetnames: del wb["Post Types"]
    wt=wb.create_sheet("Post Types"); wt.append(["POST_TYPE","CONTENT_TYPE","TEMPLATE_SOURCE","NOTES"])
    for pt,ct in TYPE_CONTENT_DEFAULT.items(): wt.append([pt,ct,"owned|inspiration|none","owned=asset client; inspiration=référence non copiée; none=libre"])
    path.parent.mkdir(parents=True,exist_ok=True); wb.save(path)


def init_client(slug: str, env_name: str = DEFAULT_ENV):
    cfg=get_client_config(slug, env_name)
    for d in ["outputs","inputs","templates"]: (cfg["drive"]/d).mkdir(parents=True,exist_ok=True)
    CONTEXT_ROOT.mkdir(parents=True,exist_ok=True)
    if not cfg["context_common"].exists():
        cfg["context_common"].write_text("""# Agent CM — Agence izilife

Règles communes à tous les clients.
- Ne jamais inventer des informations non fournies.
- Respecter le ton du client.
- 1 seul CTA par post.

## Règle TEMPLATE_SOURCE v3
- owned : asset/template du client ou de l'agence, utilisable comme base fixe.
- inspiration : référence externe, ancienne créa ou benchmark. S'en inspirer sans copier.
- none : aucun template, génération libre.
""", encoding="utf-8")
    if not cfg["context"].exists():
        cfg["context"].write_text(f"""# Agent CM — {slug}

## Identité
- Nom :
- Secteur :
- Ville :
- Compte Instagram :
- Compte Facebook :

## Ton & style

## Thèmes éditoriaux

## Ce qu'on ne fait PAS

## Hashtags du client

## Templates & inspirations
- TEMPLATE_SOURCE = owned : template client/agence utilisable comme base.
- TEMPLATE_SOURCE = inspiration : référence à ne pas copier.
- TEMPLATE_SOURCE = none : génération libre.

## Types de posts
POST_STANDARD | post | — | —
POST_EVENT    | post | templates/event/ | —
CARROUSEL     | carrousel | templates/carrousel/ | —
STORY         | story | templates/story/ | —
""", encoding="utf-8")
    ensure_excel(cfg["excel"])
    print(f"✅ Client prêt : {slug}\n   Planning : {cfg['excel']}\n   Contexte : {cfg['context']}")


def list_clients(env_name: str = DEFAULT_ENV):
    root=agence_client("", env_name).parent
    print("\n👥 Clients agence :\n")
    if not root.exists(): print(f"❌ Drive introuvable : {root}"); return
    for d in sorted(root.iterdir()):
        if d.is_dir(): print(f"  • {d.name}")


def load_context(cfg: dict) -> str:
    parts=[]
    if cfg["context_common"].exists(): parts.append(cfg["context_common"].read_text(encoding="utf-8"))
    if cfg["context"].exists(): parts.append(cfg["context"].read_text(encoding="utf-8"))
    return "\n\n---\n\n".join(parts)


def read_input(input_id: str, inputs_dir: Path):
    if not input_id: return "", []
    folder=inputs_dir/input_id
    if not folder.exists(): return "", []
    desc=(folder/"description.txt").read_text(encoding="utf-8").strip() if (folder/"description.txt").exists() else ""
    imgs=sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg",".jpeg",".png",".webp")])
    return desc,imgs


def get_content_type(post_type: str, context: str) -> str:
    for line in context.splitlines():
        if post_type in line and "|" in line:
            parts=[p.strip() for p in line.split("|")]
            if len(parts)>=2: return parts[1]
    return TYPE_CONTENT_DEFAULT.get(post_type,"post")


def build_prompt(row, content_type: str, input_desc: str, account: str) -> str:
    def v(c): val=row[c-1].value; return str(val).strip() if val else ""
    slides="\n=== SLIDES ===\nSlide 1 : ...\nSlide 2 : ..." if content_type=="carrousel" else ""
    return f"""Tu es l'agent CM du client. Compte: {account}
CONTENT_TYPE: {content_type}

TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
THEME: {v(COL_THEME)} | ARTISTE: {v(COL_ARTISTE)} | STYLE: {v(COL_STYLE)}
DATA: {v(COL_DATA)} | DATE EVENT: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
TEMPLATE_SOURCE: {v(COL_TEMPLATE_SOURCE) or 'none'}
TEMPLATE_REF: {v(COL_TEMPLATE_REF)}
IMAGE_PROMPT: {v(COL_IMAGE_PROMPT)}

Règle template : owned = base fixe utilisable ; inspiration = ne pas copier ; none = libre.

DESCRIPTION INPUT:
{input_desc}

Format :
=== TEXTE DU POST ===
[caption + hashtags]{slides}
"""


def call_llm(system: str, user: str) -> str:
    provider=os.getenv("LLM_PROVIDER","openai").lower()
    if provider=="claude":
        client=anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        msg=client.messages.create(model=os.getenv("CLAUDE_MODEL","claude-sonnet-4-6"),max_tokens=1500,system=system,messages=[{"role":"user","content":user}])
        return msg.content[0].text
    from openai import OpenAI
    client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    r=client.chat.completions.create(model=os.getenv("OPENAI_MODEL","gpt-4o"),messages=[{"role":"system","content":system},{"role":"user","content":user}],max_tokens=1500)
    return r.choices[0].message.content


def generate_image_gpt(prompt: str, folder: Path, filename: str) -> Path | None:
    if not prompt: return None
    try:
        from openai import OpenAI
        client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        result=client.images.generate(model=os.getenv("IMAGE_MODEL","gpt-image-1"),prompt=f"{prompt}. Format vertical Instagram, haute qualité.",n=1,size=os.getenv("IMAGE_SIZE","1024x1536"))
        b64=getattr(result.data[0],"b64_json",None)
        if b64:
            out=folder/f"{filename}_gpt.png"; out.write_bytes(base64.b64decode(b64)); return out
        url=getattr(result.data[0],"url",None)
        if url:
            out=folder/f"{filename}_gpt.jpg"; out.write_bytes(requests.get(url,timeout=60).content); return out
    except Exception as e: print(f"    ❌ GPT Image : {e}")
    return None


def compose_pillow(bg: Path, text: str, folder: Path, filename: str) -> Path | None:
    try:
        from PIL import Image, ImageDraw, ImageFont
        img=Image.open(bg).convert("RGBA"); W,H=img.size
        img=Image.alpha_composite(img, Image.new("RGBA",(W,H),(0,0,0,125)))
        draw=ImageDraw.Draw(img)
        try: font=ImageFont.truetype("arial.ttf",max(34,W//22))
        except Exception: font=ImageFont.load_default()
        y=H//4
        for line in [l.strip() for l in text.splitlines() if l.strip() and not l.startswith("===")][:7]:
            line=line[:55]; bb=draw.textbbox((0,0),line,font=font); x=max(30,(W-(bb[2]-bb[0]))//2)
            draw.text((x+2,y+2),line,font=font,fill=(0,0,0,180)); draw.text((x,y),line,font=font,fill=(255,255,255,255)); y+=(bb[3]-bb[1])+18
        out=folder/f"{filename}_pillow.jpg"; img.convert("RGB").save(out,quality=95); return out
    except Exception as e: print(f"    ❌ Pillow : {e}")
    return None


def resolve_template(src: str, ref: str, cfg: dict) -> Path | None:
    if src != "owned": return None
    candidates=[]
    if ref:
        p=Path(ref); candidates += [p, cfg["templates"]/ref, cfg["drive"]/ref]
    candidates.append(cfg["templates"])
    for c in candidates:
        if c.is_file(): return c
        if c.is_dir():
            files=sorted(list(c.glob("*.png"))+list(c.glob("*.jpg"))+list(c.glob("*.jpeg")))
            if files: return files[0]
    return None


def process_images(row,text,input_images,folder,filename,cfg):
    def v(c): val=row[c-1].value; return str(val).strip() if val else ""
    src=(v(COL_TEMPLATE_SOURCE) or "none").lower(); ref=v(COL_TEMPLATE_REF); prompt=v(COL_IMAGE_PROMPT)
    provider=(v(COL_IMAGE_PROVIDER) or os.getenv("IMAGE_LLM","gpt")).lower(); engine=os.getenv("IMAGE_ENGINE","gpt").lower(); generated=[]
    if src=="inspiration" and ref: prompt=f"{prompt}\nInspiration visuelle à respecter sans copier : {ref}".strip()
    tpl=resolve_template(src, ref, cfg); bg=input_images[0] if input_images else tpl
    if bg and engine in ("pillow","both"):
        out=compose_pillow(bg,text,folder,filename)
        if out: generated.append(out)
    if provider=="gpt" and engine in ("gpt","both") and prompt:
        out=generate_image_gpt(prompt,folder,filename)
        if out: generated.append(out)
    return generated


def save_post(text,input_images,generated,content_type,folder):
    lines=[text.strip(),"",f"=== CONTENT_TYPE ===\n{content_type}"]
    if input_images: lines += ["","=== IMAGES SOURCE ==="]+[f"  {p.name}" for p in input_images]
    if generated: lines += ["","=== IMAGES GENEREES ==="]+[f"  {p.name}" for p in generated]
    (folder/"post.txt").write_text("\n".join(lines),encoding="utf-8")


def run_client(slug: str, env_name: str = DEFAULT_ENV, dry_run: bool = False):
    cfg=get_client_config(slug, env_name); load_client_env(cfg)
    if not cfg["excel"].exists(): print(f"❌ Planning introuvable. Lance : python cm_agence.py --client {slug} --init"); return
    system=load_context(cfg); wb=openpyxl.load_workbook(cfg["excel"]); ws=wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    rows=[r for r in ws.iter_rows(min_row=3) if r[COL_STATUT-1].value in (STATUT_A_FAIRE,STATUT_RELANCER)]
    print(f"\n🤖 CM Agence v3 — {slug} — {len(rows)} ligne(s)" + (" [DRY RUN]" if dry_run else ""))
    cfg["outputs"].mkdir(parents=True,exist_ok=True); cfg["inputs"].mkdir(parents=True,exist_ok=True)
    account="@"+slug
    for line in system.splitlines():
        if "instagram" in line.lower() and "@" in line: account="@"+line.split("@")[-1].split()[0]; break
    for idx,row in enumerate(rows,1):
        def v(c): val=row[c-1].value; return str(val).strip() if val else ""
        post_type=v(COL_TYPE); theme=v(COL_THEME); ville=v(COL_VILLE); ct=get_content_type(post_type,system)
        print(f"  [{idx}/{len(rows)}] {post_type} — {ville} — {theme}")
        if dry_run: continue
        folder=cfg["outputs"]/f"{datetime.now().strftime('%Y%m%d')}_{slugify(post_type)}_{slugify(theme or ville)}"; folder.mkdir(parents=True,exist_ok=True)
        desc,imgs=read_input(v(COL_INPUT_ID),cfg["inputs"]); text=call_llm(system,build_prompt(row,ct,desc,account))
        gen=process_images(row,text,imgs,folder,slugify(theme or post_type),cfg); save_post(text,imgs,gen,ct,folder)
        row[COL_POST_OUTPUT-1].value=text[:32000]; row[COL_IMAGE_OUTPUT-1].value="\n".join(str(p) for p in gen)
        row[COL_IMAGE_STATUS-1].value="généré" if gen else ("source fournie" if imgs else "non demandé"); row[COL_STATUT-1].value=STATUT_GENERE
    if not dry_run: wb.save(cfg["excel"]); print(f"✅ Planning mis à jour : {cfg['excel']}")


def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--client","-c"); ap.add_argument("--env",default=DEFAULT_ENV,choices=["local","staging","prod"])
    ap.add_argument("--init",action="store_true"); ap.add_argument("--list",action="store_true"); ap.add_argument("--dry-run",action="store_true")
    args=ap.parse_args()
    if args.list: list_clients(args.env); return
    if not args.client: ap.error("--client requis sauf --list")
    if args.init: init_client(args.client,args.env); return
    run_client(args.client,args.env,args.dry_run)
if __name__=="__main__": main()
