"""
Agent Community Manager — agence v4 simplifié

Usage :
  python cm_agence.py --zone=lille --env=local --init
  python cm_agence.py --zone=lille --env=local --dry-run
  python cm_agence.py --zone=lille --env=local

V4 :
- Planning simplifié : plus de TEMPLATE_SOURCE, TEMPLATE_REF, IMAGE_PROVIDER par ligne.
- Ces réglages sont dans l'onglet "Post Types" une seule fois par type de post.
- IMAGE_PROVIDER est piloté par Post Types, fallback .env.
- Génération image branchée : GPT si provider=gpt, Pillow si provider=pillow, Canva stub si provider=canva.
"""

from __future__ import annotations
import os, sys, re, argparse, base64, requests, anthropic, openpyxl
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv


def _ensure_core_import_path():
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "scripts" / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent / "scripts")); return
        if (parent / "core" / "paths.py").exists():
            sys.path.insert(0, str(parent)); return
_ensure_core_import_path()

try:
    from core.paths import PROJECT_ROOT, ENV_GLOBAL, ENV_AGENCE, drive_workspace_root
    HAS_CORE_PATHS = True
except Exception:
    HAS_CORE_PATHS = False
    PROJECT_ROOT = Path(os.getenv("AGENTIC_WORKSPACE_ROOT", Path.home() / "Documents" / "agentic_Workspace"))
    ENV_GLOBAL = PROJECT_ROOT / ".env"
    ENV_AGENCE = PROJECT_ROOT / "agence" / ".env.agence"
    def normalize_zone(zone: str) -> str:
        z = str(zone or "").strip().lower()
        return z if z.endswith("-zone") else f"{z}-zone"

DEFAULT_SOCIAL_ENV = os.getenv("SOCIAL_ENV", "prod")
if HAS_CORE_PATHS:
    DRIVE_ROOT = drive_workspace_root(DEFAULT_SOCIAL_ENV) / "agence" / "clients"
else:
    drive_root_value = os.getenv("AGENTIC_DRIVE_ROOT", "").strip()
    if not drive_root_value:
        raise RuntimeError("AGENTIC_DRIVE_ROOT doit pointer vers la racine du Drive synchronise.")
    workspace_name = {
        "local": "agentic_workspace_local",
        "staging": "agentic_workspace_staging",
        "prod": "agentic_workspace",
    }.get(DEFAULT_SOCIAL_ENV, "agentic_workspace")
    DRIVE_ROOT = Path(drive_root_value).expanduser() / workspace_name / "agence" / "clients"
CONTEXT_ROOT = PROJECT_ROOT / "agence" / "context"
TEMPLATES_DIR = PROJECT_ROOT / "agence" / "templates"

TYPE_CONTENT = {
    "POST_STANDARD":"post", "POST_PROMO":"post", "POST_EVENT":"post", "POST_HUMOUR":"post",
    "POST_HISTOIRE":"post", "POST_PRODUIT":"post", "POST_TEMOIGNAGE":"post",
    "EVENT_PHYSIQUE":"post", "NOUVEAUTE":"post", "CARROUSEL":"carrousel", "STORY":"story",
    "REEL":"reel", "VIDEO":"video",
}
TYPE_TEMPLATE = {
    "POST_STANDARD":"templates/standard/",
    "POST_PROMO":"templates/promo/",
    "POST_EVENT":"templates/event/",
    "POST_HUMOUR":"templates/humour/",
    "POST_HISTOIRE":"templates/histoire/",
    "POST_PRODUIT":"templates/produit/",
    "POST_TEMOIGNAGE":"templates/temoignage/",
    "EVENT_PHYSIQUE":"templates/event/",
    "NOUVEAUTE":"templates/nouveaute/",
    "CARROUSEL":"templates/carrousel/",
    "STORY":"templates/story/",
}

# Planning simplifié
COL_DATE=1; COL_RESEAU=2; COL_TYPE=3; COL_VILLE=4; COL_LIEU=5; COL_SUJET=6
COL_DATA=7; COL_DATE_EVENT=8; COL_HEURE=9; COL_INPUT_ID=10; COL_IMAGE_PROMPT=11
COL_POST_OUTPUT=12; COL_IMAGE_OUTPUT=13; COL_IMAGE_STATUS=14; COL_STATUT=15
STATUT_A_FAIRE="À faire"; STATUT_RELANCER="Relancer"; STATUT_GENERE="Généré"

POST_TYPES_HEADERS = ["POST_TYPE","CONTENT_TYPE","TEMPLATE_LOCAL","TEMPLATE_SOURCE","TEMPLATE_REF","IMAGE_PROVIDER","IMAGE_MODE","NOTES"]


def zone_config(zone: str, env_name: str = DEFAULT_SOCIAL_ENV) -> dict:
    z = str(zone or "").strip().lower()
    drive = DRIVE_ROOT / z
    return {"zone": z, "drive": drive, "excel": drive / f"planning_{z}.xlsx", "outputs": drive / "outputs", "inputs": drive / "inputs", "context_zone": CONTEXT_ROOT / f"community-manager-{z}.md"}


def load_env():
    if ENV_GLOBAL.exists(): load_dotenv(ENV_GLOBAL, override=False)
    if ENV_AGENCE.exists(): load_dotenv(ENV_AGENCE, override=False)


def slugify(t: str) -> str:
    t = str(t or "").encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9_]", "", t.lower().replace(" ", "_").replace("-", "_"))[:45] or "post"


def ensure_planning(path: Path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    ws.title = "Planning"

    # Nettoie les anciennes fusions avant restyle
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    thin = Side(style='thin', color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hdr(r, c, v, bg="1A1A2E", sz=9):
        x = ws.cell(r, c)
        if isinstance(x, MergedCell):
            return
        x.value = v
        x.font = Font(name="Arial", size=sz, bold=True, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        x.border = brd

    groups=[(1,3,"PLANIFICATION","1A1A2E"),(4,7,"CONTENU","0F3460"),(8,9,"EVENT","1B4332"),(10,11,"INPUT / IMAGE","0D6986"),(12,14,"OUTPUT","1A5276"),(15,15,"STATUT","6C3483")]
    for s, e, label, bg in groups:
        if s < e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        hdr(1, s, label, bg=bg, sz=10)
        
    headers=[("DATE",12),("RÉSEAU",12),("TYPE",24),("VILLE",14),("LIEU",18),("SUJET",30),("DATA / INFOS",52),("DATE EVENT",14),("HEURE / DURÉE",16),("INPUT_ID",18),("IMAGE_PROMPT",48),("POST OUTPUT",70),("IMAGE_OUTPUT",40),("IMAGE_STATUS",18),("STATUT",14)]
    for i,(label,width) in enumerate(headers,1):
        hdr(2,i,label); ws.column_dimensions[get_column_letter(i)].width=width
    ws.freeze_panes="D3"; ws.row_dimensions[1].height=18; ws.row_dimensions[2].height=42

    validations=[
        (DataValidation(type="list", formula1='"Instagram,Facebook,Les deux"', allow_blank=True), COL_RESEAU),
        (DataValidation(type="list", formula1='"' + ','.join(TYPE_CONTENT.keys()) + '"', allow_blank=True), COL_TYPE),
        (DataValidation(type="list", formula1='"non demandé,généré,erreur,source fournie,template utilisé,canva à faire"', allow_blank=True), COL_IMAGE_STATUS),
        (DataValidation(type="list", formula1='"À faire,Généré,Relancer,Validé,Publié,Skip"', allow_blank=True), COL_STATUT),
    ]
    for dv,col in validations:
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}500")

    if "Post Types" in wb.sheetnames: del wb["Post Types"]
    wt = wb.create_sheet("Post Types")
    for i,h in enumerate(POST_TYPES_HEADERS,1):
        hdr(1,i,h,"1A1A2E")
        wt.column_dimensions[get_column_letter(i)].width=[28,14,30,18,36,18,18,58][i-1]
    default_provider = os.getenv("IMAGE_LLM", os.getenv("IMAGE_PROVIDER", "canva"))
    for pt,ct in TYPE_CONTENT.items():
        wt.append([pt, ct, TYPE_TEMPLATE.get(pt,""), "owned" if pt in TYPE_TEMPLATE else "none", "", default_provider, "auto", "owned=template interne ; inspiration=référence non copiée ; none=libre"])
    # Validations Post Types
    for col, values in [(4,"owned,inspiration,none"),(6,"canva,gpt,pillow,none"),(7,"auto,template,input,generate,none")]:
        dv=DataValidation(type="list", formula1='"'+values+'"', allow_blank=True); wt.add_data_validation(dv); dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}200")

    # Mode d'emploi
    if "Mode d'emploi" in wb.sheetnames: del wb["Mode d'emploi"]
    wi = wb.create_sheet("Mode d'emploi", 0)
    lines=[
        ["CM izilife — mode d'emploi rapide"],
        ["1. Remplir uniquement l'onglet Planning."],
        ["2. Les réglages visuels fixes sont dans Post Types, une seule fois par TYPE."],
        ["3. TEMPLATE_SOURCE dans Post Types : owned = template interne ; inspiration = référence non copiée ; none = génération libre."],
        ["4. TEMPLATE_REF dans Post Types : chemin local, Canva ID, URL ou note. Jamais à répéter ligne par ligne."],
        ["5. IMAGE_PROVIDER dans Post Types : canva, gpt, pillow, none. Fallback possible via .env."],
        ["6. INPUT_ID pointe vers inputs/[INPUT_ID]/description.txt et/ou des images."],
        ["7. IMAGE_PROMPT est optionnel. S'il est vide, le script tente description.txt/DATA/template."],
        ["8. Le script ne touche pas aux lignes Validé / Publié / Skip."],
        ["9. Pour relancer une ligne : mettre STATUT = Relancer."],
    ]
    for r,line in enumerate(lines,1):
        wi.cell(r,1,line[0]); wi.column_dimensions['A'].width=120
        if r==1: wi.cell(r,1).font=Font(name="Arial",size=14,bold=True); wi.cell(r,1).fill=PatternFill("solid",fgColor="1A1A2E"); wi.cell(r,1).font=Font(name="Arial",size=14,bold=True,color="FFFFFF")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"✅ Planning V4 prêt : {path}")


def load_post_type_config(wb, post_type: str) -> dict:
    if "Post Types" not in wb.sheetnames:
        return {}
    ws = wb["Post Types"]
    headers=[str(c.value or "").strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d={headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if str(d.get("POST_TYPE") or "").strip() == post_type:
            return d
    return {}


def init_zone(zone: str, env_name: str = DEFAULT_SOCIAL_ENV):
    cfg=zone_config(zone, env_name)
    for d in [cfg["outputs"], cfg["inputs"]]: d.mkdir(parents=True, exist_ok=True)
    cfg["context_zone"].parent.mkdir(parents=True, exist_ok=True)
    if not cfg["context_zone"].exists():
        cfg["context_zone"].write_text(f"""# Agent CM client — {cfg['zone']}
## Spécificités
- Ville principale :
- Hashtags client :
- Compte Instagram :
- Offres / événements :
- Ton / identité :
""", encoding="utf-8")
    ensure_planning(cfg["excel"])


def load_context(zone: str) -> str:
    parts=[]
    for fn in ["community-manager-agence.md"]:
        p=CONTEXT_ROOT/fn
        if p.exists(): parts.append(f"### {fn}\n{p.read_text(encoding='utf-8')}")
    zp=CONTEXT_ROOT / f"community-manager-{zone}.md"
    if zp.exists(): parts.append(f"### {zone}.md\n{zp.read_text(encoding='utf-8')}")
    return "\n\n---\n\n".join(parts)


def read_input(input_id: str, inputs_dir: Path):
    if not input_id: return "", []
    folder=inputs_dir/input_id
    if not folder.exists(): return "", []
    desc=(folder/"description.txt").read_text(encoding="utf-8").strip() if (folder/"description.txt").exists() else ""
    imgs=sorted([x for x in folder.iterdir() if x.suffix.lower() in (".jpg",".jpeg",".png",".webp")])
    return desc, imgs


def build_prompt(row, content_type: str, input_desc: str, cfg: dict) -> str:
    def v(c):
        val=row[c-1].value; return str(val).strip() if val else ""
    slides="\n=== SLIDES ===\nSlide 1 : ...\nSlide 2 : ..." if content_type=="carrousel" else ""
    return f"""Tu es l'agent Community Manager agence pour ce client. CONTENT_TYPE: {content_type}

CONFIG VISUELLE DU TYPE :
TEMPLATE_SOURCE: {cfg.get('TEMPLATE_SOURCE','none')}
TEMPLATE_REF: {cfg.get('TEMPLATE_REF','')}
IMAGE_PROVIDER: {cfg.get('IMAGE_PROVIDER','')}
IMAGE_MODE: {cfg.get('IMAGE_MODE','auto')}

DONNÉES :
TYPE: {v(COL_TYPE)} | VILLE: {v(COL_VILLE)} | LIEU: {v(COL_LIEU)}
SUJET: {v(COL_SUJET)}
DATA: {v(COL_DATA)}
DATE EVENT: {v(COL_DATE_EVENT)} | HEURE: {v(COL_HEURE)}
DESCRIPTION INPUT:
{input_desc}

Format OBLIGATOIRE :
=== TEXTE DU POST ===
[caption complète avec emojis + hashtags]{slides}
"""


def call_llm(system: str, user: str) -> str:
    provider=os.getenv("LLM_PROVIDER","claude").lower()
    if provider == "openai":
        from openai import OpenAI
        client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        r=client.chat.completions.create(model=os.getenv("OPENAI_MODEL","gpt-4o"), messages=[{"role":"system","content":system},{"role":"user","content":user}], max_tokens=1500)
        return r.choices[0].message.content
    client=anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    msg=client.messages.create(model=os.getenv("CLAUDE_MODEL","claude-sonnet-4-6"), max_tokens=1500, system=system, messages=[{"role":"user","content":user}])
    return msg.content[0].text


def find_template(template_local: str, template_ref: str):
    candidates=[]
    if template_ref and not str(template_ref).lower().startswith(("http://","https://","dag")):
        candidates.append(Path(template_ref))
        candidates.append(TEMPLATES_DIR / template_ref)
    if template_local:
        candidates.append(TEMPLATES_DIR / str(template_local).replace("templates/", ""))
    for c in candidates:
        if c.is_file(): return c
        if c.is_dir():
            files=[]
            for ext in ("*.png","*.jpg","*.jpeg","*.webp"):
                files += sorted(c.glob(ext))
            if files: return files[0]
    return None


def generate_image_gpt(prompt: str, folder: Path, filename: str):
    from openai import OpenAI
    client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    result=client.images.generate(model=os.getenv("IMAGE_MODEL","gpt-image-1"), prompt=f"{prompt}. Style Instagram, format portrait.", n=1, size=os.getenv("IMAGE_SIZE","1024x1536"))
    data=result.data[0]
    if getattr(data,"b64_json",None):
        out=folder/f"{filename}_gpt.png"; out.write_bytes(base64.b64decode(data.b64_json)); return out
    if getattr(data,"url",None):
        out=folder/f"{filename}_gpt.jpg"; out.write_bytes(requests.get(data.url, timeout=60).content); return out
    return None


def compose_pillow(bg: Path, text: str, folder: Path, filename: str):
    from PIL import Image, ImageDraw, ImageFont
    img=Image.open(bg).convert("RGBA"); W,H=img.size
    overlay=Image.new("RGBA",(W,H),(0,0,0,130)); img=Image.alpha_composite(img,overlay)
    draw=ImageDraw.Draw(img)
    try: font=ImageFont.truetype("arial.ttf", max(34, W//22))
    except Exception: font=ImageFont.load_default()
    lines=[l.strip() for l in text.split("\n") if l.strip() and not l.startswith("===")][:8]
    y=H//4
    for line in lines:
        bb=draw.textbbox((0,0), line, font=font); x=(W-(bb[2]-bb[0]))//2
        draw.text((x+2,y+2), line, font=font, fill=(0,0,0,180)); draw.text((x,y), line, font=font, fill=(255,255,255,255))
        y += (bb[3]-bb[1])+16
    out=folder/f"{filename}_pillow.jpg"; img.convert("RGB").save(out, quality=95); return out


def process_images(post_type: str, cfg: dict, img_prompt: str, input_images: list[Path], text: str, folder: Path, filename: str):
    provider=str(cfg.get("IMAGE_PROVIDER") or os.getenv("IMAGE_LLM") or os.getenv("IMAGE_PROVIDER") or "canva").lower()
    mode=str(cfg.get("IMAGE_MODE") or "auto").lower()
    tpl=find_template(str(cfg.get("TEMPLATE_LOCAL") or ""), str(cfg.get("TEMPLATE_REF") or ""))
    bg=input_images[0] if input_images else tpl
    generated=[]
    if provider == "none" or mode == "none": return generated, "non demandé"
    if provider == "canva":
        # API Canva non branchée ici : on laisse une trace actionnable.
        return generated, "canva à faire"
    if provider in ("pillow", "both") and bg:
        try: generated.append(compose_pillow(bg, text, folder, filename))
        except Exception as e: print(f"    ❌ Pillow : {e}")
    if provider in ("gpt", "both"):
        prompt=img_prompt or "\n".join([str(cfg.get("TEMPLATE_REF") or ""), text[:500]])
        if prompt.strip():
            try: generated.append(generate_image_gpt(prompt, folder, filename))
            except Exception as e: print(f"    ❌ GPT Image : {e}")
    if generated: return generated, "généré"
    if bg: return [], "template utilisé"
    return [], "non demandé"


def save_post(text: str, generated: list[Path], folder: Path, content_type: str):
    lines=[text.strip(), "", f"=== CONTENT_TYPE ===\n{content_type}"]
    if generated: lines += ["", "=== IMAGES GENEREES ==="] + [f"  {p.name}" for p in generated if p]
    (folder/"post.txt").write_text("\n".join(lines), encoding="utf-8")


def run_zone(zone: str, env_name: str = DEFAULT_SOCIAL_ENV, dry_run=False):
    load_env(); cfg=zone_config(zone, env_name); zone=cfg["zone"]
    if not cfg["excel"].exists():
        print(f"❌ Planning introuvable. Lance --init : {cfg['excel']}"); return
    system=load_context(zone)
    wb=openpyxl.load_workbook(cfg["excel"]); ws=wb["Planning"] if "Planning" in wb.sheetnames else wb.active
    rows=[r for r in ws.iter_rows(min_row=3) if str(r[COL_STATUT-1].value or "").strip() in (STATUT_A_FAIRE, STATUT_RELANCER)]
    print(f"🤖 CM agence V4 — client:{zone} — {len(rows)} ligne(s)" + (" [DRY RUN]" if dry_run else ""))
    cfg["outputs"].mkdir(parents=True,exist_ok=True); cfg["inputs"].mkdir(parents=True,exist_ok=True)
    for idx,row in enumerate(rows,1):
        def v(c): val=row[c-1].value; return str(val).strip() if val else ""
        post_type=v(COL_TYPE); subject=v(COL_SUJET); city=v(COL_VILLE)
        pt_cfg=load_post_type_config(wb, post_type)
        content_type=str(pt_cfg.get("CONTENT_TYPE") or TYPE_CONTENT.get(post_type,"post"))
        print(f"  [{idx}/{len(rows)}] {post_type} [{content_type}] — {city} — {subject}")
        input_desc, input_images=read_input(v(COL_INPUT_ID), cfg["inputs"])
        if dry_run:
            print(f"    CONFIG: provider={pt_cfg.get('IMAGE_PROVIDER')} template_source={pt_cfg.get('TEMPLATE_SOURCE')} template_ref={pt_cfg.get('TEMPLATE_REF')}")
            continue
        folder=cfg["outputs"]/f"{datetime.now().strftime('%Y%m%d')}_{slugify(post_type)}_{slugify(subject or city)}"
        folder.mkdir(parents=True,exist_ok=True)
        try:
            text=call_llm(system, build_prompt(row, content_type, input_desc, pt_cfg))
            generated, image_status=process_images(post_type, pt_cfg, v(COL_IMAGE_PROMPT), input_images, text, folder, slugify(post_type))
            save_post(text, generated, folder, content_type)
            row[COL_POST_OUTPUT-1].value=text[:30000]
            row[COL_IMAGE_OUTPUT-1].value="; ".join([str(p) for p in generated if p])
            row[COL_IMAGE_STATUS-1].value=image_status
            row[COL_STATUT-1].value=STATUT_GENERE
            print(f"    ✅ généré → {folder}")
        except Exception as e:
            row[COL_STATUT-1].value="Erreur"; row[COL_IMAGE_STATUS-1].value="erreur"; row[COL_POST_OUTPUT-1].value=str(e)[:1000]
            print(f"    ❌ {e}")
    wb.save(cfg["excel"])


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--zone", required=False)
    ap.add_argument("--env", default=DEFAULT_SOCIAL_ENV, choices=["local","staging","prod"])
    ap.add_argument("--init", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args=ap.parse_args()
    if not args.zone: raise SystemExit("--zone requis")
    if args.init: init_zone(args.zone, args.env)
    else: run_zone(args.zone, args.env, args.dry_run)

if __name__ == "__main__": main()
